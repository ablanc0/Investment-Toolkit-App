"""
InvToolkit — Excel budget & net-worth importer.
Parses '2026 Presupuesto + Patrimonio Neto.xlsx' into JSON for portfolio.json.
"""

import uuid

import openpyxl

# ── Constants ─────────────────────────────────────────────────────────

MONTH_SHEETS = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
MONTH_KEYS = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]

# Presupuesto: 6 category columns (name_col, amount_col)
CATEGORY_COLS = [
    ("income",        "Income",                   "income",      "A", "C"),
    ("essential",     "Essential Expenses",       "expense",     "E", "G"),
    ("discretionary", "Discretionary Expenses",   "expense",     "I", "K"),
    ("debt",          "Debt Payments",            "expense",     "M", "O"),
    ("savings",       "Savings",                  "savings",     "Q", "S"),
    ("investments",   "Investments",              "investments", "U", "W"),
]

# Monthly sheets: same 6 columns, each with (name, estimated, actual)
MONTH_CAT_COLS = [
    ("income",        "A", "B", "C"),
    ("essential",     "E", "F", "G"),
    ("discretionary", "I", "J", "K"),
    ("debt",          "M", "N", "O"),
    ("savings",       "Q", "R", "S"),
    ("investments",   "U", "V", "W"),
]

# Transaction log columns in monthly sheets (rows 59-158)
TXN_LOG_COLS = [
    ("income",        "A", "B", "C"),      # subcategory, date, amount
    ("essential",     "E", "F", "G"),
    ("discretionary", "I", "J", "K"),
    ("debt",          "M", "N", "O"),
    ("savings",       "Q", "R", "S"),
    ("investments",   "U", "V", "W"),
]
TXN_START_ROW = 59
TXN_END_ROW = 158

# Patrimonio Neto: asset/liability sections
NW_MONTH_COLS = {
    "start": "C", "january": "E", "february": "G", "march": "I",
    "april": "K", "may": "M", "june": "O", "july": "Q",
    "august": "S", "september": "U", "october": "W", "november": "Y",
    "december": "AA",
}

ASSET_SECTIONS = [
    {"group": "bankAccounts", "label_row": 42, "data_start": 43, "data_end": 52, "total_row": 53},
    {"group": "investments",  "label_row": 54, "data_start": 55, "data_end": 69, "total_row": 70},
    {"group": "property",     "label_row": 71, "data_start": 72, "data_end": 81, "total_row": 82},
    {"group": "otherAssets",  "label_row": 83, "data_start": 84, "data_end": 98, "total_row": 99},
]

LIABILITY_SECTIONS = [
    {"group": "debt",            "label_row": 107, "data_start": 108, "data_end": 127, "total_row": 128},
    {"group": "otherLiabilities", "label_row": 129, "data_start": 130, "data_end": 139, "total_row": 140},
]


def _cell(ws, col, row):
    """Read a cell value, return None-safe float or string."""
    v = ws[f"{col}{row}"].value
    if v is None:
        return None
    if isinstance(v, str) and v.strip() in ("", "$"):
        return None
    return v


def _num(v):
    """Coerce to float, default 0."""
    if v is None:
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


# ── Presupuesto (master budget) ──────────────────────────────────────

def _parse_categories(ws):
    """Extract 6 categories with subcategories and budgeted amounts."""
    categories = []
    for cat_id, name, cat_type, name_col, amt_col in CATEGORY_COLS:
        subcategories = []
        for row in range(10, 35):
            sub_name = _cell(ws, name_col, row)
            if sub_name and isinstance(sub_name, str) and sub_name.strip():
                amt = _num(_cell(ws, amt_col, row))
                subcategories.append({"name": sub_name.strip(), "budgeted": amt})
        categories.append({
            "id": cat_id,
            "name": name,
            "type": cat_type,
            "subcategories": subcategories,
        })
    return categories


def _parse_goals(ws):
    """Read goals from M1:M3."""
    goals = []
    for row in range(1, 4):
        v = _cell(ws, "M", row)
        if v and isinstance(v, str) and v.strip():
            goals.append(v.strip())
    return goals


# ── Monthly sheets ───────────────────────────────────────────────────

def _parse_transactions(ws, cat_id, sub_col, date_col, amt_col):
    """Extract individual transactions from the transaction log area."""
    txns = []
    for row in range(TXN_START_ROW, TXN_END_ROW + 1):
        sub_name = _cell(ws, sub_col, row)
        if not sub_name or not isinstance(sub_name, str) or not sub_name.strip():
            continue
        amount = _num(_cell(ws, amt_col, row))
        if amount == 0:
            continue

        date_val = _cell(ws, date_col, row)
        date_str = ""
        if date_val is not None:
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, str):
                date_str = date_val.strip()

        txns.append({
            "id": uuid.uuid4().hex[:8],
            "subcategory": sub_name.strip(),
            "date": date_str,
            "amount": amount,
            "notes": "",
        })
    return txns


def _synthetic_transactions(actuals):
    """Create synthetic transactions from aggregated actuals (fallback)."""
    transactions = {}
    for cat_id, cat_actuals in actuals.items():
        cat_txns = []
        for sub_name, amount in cat_actuals.items():
            if amount == 0:
                continue
            cat_txns.append({
                "id": uuid.uuid4().hex[:8],
                "subcategory": sub_name,
                "date": "",
                "amount": float(amount),
                "notes": "Imported from actuals",
            })
        if cat_txns:
            transactions[cat_id] = cat_txns
    return transactions


def _parse_month(ws, categories):
    """Parse a monthly sheet: actuals, overrides, and transactions."""
    actuals = {}
    overrides = {}

    # Build master budget lookup from categories
    master = {}
    for cat in categories:
        master[cat["id"]] = {s["name"]: s["budgeted"] for s in cat["subcategories"]}

    for cat_id, name_col, est_col, act_col in MONTH_CAT_COLS:
        cat_actuals = {}
        cat_overrides = {}
        for row in range(25, 50):
            sub_name = _cell(ws, name_col, row)
            if not sub_name or not isinstance(sub_name, str) or not sub_name.strip():
                continue
            sub_name = sub_name.strip()
            actual = _num(_cell(ws, act_col, row))
            estimated = _num(_cell(ws, est_col, row))

            cat_actuals[sub_name] = actual

            # Detect override: estimated differs from master budget
            master_amt = master.get(cat_id, {}).get(sub_name)
            if master_amt is not None and estimated != master_amt:
                cat_overrides[sub_name] = estimated

        if cat_actuals:
            actuals[cat_id] = cat_actuals
        if cat_overrides:
            overrides[cat_id] = cat_overrides

    # Try to extract transaction-level data from the log area
    transactions = {}
    has_txn_data = False
    for cat_id, sub_col, date_col, amt_col in TXN_LOG_COLS:
        txns = _parse_transactions(ws, cat_id, sub_col, date_col, amt_col)
        if txns:
            transactions[cat_id] = txns
            has_txn_data = True

    # Fall back to synthetic transactions from actuals if no log data
    if not has_txn_data:
        transactions = _synthetic_transactions(actuals)

    rollover = _cell(ws, "S", 2)
    rollover = bool(rollover) if rollover is not None else False

    month_data = {"actuals": actuals, "rollover": rollover}
    if overrides:
        month_data["overrides"] = overrides
    if transactions:
        month_data["transactions"] = transactions
    return month_data


# ── Patrimonio Neto ──────────────────────────────────────────────────

def _parse_nw_section(ws, sections, month_col):
    """Read asset or liability values for a given month column."""
    items = {}
    for sec in sections:
        for row in range(sec["data_start"], sec["data_end"] + 1):
            name = _cell(ws, "A", row)
            if not name or not isinstance(name, str) or not name.strip():
                continue
            val = _num(_cell(ws, month_col, row))
            items[name.strip()] = val
    return items


def _parse_net_worth(ws):
    """Extract asset/liability definitions and monthly snapshots."""
    # Collect asset items with groups
    assets = []
    for sec in ASSET_SECTIONS:
        for row in range(sec["data_start"], sec["data_end"] + 1):
            name = _cell(ws, "A", row)
            if name and isinstance(name, str) and name.strip():
                assets.append({"name": name.strip(), "group": sec["group"]})

    # Collect liability items with groups
    liabilities = []
    for sec in LIABILITY_SECTIONS:
        for row in range(sec["data_start"], sec["data_end"] + 1):
            name = _cell(ws, "A", row)
            if name and isinstance(name, str) and name.strip():
                liabilities.append({"name": name.strip(), "group": sec["group"]})

    # Build snapshots
    snapshots = []
    for month_key, col in NW_MONTH_COLS.items():
        asset_vals = _parse_nw_section(ws, ASSET_SECTIONS, col)
        liability_vals = _parse_nw_section(ws, LIABILITY_SECTIONS, col)
        # Skip months with no data
        if not any(v != 0 for v in asset_vals.values()) and not any(v != 0 for v in liability_vals.values()):
            continue
        snap = {"month": month_key, "year": 2026, "assets": asset_vals, "liabilities": liability_vals}
        snapshots.append(snap)

    return {"assets": assets, "liabilities": liabilities, "snapshots": snapshots}


# ── Public API ───────────────────────────────────────────────────────

def import_budget(filepath):
    """Parse Excel file and return {budget: {...}, netWorth: {...}} dict."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 1. Parse master budget from Presupuesto
    ws_pres = wb["Presupuesto"]
    categories = _parse_categories(ws_pres)
    goals = _parse_goals(ws_pres)

    # 2. Parse monthly actuals and transactions
    months = {}
    for sheet_name, month_key in zip(MONTH_SHEETS, MONTH_KEYS):
        if sheet_name in wb.sheetnames:
            ws_month = wb[sheet_name]
            month_data = _parse_month(ws_month, categories)
            # Only include months that have actual data
            has_data = any(
                v != 0 for cat_acts in month_data["actuals"].values() for v in cat_acts.values()
            )
            if has_data:
                months[month_key] = month_data

    budget = {
        "year": 2026,
        "currency": "$",
        "goals": goals,
        "categories": categories,
        "months": months,
    }

    # 3. Parse net worth
    ws_nw = wb["Patrimonio Neto"]
    net_worth = _parse_net_worth(ws_nw)

    return {"budget": budget, "netWorth": net_worth}
