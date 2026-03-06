"""
InvToolkit — Historical Monte Carlo / Rule 4% simulation engine.
Runs retirement withdrawal simulations across all historical S&P 500 periods.
"""

import os

from config import DATA_DIR
from services.data_store import load_portfolio, save_portfolio


def load_historic_data():
    """Load S&P 500 historic data from portfolio or import from Excel."""
    portfolio = load_portfolio()
    historic = portfolio.get("historicData")
    if historic and len(historic) > 50:
        return historic

    # Try importing from Excel
    xlsx_path = os.path.join(DATA_DIR, "Investments Toolkit-v1.0.xlsx")
    if not os.path.exists(xlsx_path):
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["Historic_Data"]
        data = []
        for r in range(2, ws.max_row + 1):
            year = ws.cell(r, 1).value
            if year is None or not isinstance(year, (int, float)) or year < 1900:
                continue
            data.append({
                "year": int(year),
                "avgClosing": ws.cell(r, 2).value or 0,
                "yearOpen": ws.cell(r, 3).value or 0,
                "yearHigh": ws.cell(r, 4).value or 0,
                "yearLow": ws.cell(r, 5).value or 0,
                "yearClose": ws.cell(r, 6).value or 0,
                "annualReturn": ws.cell(r, 7).value or 0,
                "cpi": ws.cell(r, 8).value or 0,
            })
        wb.close()
        # Cache in portfolio
        portfolio["historicData"] = data
        save_portfolio(portfolio)
        return data
    except Exception as e:
        print(f"Error importing historic data: {e}")
        return []


def _run_simulation(returns_by_year, cpi_by_year, all_years, max_year,
                     starting_balance, withdrawal_rate, horizon,
                     strategy="fixed", guardrail_floor=None, guardrail_ceiling=None,
                     cash_buffer_years=0, div_yield=0, div_growth=0):
    """Core simulation engine supporting multiple strategies."""
    scenarios = []
    success_count = 0
    total_count = 0

    for start_year in all_years:
        end_year = start_year + horizon - 1
        if end_year > max_year:
            break

        total_count += 1
        balance = starting_balance
        base_withdrawal = starting_balance * withdrawal_rate
        annual_withdrawal = base_withdrawal
        cash_reserve = base_withdrawal * cash_buffer_years
        yearly_data = []
        survived = True
        cumulative_inflation = 1.0

        for yr_offset in range(horizon):
            yr = start_year + yr_offset
            ret = returns_by_year.get(yr, 0)
            cpi = cpi_by_year.get(yr, 0.03)
            cumulative_inflation *= (1 + cpi)

            if strategy == "dividend":
                # Dividend strategy: yield on current balance, no selling
                div_income = balance * div_yield
                balance = balance * (1 + ret)
                actual_withdrawal = div_income
            elif strategy == "combined":
                # Combined: dividend income + sell remainder
                div_income = balance * div_yield
                balance = balance * (1 + ret)
                sell_amount = max(0, annual_withdrawal - div_income)
                balance -= sell_amount
                actual_withdrawal = div_income + sell_amount
            elif strategy == "guardrails":
                # Guardrails: adjust withdrawal based on portfolio performance
                balance = balance * (1 + ret)
                floor_amount = base_withdrawal * cumulative_inflation * (guardrail_floor or 0.8)
                ceiling_amount = base_withdrawal * cumulative_inflation * (guardrail_ceiling or 1.2)
                # Target: withdrawal_rate of current balance
                target = balance * withdrawal_rate
                annual_withdrawal = max(floor_amount, min(ceiling_amount, target))
                # Use cash buffer during down years
                if ret < -0.1 and cash_reserve > 0:
                    from_cash = min(cash_reserve, annual_withdrawal)
                    cash_reserve -= from_cash
                    balance -= (annual_withdrawal - from_cash)
                else:
                    balance -= annual_withdrawal
                actual_withdrawal = annual_withdrawal
            else:
                # Fixed (classic Rule 4%)
                balance = balance * (1 + ret)
                # Use cash buffer during down years
                if cash_buffer_years > 0 and ret < -0.1 and cash_reserve > 0:
                    from_cash = min(cash_reserve, annual_withdrawal)
                    cash_reserve -= from_cash
                    balance -= (annual_withdrawal - from_cash)
                else:
                    balance -= annual_withdrawal
                actual_withdrawal = annual_withdrawal

            yearly_data.append({
                "year": yr,
                "retirementYear": yr_offset + 1,
                "balance": round(balance, 2),
                "returnPct": ret,
                "withdrawalAmount": round(actual_withdrawal, 2),
                "inflationPct": cpi,
                "cumulativeInflation": round(cumulative_inflation, 4),
                "cashReserve": round(cash_reserve, 2) if cash_buffer_years > 0 else None,
            })

            if balance <= 0:
                survived = False
                for remaining in range(yr_offset + 1, horizon):
                    yearly_data.append({
                        "year": start_year + remaining,
                        "retirementYear": remaining + 1,
                        "balance": 0,
                        "returnPct": returns_by_year.get(start_year + remaining, 0),
                        "withdrawalAmount": 0,
                        "inflationPct": cpi_by_year.get(start_year + remaining, 0),
                        "cumulativeInflation": round(cumulative_inflation, 4),
                        "cashReserve": 0,
                    })
                break

            # Adjust withdrawal for inflation (fixed & combined strategies)
            if strategy in ("fixed", "combined"):
                annual_withdrawal *= (1 + cpi)
            elif strategy == "dividend":
                # Dividend growth replaces inflation adjustment
                div_yield_adj = div_yield  # yield stays same, applied to growing balance

        if survived:
            success_count += 1

        scenarios.append({
            "startYear": start_year,
            "endYear": end_year,
            "survived": survived,
            "finalBalance": round(yearly_data[-1]["balance"], 2),
            "data": yearly_data,
        })

    success_rate = round(success_count / total_count * 100, 1) if total_count > 0 else 0
    avg_final = round(sum(s["finalBalance"] for s in scenarios if s["survived"]) / max(success_count, 1), 2)
    worst = min(scenarios, key=lambda s: s["finalBalance"]) if scenarios else None
    best = max(scenarios, key=lambda s: s["finalBalance"]) if scenarios else None

    return {
        "horizon": horizon,
        "totalScenarios": total_count,
        "successCount": success_count,
        "failureCount": total_count - success_count,
        "successRate": success_rate,
        "avgFinalBalance": avg_final,
        "worstStartYear": worst["startYear"] if worst else None,
        "worstFinalBalance": worst["finalBalance"] if worst else None,
        "bestStartYear": best["startYear"] if best else None,
        "bestFinalBalance": best["finalBalance"] if best else None,
        "scenarios": scenarios,
    }
