"""Tests for Budget & Net Worth routes and import logic."""

import json
import pytest


# ── Sample budget data ────────────────────────────────────────────────

SAMPLE_BUDGET = {
    "year": 2026,
    "currency": "$",
    "goals": ["Save 24k"],
    "categories": [
        {
            "id": "income", "name": "Income", "type": "income",
            "subcategories": [
                {"name": "Salary A", "budgeted": 4000},
                {"name": "Salary B", "budgeted": 3000},
            ]
        },
        {
            "id": "essential", "name": "Essential Expenses", "type": "expense",
            "subcategories": [
                {"name": "Rent", "budgeted": 1000},
                {"name": "Groceries", "budgeted": 500},
            ]
        },
        {
            "id": "discretionary", "name": "Discretionary Expenses", "type": "expense",
            "subcategories": [{"name": "Fun", "budgeted": 300}]
        },
        {
            "id": "debt", "name": "Debt Payments", "type": "expense",
            "subcategories": [{"name": "Phone", "budgeted": 55}]
        },
        {
            "id": "savings", "name": "Savings", "type": "savings",
            "subcategories": [{"name": "Emergency", "budgeted": 500}]
        },
        {
            "id": "investments", "name": "Investments", "type": "investments",
            "subcategories": [{"name": "Stocks", "budgeted": 200}]
        },
    ],
    "months": {
        "january": {
            "actuals": {
                "income": {"Salary A": 4200, "Salary B": 3000},
                "essential": {"Rent": 1000, "Groceries": 450},
                "discretionary": {"Fun": 350},
                "debt": {"Phone": 55},
                "savings": {"Emergency": 500},
                "investments": {"Stocks": 200},
            },
            "rollover": False,
        }
    },
}

SAMPLE_NET_WORTH = {
    "assets": [
        {"name": "HYSA", "group": "bankAccounts"},
        {"name": "Stocks", "group": "investments"},
    ],
    "liabilities": [
        {"name": "Phone Loan", "group": "debt"},
    ],
    "snapshots": [
        {
            "month": "start", "year": 2026,
            "assets": {"HYSA": 10000, "Stocks": 30000},
            "liabilities": {"Phone Loan": 1000},
        },
        {
            "month": "january", "year": 2026,
            "assets": {"HYSA": 12000, "Stocks": 32000},
            "liabilities": {"Phone Loan": 945},
        },
    ],
}


def _seed_budget(client):
    """Seed portfolio with budget and net worth data."""
    from services.data_store import load_portfolio, save_portfolio
    portfolio = load_portfolio()
    portfolio["budget"] = json.loads(json.dumps(SAMPLE_BUDGET))
    portfolio["netWorth"] = json.loads(json.dumps(SAMPLE_NET_WORTH))
    save_portfolio(portfolio)


# ── Budget API tests ──────────────────────────────────────────────────

class TestBudgetApi:

    def test_get_budget_empty(self, client):
        res = client.get("/api/budget")
        assert res.status_code == 200
        data = res.get_json()
        assert "budget" in data

    def test_get_budget_with_data(self, client):
        _seed_budget(client)
        res = client.get("/api/budget")
        data = res.get_json()
        budget = data["budget"]
        assert budget["year"] == 2026
        assert len(budget["categories"]) == 6
        assert "january" in budget["summaries"]
        summary = budget["summaries"]["january"]
        assert summary["totalIncome"] == 7200
        assert summary["totalExpenses"] == 1855  # 1450 + 350 + 55
        assert summary["savingsRate"] > 0
        # Rollover amounts should be present
        assert "rolloverAmounts" in budget

    def test_update_actual(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/actual", json={
            "month": "january", "categoryId": "essential",
            "subcategory": "Rent", "amount": 1050
        })
        assert res.get_json()["ok"]
        # Verify it persisted
        res2 = client.get("/api/budget")
        actuals = res2.get_json()["budget"]["months"]["january"]["actuals"]
        assert actuals["essential"]["Rent"] == 1050

    def test_update_actual_missing_fields(self, client):
        res = client.post("/api/budget/actual", json={"month": "january"})
        assert res.status_code == 400

    def test_update_override(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/override", json={
            "month": "january", "categoryId": "income",
            "subcategory": "Salary A", "amount": 5000
        })
        assert res.get_json()["ok"]
        res2 = client.get("/api/budget")
        overrides = res2.get_json()["budget"]["months"]["january"]["overrides"]
        assert overrides["income"]["Salary A"] == 5000

    def test_clear_override(self, client):
        _seed_budget(client)
        # Set an override first
        client.post("/api/budget/override", json={
            "month": "january", "categoryId": "income",
            "subcategory": "Salary A", "amount": 5000
        })
        # Clear it by passing null
        res = client.post("/api/budget/override", json={
            "month": "january", "categoryId": "income",
            "subcategory": "Salary A", "amount": None
        })
        assert res.get_json()["ok"]

    def test_update_goals(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/goals", json={"goals": ["New goal 1", "New goal 2"]})
        assert res.get_json()["ok"]

    def test_add_subcategory(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/category", json={
            "categoryId": "essential",
            "subcategory": {"name": "Internet", "budgeted": 50}
        })
        assert res.get_json()["ok"]
        assert res.get_json()["action"] == "added"

    def test_delete_subcategory(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/subcategory/delete", json={
            "categoryId": "essential", "name": "Groceries"
        })
        assert res.get_json()["ok"]
        # Verify removed
        res2 = client.get("/api/budget")
        essential = [c for c in res2.get_json()["budget"]["categories"] if c["id"] == "essential"][0]
        names = [s["name"] for s in essential["subcategories"]]
        assert "Groceries" not in names

    def test_delete_subcategory_not_found(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/subcategory/delete", json={
            "categoryId": "essential", "name": "NonExistent"
        })
        assert res.status_code == 404

    def test_annual_summary(self, client):
        _seed_budget(client)
        res = client.get("/api/budget/annual")
        data = res.get_json()
        assert "annual" in data
        assert data["annual"]["monthCount"] == 1
        assert len(data["annual"]["monthly"]) == 1
        # Detail grids and savings rates
        assert "detailGrids" in data["annual"]
        assert "savingsRates" in data["annual"]
        assert "annualNotes" in data["annual"]

    def test_annual_notes(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/annual/notes", json={"notes": "Test note for the year"})
        assert res.get_json()["ok"]
        # Verify saved
        res2 = client.get("/api/budget/annual")
        assert res2.get_json()["annual"]["annualNotes"] == "Test note for the year"


# ── Transaction CRUD tests ───────────────────────────────────────────

class TestTransactionCrud:

    def test_add_transaction(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/transaction", json={
            "month": "january", "categoryId": "essential",
            "subcategory": "Rent", "date": "2026-01-15",
            "amount": 1000, "notes": "January rent"
        })
        data = res.get_json()
        assert data["ok"]
        assert "id" in data

    def test_add_transaction_recomputes_actuals(self, client):
        _seed_budget(client)
        # Add transaction
        client.post("/api/budget/transaction", json={
            "month": "february", "categoryId": "essential",
            "subcategory": "Rent", "date": "2026-02-01", "amount": 1000
        })
        client.post("/api/budget/transaction", json={
            "month": "february", "categoryId": "essential",
            "subcategory": "Groceries", "date": "2026-02-05", "amount": 200
        })
        # Check actuals were recomputed
        res = client.get("/api/budget")
        feb = res.get_json()["budget"]["months"]["february"]
        assert feb["actuals"]["essential"]["Rent"] == 1000
        assert feb["actuals"]["essential"]["Groceries"] == 200

    def test_update_transaction(self, client):
        _seed_budget(client)
        # Add first
        res = client.post("/api/budget/transaction", json={
            "month": "january", "categoryId": "income",
            "subcategory": "Salary A", "amount": 4200
        })
        txn_id = res.get_json()["id"]
        # Update amount
        res2 = client.put(f"/api/budget/transaction/{txn_id}", json={
            "month": "january", "categoryId": "income", "amount": 4500
        })
        assert res2.get_json()["ok"]

    def test_delete_transaction(self, client):
        _seed_budget(client)
        # Add first
        res = client.post("/api/budget/transaction", json={
            "month": "january", "categoryId": "essential",
            "subcategory": "Rent", "amount": 1000
        })
        txn_id = res.get_json()["id"]
        # Delete
        res2 = client.delete(f"/api/budget/transaction/{txn_id}", json={
            "month": "january", "categoryId": "essential"
        })
        assert res2.get_json()["ok"]

    def test_delete_transaction_not_found(self, client):
        _seed_budget(client)
        res = client.delete("/api/budget/transaction/nonexistent", json={
            "month": "january", "categoryId": "essential"
        })
        assert res.status_code == 404

    def test_add_transaction_missing_fields(self, client):
        res = client.post("/api/budget/transaction", json={"month": "january"})
        assert res.status_code == 400


# ── Rollover tests ───────────────────────────────────────────────────

class TestRollover:

    def test_toggle_rollover(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/rollover", json={
            "month": "january", "enabled": True
        })
        assert res.get_json()["ok"]
        # Verify
        res2 = client.get("/api/budget")
        assert res2.get_json()["budget"]["months"]["january"]["rollover"] is True

    def test_rollover_missing_month(self, client):
        res = client.post("/api/budget/rollover", json={"enabled": True})
        assert res.status_code == 400


# ── Migration tests ──────────────────────────────────────────────────

class TestMigration:

    def test_migrate_actuals_to_transactions(self, client):
        _seed_budget(client)
        res = client.post("/api/budget/transactions/migrate", json={})
        data = res.get_json()
        assert data["ok"]
        assert data["migrated"] > 0

        # Verify transactions were created
        res2 = client.get("/api/budget")
        jan = res2.get_json()["budget"]["months"]["january"]
        assert "transactions" in jan
        # Should have transactions for income
        assert len(jan["transactions"]["income"]) >= 2

    def test_migrate_skips_existing_transactions(self, client):
        _seed_budget(client)
        # First migration
        res1 = client.post("/api/budget/transactions/migrate", json={})
        count1 = res1.get_json()["migrated"]
        # Second migration should skip
        res2 = client.post("/api/budget/transactions/migrate", json={})
        assert res2.get_json()["migrated"] == 0


# ── Net Worth API tests ──────────────────────────────────────────────

class TestNetWorthApi:

    def test_get_net_worth_empty(self, client):
        res = client.get("/api/net-worth")
        assert res.status_code == 200

    def test_get_net_worth_with_data(self, client):
        _seed_budget(client)
        res = client.get("/api/net-worth")
        nw = res.get_json()["netWorth"]
        assert len(nw["assets"]) == 2
        assert len(nw["liabilities"]) == 1
        assert len(nw["snapshots"]) == 2
        # Check computed fields
        jan = nw["snapshots"][1]
        assert jan["totalAssets"] == 44000
        assert jan["totalLiabilities"] == 945
        assert jan["netWorth"] == 43055
        assert jan["monthlyGrowth"] > 0

    def test_update_snapshot(self, client):
        _seed_budget(client)
        res = client.post("/api/net-worth/snapshot", json={
            "month": "february", "year": 2026,
            "assets": {"HYSA": 13000, "Stocks": 33000},
            "liabilities": {"Phone Loan": 890}
        })
        assert res.get_json()["ok"]
        assert res.get_json()["action"] == "added"

    def test_update_existing_snapshot(self, client):
        _seed_budget(client)
        res = client.post("/api/net-worth/snapshot", json={
            "month": "january", "year": 2026,
            "assets": {"HYSA": 12500}
        })
        assert res.get_json()["action"] == "updated"

    def test_update_snapshot_cell(self, client):
        _seed_budget(client)
        res = client.post("/api/net-worth/snapshot/cell", json={
            "month": "january", "type": "asset", "name": "HYSA", "value": 13000
        })
        assert res.get_json()["ok"]
        # Verify
        res2 = client.get("/api/net-worth")
        jan = [s for s in res2.get_json()["netWorth"]["snapshots"] if s["month"] == "january"][0]
        assert jan["assets"]["HYSA"] == 13000

    def test_update_snapshot_cell_creates_snapshot(self, client):
        _seed_budget(client)
        res = client.post("/api/net-worth/snapshot/cell", json={
            "month": "february", "type": "asset", "name": "HYSA", "value": 14000
        })
        assert res.get_json()["ok"]

    def test_update_snapshot_cell_missing_fields(self, client):
        res = client.post("/api/net-worth/snapshot/cell", json={
            "month": "january", "type": "asset"
        })
        assert res.status_code == 400

    def test_add_asset(self, client):
        _seed_budget(client)
        res = client.post("/api/net-worth/asset", json={
            "type": "asset", "name": "Crypto", "group": "investments"
        })
        assert res.get_json()["ok"]

    def test_add_duplicate_asset(self, client):
        _seed_budget(client)
        res = client.post("/api/net-worth/asset", json={
            "type": "asset", "name": "HYSA", "group": "bankAccounts"
        })
        assert res.status_code == 400

    def test_remove_liability(self, client):
        _seed_budget(client)
        res = client.post("/api/net-worth/asset", json={
            "type": "liability", "name": "Phone Loan", "action": "remove"
        })
        assert res.get_json()["ok"]


# ── Import function tests ────────────────────────────────────────────

class TestBudgetImport:

    def test_import_categories_structure(self):
        """Verify import produces correct category structure."""
        from services.budget_import import _parse_categories
        import openpyxl

        # Create a mock workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        # Set up income category (cols A, C)
        ws["A8"] = "INGRESOS"
        ws["A10"] = "Salary"
        ws["C10"] = 5000
        ws["A11"] = "Bonus"
        ws["C11"] = 0

        categories = _parse_categories(ws)
        assert len(categories) == 6
        assert categories[0]["id"] == "income"
        assert len(categories[0]["subcategories"]) == 2
        assert categories[0]["subcategories"][0]["name"] == "Salary"
        assert categories[0]["subcategories"][0]["budgeted"] == 5000

    def test_synthetic_transactions(self):
        """Verify synthetic transaction creation from actuals."""
        from services.budget_import import _synthetic_transactions
        actuals = {
            "income": {"Salary": 5000, "Bonus": 0},
            "essential": {"Rent": 1000},
        }
        txns = _synthetic_transactions(actuals)
        assert "income" in txns
        assert len(txns["income"]) == 1  # Bonus=0 is skipped
        assert txns["income"][0]["subcategory"] == "Salary"
        assert txns["income"][0]["amount"] == 5000
        assert "id" in txns["income"][0]
