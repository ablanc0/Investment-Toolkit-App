"""
Portfolio Fidelity 2024 — Standalone Investment Dashboard Server
Pulls live market data from Yahoo Finance via yfinance.
"""

import json
import time
import threading
from concurrent.futures import ThreadPoolExecutor
import requests as http_requests
from pathlib import Path
from datetime import datetime, timedelta
from flask import Flask, jsonify, request, send_from_directory

import re
import yfinance as yf
from finvizfinance.quote import finvizfinance

# ── Config ──────────────────────────────────────────────────────────────
import os
import platform

BASE_DIR = Path(__file__).parent

def _resolve_data_dir():
    """Resolve the data directory: env var > config.json > Google Drive default > local."""
    # 1. Environment variable override
    env_dir = os.environ.get("INVTOOLKIT_DATA_DIR")
    if env_dir:
        p = Path(env_dir)
        if p.exists():
            return p

    # 2. config.json in the app folder
    config_file = BASE_DIR / "config.json"
    if config_file.exists():
        try:
            cfg = json.loads(config_file.read_text())
            p = Path(cfg.get("dataDir", ""))
            if p.exists():
                return p
        except Exception:
            pass

    # 3. Google Drive default paths (Mac / Windows)
    home = Path.home()
    gdrive_candidates = [
        # macOS Google Drive
        home / "Library" / "CloudStorage" / "GoogleDrive-ale.blancoglez91@gmail.com" / "My Drive" / "Investments" / "portfolio-app",
        # Windows Google Drive (stream)
        Path("G:/My Drive/Investments/portfolio-app"),
        # Windows Google Drive (mirror)
        home / "Google Drive" / "My Drive" / "Investments" / "portfolio-app",
    ]
    for candidate in gdrive_candidates:
        if candidate.exists():
            return candidate

    # 4. Fallback: local directory (for development)
    return BASE_DIR

DATA_DIR = _resolve_data_dir()
PORTFOLIO_FILE = DATA_DIR / "portfolio.json"
CACHE_FILE = DATA_DIR / "cache.json"
CACHE_TTL = 300  # 5 minutes

app = Flask(__name__, static_folder="static")

# ── Cache ───────────────────────────────────────────────────────────────
_cache = {}
_cache_lock = threading.Lock()


def load_disk_cache():
    global _cache
    if CACHE_FILE.exists():
        try:
            _cache = json.loads(CACHE_FILE.read_text())
        except Exception:
            _cache = {}


def save_disk_cache():
    try:
        CACHE_FILE.write_text(json.dumps(_cache, default=str))
    except Exception:
        pass


def cache_get(key):
    with _cache_lock:
        entry = _cache.get(key)
        if entry and (time.time() - entry.get("ts", 0)) < CACHE_TTL:
            return entry["data"]
    return None


def cache_set(key, data):
    with _cache_lock:
        _cache[key] = {"ts": time.time(), "data": data}
        save_disk_cache()


# ── Portfolio I/O ───────────────────────────────────────────────────────
def load_portfolio():
    if PORTFOLIO_FILE.exists():
        return json.loads(PORTFOLIO_FILE.read_text())
    return {"positions": [], "watchlist": [], "cash": 0, "goals": {}, "targets": {}, "strategy": []}


def save_portfolio(data):
    PORTFOLIO_FILE.write_text(json.dumps(data, indent=2))


# ── yfinance helpers ────────────────────────────────────────────────────
def fetch_ticker_data(ticker):
    """Fetch quote data for a single ticker via yfinance, with cache."""
    cached = cache_get(f"yf_{ticker}")
    if cached:
        return cached

    try:
        t = yf.Ticker(ticker)
        info = t.info or {}

        # yfinance returns a lot — we extract what we need
        data = {
            "price": info.get("currentPrice") or info.get("regularMarketPrice") or info.get("previousClose", 0),
            "previousClose": info.get("previousClose") or info.get("regularMarketPreviousClose", 0),
            "name": info.get("longName") or info.get("shortName", ticker),
            "marketCap": info.get("marketCap", 0),
            "pe": info.get("trailingPE", 0),
            "forwardPE": info.get("forwardPE", 0),
            "sector": info.get("sector", ""),
            "industry": info.get("industry", ""),
            "divYield": round(info.get("dividendYield") or info.get("trailingAnnualDividendYield") or 0, 2),
            "divRate": info.get("dividendRate") or info.get("trailingAnnualDividendRate", 0),
            "beta": info.get("beta", 0),
            "fiftyTwoWeekHigh": info.get("fiftyTwoWeekHigh", 0),
            "fiftyTwoWeekLow": info.get("fiftyTwoWeekLow", 0),
            "targetMeanPrice": info.get("targetMeanPrice", 0),
        }

        # Calculate day change %
        price = data["price"]
        prev = data["previousClose"]
        if prev and prev > 0:
            data["changePercent"] = round((price - prev) / prev * 100, 2)
        else:
            data["changePercent"] = 0

        cache_set(f"yf_{ticker}", data)
        return data

    except Exception as e:
        print(f"[yfinance] Error fetching {ticker}: {e}")
        # Return stale cache if available
        with _cache_lock:
            entry = _cache.get(f"yf_{ticker}")
            if entry:
                return entry["data"]
        return {"price": 0, "previousClose": 0, "name": ticker, "changePercent": 0}


def fetch_all_quotes(tickers):
    """Fetch quotes for a list of tickers."""
    results = {}
    for ticker in tickers:
        results[ticker] = fetch_ticker_data(ticker)
    return results


def fetch_dividends(ticker):
    """Fetch dividend history for a ticker."""
    cached = cache_get(f"divs_{ticker}")
    if cached is not None:
        return cached

    try:
        t = yf.Ticker(ticker)
        divs = t.dividends  # pandas Series indexed by date
        result = []
        for date, amount in divs.items():
            result.append({
                "date": date.strftime("%Y-%m-%d"),
                "dividend": float(amount),
            })
        cache_set(f"divs_{ticker}", result)
        return result
    except Exception as e:
        print(f"[yfinance] Error fetching dividends for {ticker}: {e}")
        return []


# ── Routes ──────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory("static", "dashboard.html")


@app.route("/api/portfolio")
def api_portfolio():
    """Main endpoint: returns enriched portfolio data with live prices."""
    portfolio = load_portfolio()
    pos_list = portfolio.get("positions", [])
    tickers = [p["ticker"] for p in pos_list]

    quotes = fetch_all_quotes(tickers)

    # Build IV lookup from intrinsicValues
    iv_list = portfolio.get("intrinsicValues", [])
    iv_map = {}
    for iv in iv_list:
        t = iv.get("ticker", "")
        if t:
            iv_map[t] = iv

    # Build total dividends received lookup from dividendLog
    div_log = portfolio.get("dividendLog", [])
    total_divs_received = {}
    for entry in div_log:
        for key, val in entry.items():
            if key not in ("year", "month", "cashInterest", "total") and val:
                total_divs_received[key] = total_divs_received.get(key, 0) + float(val or 0)

    enriched = []
    total_market_value = 0
    total_cost_basis = 0
    total_day_change = 0
    total_annual_div_income = 0

    for p in pos_list:
        ticker = p["ticker"]
        q = quotes.get(ticker, {})

        price = q.get("price", 0)
        prev_close = q.get("previousClose", price)
        day_change_pct = q.get("changePercent", 0)
        name = q.get("name", ticker)

        shares = p["shares"]
        avg_cost = p["avgCost"]
        cost_basis = shares * avg_cost
        market_value = shares * price
        market_return = market_value - cost_basis
        market_return_pct = (market_return / cost_basis * 100) if cost_basis > 0 else 0
        day_change_share = price - prev_close
        day_change_val = shares * day_change_share

        # Dividend fields
        div_rate = q.get("divRate", 0) or 0
        div_yield = q.get("divYield", 0) or 0
        # For ETFs where divRate is missing but divYield exists, derive rate from price
        if div_rate == 0 and div_yield > 0 and price > 0:
            div_rate = round(price * div_yield / 100, 4)
        annual_div_income = div_rate * shares
        yield_on_cost = (div_rate / avg_cost * 100) if avg_cost > 0 else 0
        divs_received = total_divs_received.get(ticker, 0)
        total_return_val = market_return + divs_received
        total_return_pct = (total_return_val / cost_basis * 100) if cost_basis > 0 else 0
        annual_shares_purch = (annual_div_income / price) if price > 0 else 0

        # IV fields
        iv_data = iv_map.get(ticker, {})
        intrinsic_value = iv_data.get("intrinsicValue", 0) or 0
        invt_score_data = iv_data.get("invtScore")
        invt_score = invt_score_data.get("score", 0) if isinstance(invt_score_data, dict) else 0
        dist_from_iv = ((price - intrinsic_value) / intrinsic_value) if intrinsic_value > 0 else 0
        dist_from_avg = ((price - avg_cost) / avg_cost) if avg_cost > 0 else 0

        # IV Signal
        if intrinsic_value > 0:
            if dist_from_iv > 0.50:
                iv_signal = "Overrated"
            elif dist_from_iv > 0.20:
                iv_signal = "Expensive"
            elif dist_from_iv < -0.05:
                iv_signal = "Strong Buy"
            elif dist_from_iv < 0.05:
                iv_signal = "Buy"
            else:
                iv_signal = "Hold"
        else:
            iv_signal = ""

        # Avg Cost Signal
        if dist_from_avg > 0.50:
            avg_cost_signal = "Overrated"
        elif dist_from_avg > 0.20:
            avg_cost_signal = "Expensive"
        elif dist_from_avg < -0.05:
            avg_cost_signal = "Strong Buy"
        elif dist_from_avg < 0.05:
            avg_cost_signal = "Buy"
        else:
            avg_cost_signal = "Hold"

        total_market_value += market_value
        total_cost_basis += cost_basis
        total_day_change += day_change_val
        total_annual_div_income += annual_div_income

        enriched.append({
            "ticker": ticker,
            "company": name,
            "name": name,
            "shares": shares,
            "avgCost": avg_cost,
            "price": round(price, 2),
            "prevClose": round(prev_close, 2),
            "costBasis": round(cost_basis, 2),
            "marketValue": round(market_value, 2),
            "mktValue": round(market_value, 2),
            "marketReturn": round(market_return, 2),
            "marketReturnPct": round(market_return_pct, 2),
            "totalReturn": round(total_return_val, 2),
            "returnPercent": round(total_return_pct, 2),
            "totalRetPct": round(total_return_pct, 2),
            "dayChangeShare": round(day_change_share, 2),
            "dayChange": round(day_change_val, 2),
            "dayChangePercent": round(day_change_pct, 2),
            "dayChangePct": round(day_change_pct, 2),
            "weight": 0,
            "allocation": 0,
            "secType": p.get("secType", "Stocks"),
            "sector": p.get("sector", ""),
            "category": p.get("category", ""),
            "signal": "",
            "divYield": div_yield,
            "divRate": div_rate,
            "yieldOnCost": round(yield_on_cost, 2),
            "annualDivIncome": round(annual_div_income, 2),
            "totalDivsReceived": round(divs_received, 2),
            "annualSharesPurch": round(annual_shares_purch, 3),
            "intrinsicValue": round(intrinsic_value, 2),
            "invtScore": round(invt_score, 2),
            "distFromIV": round(dist_from_iv * 100, 2),
            "ivSignal": iv_signal,
            "distFromAvgCost": round(dist_from_avg * 100, 2),
            "avgCostSignal": avg_cost_signal,
            "pe": q.get("pe", 0),
            "marketCap": q.get("marketCap", 0),
            "beta": q.get("beta", 0),
            "fiftyTwoWeekHigh": q.get("fiftyTwoWeekHigh", 0),
            "fiftyTwoWeekLow": q.get("fiftyTwoWeekLow", 0),
            "targetPrice": q.get("targetMeanPrice", 0),
        })

    cash = portfolio.get("cash", 0)
    total_portfolio = total_market_value + cash
    total_return = total_market_value - total_cost_basis
    total_return_pct = (total_return / total_cost_basis * 100) if total_cost_basis > 0 else 0

    # Calculate allocations and weights
    for pos in enriched:
        alloc = round((pos["marketValue"] / total_portfolio * 100) if total_portfolio > 0 else 0, 2)
        pos["allocation"] = alloc
        pos["weight"] = alloc

    # Category allocation
    cat_alloc = {}
    for pos in enriched:
        cat = pos["category"]
        cat_alloc[cat] = round(cat_alloc.get(cat, 0) + pos["allocation"], 2)

    # Sector allocation
    sec_alloc = {}
    for pos in enriched:
        sec = pos["sector"]
        sec_alloc[sec] = round(sec_alloc.get(sec, 0) + pos["allocation"], 2)

    # Security type allocation
    type_alloc = {}
    for pos in enriched:
        st = pos["secType"]
        type_alloc[st] = round(type_alloc.get(st, 0) + pos["allocation"], 2)

    # Signals based on valuation
    for pos in enriched:
        ret_pct = pos["returnPercent"]
        if ret_pct > 50:
            pos["signal"] = "Overrated"
        elif ret_pct > 20:
            pos["signal"] = "Expensive"
        elif ret_pct < -5:
            pos["signal"] = "Strong Buy"
        elif ret_pct < 5:
            pos["signal"] = "Buy"
        else:
            pos["signal"] = "Hold"

    # Percent of total dividend income
    for pos in enriched:
        pos["pctOfTotalIncome"] = round((pos["annualDivIncome"] / total_annual_div_income * 100) if total_annual_div_income > 0 else 0, 2)

    # Goals array
    raw_goals = portfolio.get("goals", {})
    goals_array = []
    if raw_goals.get("portfolioTarget"):
        goals_array.append({"name": f"${raw_goals['portfolioTarget']:,} Portfolio Goal", "current": round(total_market_value, 2), "target": raw_goals["portfolioTarget"]})
    if raw_goals.get("dividendTarget"):
        goals_array.append({"name": f"${raw_goals['dividendTarget']:,} Annual Dividend Goal", "current": 0, "target": raw_goals["dividendTarget"]})
    if raw_goals.get("maxHoldings"):
        goals_array.append({"name": f"Diversification ({raw_goals['maxHoldings']} Holdings Max)", "current": len(enriched), "target": raw_goals["maxHoldings"]})

    day_change_pct_total = round((total_day_change / (total_market_value - total_day_change) * 100) if (total_market_value - total_day_change) > 0 else 0, 2)

    # Portfolio-level dividend metrics
    cash_weight = round((cash / total_portfolio * 100) if total_portfolio > 0 else 0, 2)
    portfolio_div_yield = round((total_annual_div_income / total_market_value * 100) if total_market_value > 0 else 0, 2)
    portfolio_yoc = round((total_annual_div_income / total_cost_basis * 100) if total_cost_basis > 0 else 0, 2)
    lifetime_divs = round(sum(total_divs_received.values()), 2)

    # Sold positions summary
    sold_list = portfolio.get("soldPositions", [])
    sold_market_return = 0
    for sp in sold_list:
        shares_s = float(sp.get("shares", 0))
        sell_price = float(sp.get("sellPrice", 0))
        buy_cost = float(sp.get("avgCost", 0))
        sold_market_return += (sell_price - buy_cost) * shares_s
    sold_total_return = round(sold_market_return, 2)

    return jsonify({
        "positions": enriched,
        "summary": {
            "marketValue": round(total_market_value, 2),
            "totalMarketValue": round(total_market_value, 2),
            "costBasis": round(total_cost_basis, 2),
            "totalCostBasis": round(total_cost_basis, 2),
            "totalReturn": round(total_return, 2),
            "totalReturnPct": round(total_return_pct, 2),
            "totalReturnPercent": round(total_return_pct, 2),
            "marketReturnPercent": round(total_return_pct, 2),
            "dayChange": round(total_day_change, 2),
            "dayChangePct": day_change_pct_total,
            "dayChangePercent": day_change_pct_total,
            "cash": cash,
            "cashWeight": cash_weight,
            "holdings": len(enriched),
            "totalPortfolio": round(total_portfolio, 2),
            "annualDivIncome": round(total_annual_div_income, 2),
            "monthlyDivIncome": round(total_annual_div_income / 12, 2),
            "weeklyDivIncome": round(total_annual_div_income / 52, 2),
            "dailyDivIncome": round(total_annual_div_income / 365, 2),
            "portfolioDivYield": portfolio_div_yield,
            "portfolioYOC": portfolio_yoc,
            "lifetimeDivsReceived": lifetime_divs,
            "soldReturn": sold_total_return,
            "soldPositionsCount": len(sold_list),
        },
        "allocations": {
            "category": cat_alloc,
            "sector": sec_alloc,
            "securityType": type_alloc,
        },
        "targets": portfolio.get("targets", {}),
        "goals": goals_array,
        "strategy": portfolio.get("strategy", []),
        "lastUpdated": datetime.now().isoformat(),
    })


@app.route("/api/watchlist")
def api_watchlist():
    """Watchlist with live quotes."""
    portfolio = load_portfolio()
    wl = portfolio.get("watchlist", [])
    tickers = [w["ticker"] if isinstance(w, dict) else w for w in wl]
    priorities = {(w["ticker"] if isinstance(w, dict) else w): (w.get("priority", "Low") if isinstance(w, dict) else "Low") for w in wl}

    quotes = fetch_all_quotes(tickers)

    result = []
    for ticker in tickers:
        q = quotes.get(ticker, {})
        price = q.get("price", 0)
        prev_close = q.get("previousClose", price)
        target_price = q.get("targetMeanPrice", 0)
        dist_pct = round(((price - target_price) / target_price * 100) if target_price > 0 else 0, 2)
        div_rate = q.get("divRate", 0) or 0
        div_yield = q.get("divYield", 0) or 0
        if div_rate == 0 and div_yield > 0 and price > 0:
            div_rate = round(price * div_yield / 100, 4)
        annual_income_100 = round(div_rate * 100, 2) if div_rate else 0
        cost_100_shares = round(price * 100, 2) if price else 0
        eps = q.get("pe", 0)
        pe = q.get("pe", 0)
        # EPS = Price / PE
        eps_val = round(price / pe, 2) if pe and pe > 0 else 0

        result.append({
            "ticker": ticker,
            "company": q.get("name", ticker),
            "name": q.get("name", ticker),
            "price": round(price, 2),
            "intrinsicValue": round(target_price, 2),
            "iv": round(target_price, 2),
            "pe": pe,
            "eps": eps_val,
            "marketCap": q.get("marketCap", 0),
            "priority": priorities.get(ticker, "Low"),
            "distance": dist_pct,
            "dist": dist_pct,
            "dayChangeShare": round(price - prev_close, 2),
            "dayChange": q.get("changePercent", 0),
            "dayChangePct": q.get("changePercent", 0),
            "sector": q.get("sector", ""),
            "industry": q.get("industry", ""),
            "divYield": div_yield,
            "divRate": div_rate,
            "annualIncome100": annual_income_100,
            "cost100Shares": cost_100_shares,
            "fiftyTwoWeekHigh": q.get("fiftyTwoWeekHigh", 0),
            "fiftyTwoWeekLow": q.get("fiftyTwoWeekLow", 0),
            "beta": q.get("beta", 0),
        })

    return jsonify({"watchlist": result, "lastUpdated": datetime.now().isoformat()})


@app.route("/api/dividends")
def api_dividends():
    """Dividend data for all holdings."""
    portfolio = load_portfolio()
    tickers = [p["ticker"] for p in portfolio.get("positions", [])]

    all_dividends = {}
    monthly_totals = {}
    total_received = 0

    for ticker in tickers:
        divs = fetch_dividends(ticker)
        shares = next((p["shares"] for p in portfolio["positions"] if p["ticker"] == ticker), 0)

        ticker_total = 0
        for d in divs:
            amount = d.get("dividend", 0) * shares
            ticker_total += amount
            total_received += amount

            date_str = d.get("date", "")
            if date_str:
                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                    key = dt.strftime("%Y-%m")
                    monthly_totals[key] = monthly_totals.get(key, 0) + amount
                except ValueError:
                    pass

        all_dividends[ticker] = round(ticker_total, 2)

    now = datetime.now()
    one_year_ago = now - timedelta(days=365)
    annual_estimate = sum(v for k, v in monthly_totals.items()
                         if k >= one_year_ago.strftime("%Y-%m"))

    return jsonify({
        "byHolding": all_dividends,
        "monthlyTotals": {k: round(v, 2) for k, v in sorted(monthly_totals.items())},
        "totalReceived": round(total_received, 2),
        "annualEstimate": round(annual_estimate, 2),
        "lastUpdated": datetime.now().isoformat(),
    })


@app.route("/api/quote/<ticker>")
def api_quote(ticker):
    """Single ticker quote."""
    data = fetch_ticker_data(ticker.upper())
    if data and data.get("price", 0) > 0:
        return jsonify(data)
    return jsonify({"error": "Not found"}), 404


# ── CRUD Endpoints for Inline Editing ───────────────────────────────────
@app.route("/api/position/update", methods=["POST"])
def api_position_update():
    """Update a field on a position. Body: {ticker, field, value}"""
    body = request.get_json()
    ticker = body.get("ticker")
    field = body.get("field")
    value = body.get("value")

    if not ticker or not field:
        return jsonify({"error": "ticker and field required"}), 400

    allowed_fields = {"shares", "avgCost", "category", "sector", "secType"}
    if field not in allowed_fields:
        return jsonify({"error": f"Field '{field}' not editable"}), 400

    portfolio = load_portfolio()
    for p in portfolio["positions"]:
        if p["ticker"] == ticker:
            if field in ("shares", "avgCost"):
                p[field] = float(value)
            else:
                p[field] = str(value)
            save_portfolio(portfolio)
            return jsonify({"ok": True, "position": p})

    return jsonify({"error": f"Ticker '{ticker}' not found"}), 404


@app.route("/api/position/add", methods=["POST"])
def api_position_add():
    """Add a new position. Body: {ticker, shares, avgCost, category, sector, secType}"""
    body = request.get_json()
    ticker = body.get("ticker", "").upper().strip()
    if not ticker:
        return jsonify({"error": "ticker required"}), 400

    portfolio = load_portfolio()

    # Check for duplicates
    for p in portfolio["positions"]:
        if p["ticker"] == ticker:
            return jsonify({"error": f"'{ticker}' already exists"}), 400

    new_pos = {
        "ticker": ticker,
        "shares": float(body.get("shares", 0)),
        "avgCost": float(body.get("avgCost", 0)),
        "category": body.get("category", "Growth"),
        "sector": body.get("sector", ""),
        "secType": body.get("secType", "Stocks"),
    }
    portfolio["positions"].append(new_pos)
    save_portfolio(portfolio)
    return jsonify({"ok": True, "position": new_pos})


@app.route("/api/position/delete", methods=["POST"])
def api_position_delete():
    """Delete a position. Body: {ticker}"""
    body = request.get_json()
    ticker = body.get("ticker", "").upper().strip()
    if not ticker:
        return jsonify({"error": "ticker required"}), 400

    portfolio = load_portfolio()
    original_len = len(portfolio["positions"])
    portfolio["positions"] = [p for p in portfolio["positions"] if p["ticker"] != ticker]

    if len(portfolio["positions"]) == original_len:
        return jsonify({"error": f"'{ticker}' not found"}), 404

    save_portfolio(portfolio)
    return jsonify({"ok": True})


@app.route("/api/watchlist/add", methods=["POST"])
def api_watchlist_add():
    """Add to watchlist. Body: {ticker, priority}"""
    body = request.get_json()
    ticker = body.get("ticker", "").upper().strip()
    if not ticker:
        return jsonify({"error": "ticker required"}), 400

    portfolio = load_portfolio()
    for w in portfolio.get("watchlist", []):
        t = w["ticker"] if isinstance(w, dict) else w
        if t == ticker:
            return jsonify({"error": f"'{ticker}' already on watchlist"}), 400

    portfolio.setdefault("watchlist", []).append({
        "ticker": ticker,
        "priority": body.get("priority", "Low"),
    })
    save_portfolio(portfolio)
    return jsonify({"ok": True})


@app.route("/api/watchlist/delete", methods=["POST"])
def api_watchlist_delete():
    """Remove from watchlist. Body: {ticker}"""
    body = request.get_json()
    ticker = body.get("ticker", "").upper().strip()
    if not ticker:
        return jsonify({"error": "ticker required"}), 400

    portfolio = load_portfolio()
    original = portfolio.get("watchlist", [])
    portfolio["watchlist"] = [w for w in original if (w["ticker"] if isinstance(w, dict) else w) != ticker]

    if len(portfolio["watchlist"]) == len(original):
        return jsonify({"error": f"'{ticker}' not on watchlist"}), 404

    save_portfolio(portfolio)
    return jsonify({"ok": True})


@app.route("/api/watchlist/update", methods=["POST"])
def api_watchlist_update():
    """Update watchlist item priority. Body: {ticker, priority}"""
    body = request.get_json()
    ticker = body.get("ticker", "").upper().strip()
    priority = body.get("priority", "Low")

    portfolio = load_portfolio()
    for w in portfolio.get("watchlist", []):
        if isinstance(w, dict) and w["ticker"] == ticker:
            w["priority"] = priority
            save_portfolio(portfolio)
            return jsonify({"ok": True})

    return jsonify({"error": f"'{ticker}' not on watchlist"}), 404


@app.route("/api/cash/update", methods=["POST"])
def api_cash_update():
    """Update cash balance. Body: {cash}"""
    body = request.get_json()
    portfolio = load_portfolio()
    portfolio["cash"] = float(body.get("cash", 0))
    save_portfolio(portfolio)
    return jsonify({"ok": True, "cash": portfolio["cash"]})


@app.route("/api/goals/update", methods=["POST"])
def api_goals_update():
    """Update goals. Body: {portfolioTarget, dividendTarget, maxHoldings}"""
    body = request.get_json()
    portfolio = load_portfolio()
    goals = portfolio.get("goals", {})
    for key in ["portfolioTarget", "dividendTarget", "maxHoldings", "cashReserveMin", "cashReserveMax"]:
        if key in body:
            goals[key] = float(body[key])
    portfolio["goals"] = goals
    save_portfolio(portfolio)
    return jsonify({"ok": True, "goals": goals})


# ── Generic CRUD helper ─────────────────────────────────────────────────
def crud_list(section):
    """GET: return a list section from portfolio.json."""
    portfolio = load_portfolio()
    return jsonify({section: portfolio.get(section, []), "lastUpdated": datetime.now().isoformat()})

def crud_add(section, item):
    """Add an item to a list section."""
    portfolio = load_portfolio()
    portfolio.setdefault(section, []).append(item)
    save_portfolio(portfolio)
    return jsonify({"ok": True, "item": item})

def crud_update(section, index, updates):
    """Update an item at index in a list section."""
    portfolio = load_portfolio()
    items = portfolio.get(section, [])
    if 0 <= index < len(items):
        items[index].update(updates)
        save_portfolio(portfolio)
        return jsonify({"ok": True, "item": items[index]})
    return jsonify({"error": "Index out of range"}), 404

def crud_delete(section, index):
    """Delete an item at index in a list section."""
    portfolio = load_portfolio()
    items = portfolio.get(section, [])
    if 0 <= index < len(items):
        removed = items.pop(index)
        save_portfolio(portfolio)
        return jsonify({"ok": True, "removed": removed})
    return jsonify({"error": "Index out of range"}), 404

def crud_replace(section, data):
    """Replace entire list section."""
    portfolio = load_portfolio()
    portfolio[section] = data
    save_portfolio(portfolio)
    return jsonify({"ok": True})


# ── Sold Positions ──────────────────────────────────────────────────────
@app.route("/api/sold-positions")
def api_sold_positions():
    return crud_list("soldPositions")

@app.route("/api/sold-positions/add", methods=["POST"])
def api_sold_positions_add():
    b = request.get_json()
    item = {
        "ticker": b.get("ticker", "").upper().strip(),
        "shares": float(b.get("shares", 0)),
        "buyDate": b.get("buyDate", ""),
        "sellDate": b.get("sellDate", ""),
        "avgCost": float(b.get("avgCost", 0)),
        "sellPrice": float(b.get("sellPrice", 0)),
        "category": b.get("category", ""),
        "notes": b.get("notes", ""),
    }
    item["gain"] = round((item["sellPrice"] - item["avgCost"]) * item["shares"], 2)
    item["gainPct"] = round(((item["sellPrice"] - item["avgCost"]) / item["avgCost"] * 100) if item["avgCost"] > 0 else 0, 2)
    return crud_add("soldPositions", item)

@app.route("/api/sold-positions/update", methods=["POST"])
def api_sold_positions_update():
    b = request.get_json()
    return crud_update("soldPositions", int(b.get("index", -1)), b.get("updates", {}))

@app.route("/api/sold-positions/delete", methods=["POST"])
def api_sold_positions_delete():
    b = request.get_json()
    return crud_delete("soldPositions", int(b.get("index", -1)))


# ── Dividend Log ────────────────────────────────────────────────────────
@app.route("/api/dividend-log")
def api_dividend_log():
    """GET: Returns full dividendLog array plus the list of active tickers from positions."""
    portfolio = load_portfolio()
    dividend_log = portfolio.get("dividendLog", [])
    
    # Get active tickers from positions
    positions = portfolio.get("positions", [])
    active_tickers = [p.get("ticker") for p in positions if p.get("ticker")]
    
    return jsonify({
        "dividendLog": dividend_log,
        "activeTickers": active_tickers,
        "lastUpdated": datetime.now().isoformat()
    })

@app.route("/api/dividend-log/update", methods=["POST"])
def api_dividend_log_update():
    """POST: Updates a specific cell in dividendLog: {year, month, ticker, value}."""
    b = request.get_json()
    year = b.get("year")
    month = b.get("month")
    ticker = b.get("ticker")
    value = float(b.get("value", 0))
    
    portfolio = load_portfolio()
    dividend_log = portfolio.get("dividendLog", [])
    
    # Find the matching entry
    entry = None
    for item in dividend_log:
        if item.get("year") == year and item.get("month") == month:
            entry = item
            break
    
    if not entry:
        return jsonify({"error": "Entry not found"}), 404
    
    # Update the ticker value
    if ticker in entry:
        entry[ticker] = value
    else:
        return jsonify({"error": f"Ticker {ticker} not found in entry"}), 404
    
    # Recalculate total for that month
    total = 0
    ticker_list = ["GOOGL", "MSFT", "DIS", "TGT", "SBUX", "JPM", "KO", "O", "VICI", "UNH", "VGT", "VOO", "VTI", "SCHD", "VIG", "VXUS", "BND"]
    for t in ticker_list:
        total += entry.get(t, 0)
    total += entry.get("cashInterest", 0)
    entry["total"] = round(total, 2)
    
    save_portfolio(portfolio)
    return jsonify({"ok": True, "entry": entry})

@app.route("/api/dividend-log/add-year", methods=["POST"])
def api_dividend_log_add_year():
    """POST: Adds a new year (12 months of empty entries): {year}."""
    b = request.get_json()
    year = b.get("year")
    
    if not year:
        return jsonify({"error": "Year is required"}), 400
    
    portfolio = load_portfolio()
    dividend_log = portfolio.get("dividendLog", [])
    
    # Check if year already exists
    months_in_year = [item for item in dividend_log if item.get("year") == year]
    if len(months_in_year) > 0:
        return jsonify({"error": f"Year {year} already exists"}), 400
    
    # Add 12 months for the new year
    month_names = ["January", "February", "March", "April", "May", "June", 
                   "July", "August", "September", "October", "November", "December"]
    ticker_list = ["GOOGL", "MSFT", "DIS", "TGT", "SBUX", "JPM", "KO", "O", "VICI", "UNH", "VGT", "VOO", "VTI", "SCHD", "VIG", "VXUS", "BND"]
    
    for month_name in month_names:
        entry = {"year": year, "month": month_name}
        for ticker in ticker_list:
            entry[ticker] = 0
        entry["cashInterest"] = 0.0
        entry["total"] = 0.0
        dividend_log.append(entry)
    
    portfolio["dividendLog"] = dividend_log
    save_portfolio(portfolio)
    return jsonify({"ok": True, "year": year, "monthsAdded": 12})


# ── Monthly Data ────────────────────────────────────────────────────────
@app.route("/api/monthly-data")
def api_monthly_data():
    """GET: Returns full monthlyData array. For each month, compute dividendIncome from dividendLog."""
    portfolio = load_portfolio()
    monthly_data = portfolio.get("monthlyData", [])
    dividend_log = portfolio.get("dividendLog", [])
    
    # Compute dividendIncome from dividendLog for each month
    for month_entry in monthly_data:
        year = month_entry.get("year")
        month_raw = month_entry.get("month", "")
        # monthlyData has "January 24", dividendLog has "January" — extract just month name
        month_name = month_raw.split(" ")[0] if month_raw else ""

        # Find matching dividend log entry by year + month name
        dividend_total = 0.0
        for div_entry in dividend_log:
            if div_entry.get("year") == year and div_entry.get("month") == month_name:
                dividend_total = div_entry.get("total", 0.0)
                break

        month_entry["dividendIncome"] = round(dividend_total, 2)
    
    # Build monthly income distribution matrix: months x years
    month_names = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
    now = datetime.now()
    current_year = now.year
    current_month_idx = now.month  # 1-based
    years = sorted(set(e.get("year") for e in monthly_data if e.get("year")))
    income_matrix = []
    for m_idx, m_name in enumerate(month_names, 1):
        row = {"month": m_name}
        prev_val = None
        for y in years:
            val = 0.0
            for e in monthly_data:
                if e.get("year") == y and e.get("month", "").startswith(m_name):
                    val = e.get("dividendIncome", 0.0)
                    break
            row[str(y)] = round(val, 2)
            # Only compute YOY if this month has already occurred in year y
            month_has_passed = y < current_year or (y == current_year and m_idx <= current_month_idx)
            if prev_val is not None and prev_val > 0 and month_has_passed:
                row[f"yoy_{y}"] = round((val - prev_val) / prev_val * 100, 1)
            prev_val = val
        income_matrix.append(row)

    # Yearly totals row
    totals = {"month": "Total"}
    prev_total = None
    for y in years:
        total = sum(r[str(y)] for r in income_matrix)
        totals[str(y)] = round(total, 2)
        if prev_total is not None and prev_total > 0 and y <= current_year:
            totals[f"yoy_{y}"] = round((total - prev_total) / prev_total * 100, 1)
        prev_total = total
    income_matrix.append(totals)

    return jsonify({
        "monthlyData": monthly_data,
        "incomeDistribution": income_matrix,
        "years": years,
        "lastUpdated": datetime.now().isoformat()
    })

@app.route("/api/monthly-data/update", methods=["POST"])
def api_monthly_data_update():
    """POST: Updates a specific cell: {index, field, value}. Only portfolioValue, contributions, accumulatedInvestment are editable."""
    b = request.get_json()
    index = int(b.get("index", -1))
    field = b.get("field")
    value = b.get("value")
    
    editable_fields = ["portfolioValue", "contributions", "accumulatedInvestment"]
    
    if field not in editable_fields:
        return jsonify({"error": f"Field {field} is not editable"}), 400
    
    portfolio = load_portfolio()
    monthly_data = portfolio.get("monthlyData", [])
    
    if not (0 <= index < len(monthly_data)):
        return jsonify({"error": "Index out of range"}), 404
    
    entry = monthly_data[index]
    entry[field] = float(value)
    
    # Recalculate totalReturn and totalReturnPct
    accumulated = entry.get("accumulatedInvestment", 0)
    portfolio_value = entry.get("portfolioValue", 0)
    
    entry["totalReturn"] = round(portfolio_value - accumulated, 2)
    
    if accumulated > 0:
        entry["totalReturnPct"] = round(((portfolio_value - accumulated) / accumulated) * 100, 2)
    else:
        entry["totalReturnPct"] = 0
    
    save_portfolio(portfolio)
    return jsonify({"ok": True, "entry": entry})


# ── Annual Data (computed from monthly and dividend log) ─────────────────────────────────
@app.route("/api/annual-data")
def api_annual_data():
    """GET: Computes annual summary from monthlyData and dividendLog."""
    portfolio = load_portfolio()
    monthly_data = portfolio.get("monthlyData", [])
    dividend_log = portfolio.get("dividendLog", [])
    existing_annual = portfolio.get("annualData", [])
    
    # Build a map of existing sp500YieldPct by year
    sp500_map = {}
    for item in existing_annual:
        year = item.get("year")
        sp500_yield = item.get("sp500YieldPct", 0)
        if year:
            sp500_map[str(year)] = sp500_yield
    
    # Group monthly data by year
    annual_by_year = {}
    for month_entry in monthly_data:
        year = month_entry.get("year")
        if not year:
            continue
        
        year_str = str(year)
        if year_str not in annual_by_year:
            annual_by_year[year_str] = []
        annual_by_year[year_str].append(month_entry)
    
    # Compute annual summaries
    result = []
    for year_str in sorted(annual_by_year.keys()):
        months = annual_by_year[year_str]
        
        # portfolioValue = last non-zero portfolioValue in that year
        portfolio_value = 0
        for month_entry in reversed(months):
            pv = month_entry.get("portfolioValue", 0)
            if pv != 0:
                portfolio_value = pv
                break
        
        # annualContributions = sum of contributions
        annual_contributions = sum(m.get("contributions", 0) for m in months)
        
        # dividendIncome = sum of dividendIncome from dividendLog for that year
        dividend_income = 0.0
        for div_entry in dividend_log:
            if div_entry.get("year") == int(year_str):
                dividend_income += div_entry.get("total", 0)
        
        # accumulatedInvestment of last month in that year
        last_accumulated = 0
        if months:
            last_accumulated = months[-1].get("accumulatedInvestment", 0)
        
        # totalReturn = portfolioValue - accumulatedInvestment
        total_return = portfolio_value - last_accumulated
        
        # totalReturnPct = totalReturn / accumulatedInvestment (as decimal, e.g. 0.06)
        if last_accumulated > 0:
            total_return_pct = total_return / last_accumulated
        else:
            total_return_pct = 0

        # Get sp500YieldPct from existing data
        sp500_yield = sp500_map.get(year_str, 0)

        annual_entry = {
            "year": year_str,
            "portfolioValue": round(portfolio_value, 2),
            "annualContributions": round(annual_contributions, 2),
            "dividendIncome": round(dividend_income, 2),
            "totalReturn": round(total_return, 2),
            "totalReturnPct": round(total_return_pct, 6),
            "dividendYield": round(dividend_income / portfolio_value if portfolio_value > 0 else 0, 6),
            "portfolioYieldPct": round(total_return_pct, 6),
            "sp500YieldPct": sp500_yield
        }
        
        result.append(annual_entry)
    
    return jsonify({"annualData": result, "lastUpdated": datetime.now().isoformat()})


# ── My Lab (Multi-Portfolio) ──────────────────────────────────────────────────
@app.route("/api/my-lab")
def api_my_lab():
    portfolio = load_portfolio()
    lab = portfolio.get("myLab", [])
    research = portfolio.get("labResearch", [])
    return jsonify({"myLab": lab, "labResearch": research, "lastUpdated": datetime.now().isoformat()})

@app.route("/api/my-lab/research", methods=["POST"])
def api_my_lab_research():
    """Compile ticker frequency across all portfolios."""
    portfolio = load_portfolio()
    lab = portfolio.get("myLab", [])
    freq = {}
    for p in lab:
        seen = set()
        for h in p.get("holdings", []):
            t = h.get("ticker", "").upper().strip()
            if t and t not in seen:
                seen.add(t)
                if t not in freq:
                    freq[t] = {"ticker": t, "companyName": h.get("companyName", t), "count": 0, "portfolios": []}
                freq[t]["count"] += 1
                freq[t]["portfolios"].append(p.get("name", ""))
    result = sorted(freq.values(), key=lambda x: x["count"], reverse=True)
    # Save updated research data
    portfolio["labResearch"] = [{"ticker": r["ticker"], "frequency": r["count"]} for r in result]
    save_portfolio(portfolio)
    return jsonify({"research": result, "totalPortfolios": len(lab)})

@app.route("/api/my-lab/add-portfolio", methods=["POST"])
def api_my_lab_add_portfolio():
    b = request.get_json()
    name = b.get("name", "").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    portfolio = load_portfolio()
    lab = portfolio.get("myLab", [])
    lab.append({"name": name, "holdings": [], "totalHoldings": 0, "totalMarketValue": 0, "totalAnnualDividend": 0})
    portfolio["myLab"] = lab
    save_portfolio(portfolio)
    return jsonify({"ok": True, "index": len(lab) - 1})

@app.route("/api/my-lab/add-holding", methods=["POST"])
def api_my_lab_add_holding():
    b = request.get_json()
    pi = int(b.get("portfolioIndex", -1))
    portfolio = load_portfolio()
    lab = portfolio.get("myLab", [])
    if pi < 0 or pi >= len(lab):
        return jsonify({"error": "Invalid portfolio index"}), 400
    holding = {
        "ticker": b.get("ticker", "").upper().strip(),
        "companyName": b.get("companyName", ""),
        "shares": float(b.get("shares", 0)),
        "sharePrice": float(b.get("sharePrice", 0)),
        "securityType": b.get("securityType", ""),
        "sector": b.get("sector", ""),
        "category": b.get("category", ""),
        "portfolioAllocation": 0, "annualDividend": 0, "dividendYield": 0,
        "marketValue": float(b.get("shares", 0)) * float(b.get("sharePrice", 0)),
        "annualDividendIncome": 0, "pctOfTotalIncome": 0,
    }
    lab[pi]["holdings"].append(holding)
    lab[pi]["totalHoldings"] = len(lab[pi]["holdings"])
    lab[pi]["totalMarketValue"] = round(sum(h.get("marketValue", 0) for h in lab[pi]["holdings"]), 2)
    lab[pi]["totalAnnualDividend"] = round(sum(h.get("annualDividendIncome", 0) for h in lab[pi]["holdings"]), 2)
    portfolio["myLab"] = lab
    save_portfolio(portfolio)
    return jsonify({"ok": True})

@app.route("/api/my-lab/delete-holding", methods=["POST"])
def api_my_lab_delete_holding():
    b = request.get_json()
    pi = int(b.get("portfolioIndex", -1))
    hi = int(b.get("holdingIndex", -1))
    portfolio = load_portfolio()
    lab = portfolio.get("myLab", [])
    if pi < 0 or pi >= len(lab):
        return jsonify({"error": "Invalid portfolio"}), 400
    holdings = lab[pi].get("holdings", [])
    if hi < 0 or hi >= len(holdings):
        return jsonify({"error": "Invalid holding"}), 400
    holdings.pop(hi)
    lab[pi]["totalHoldings"] = len(holdings)
    lab[pi]["totalMarketValue"] = round(sum(h.get("marketValue", 0) for h in holdings), 2)
    lab[pi]["totalAnnualDividend"] = round(sum(h.get("annualDividendIncome", 0) for h in holdings), 2)
    save_portfolio(portfolio)
    return jsonify({"ok": True})

@app.route("/api/my-lab/update-portfolio", methods=["POST"])
def api_my_lab_update_portfolio():
    b = request.get_json()
    pi = int(b.get("portfolioIndex", -1))
    portfolio = load_portfolio()
    lab = portfolio.get("myLab", [])
    if pi < 0 or pi >= len(lab):
        return jsonify({"error": "Invalid portfolio index"}), 400
    for field in ("name", "lastUpdate", "source"):
        if field in b:
            lab[pi][field] = b[field]
    portfolio["myLab"] = lab
    save_portfolio(portfolio)
    return jsonify({"ok": True})


# ── Intrinsic Values ───────────────────────────────────────────────────
@app.route("/api/intrinsic-values")
def api_intrinsic_values():
    return crud_list("intrinsicValues")

@app.route("/api/intrinsic-values/add", methods=["POST"])
def api_intrinsic_values_add():
    b = request.get_json()
    item = {
        "ticker": b.get("ticker", "").upper().strip(),
        "method": b.get("method", "DCF"),
        "intrinsicValue": float(b.get("intrinsicValue", 0)),
        "currentPrice": float(b.get("currentPrice", 0)),
        "marginOfSafety": float(b.get("marginOfSafety", 0)),
        "notes": b.get("notes", ""),
        "lastUpdated": b.get("lastUpdated", datetime.now().strftime("%Y-%m-%d")),
    }
    if item["intrinsicValue"] > 0 and item["currentPrice"] > 0:
        item["marginOfSafety"] = round((1 - item["currentPrice"] / item["intrinsicValue"]) * 100, 2)
    return crud_add("intrinsicValues", item)

@app.route("/api/intrinsic-values/update", methods=["POST"])
def api_intrinsic_values_update():
    b = request.get_json()
    return crud_update("intrinsicValues", int(b.get("index", -1)), b.get("updates", {}))

@app.route("/api/intrinsic-values/delete", methods=["POST"])
def api_intrinsic_values_delete():
    b = request.get_json()
    return crud_delete("intrinsicValues", int(b.get("index", -1)))


@app.route("/api/intrinsic-values/upsert", methods=["POST"])
def api_intrinsic_values_upsert():
    """Insert or update an IV entry by ticker. Used by Stock Analyzer Summary tab."""
    b = request.get_json()
    ticker = b.get("ticker", "").upper().strip()
    if not ticker:
        return jsonify({"error": "ticker required"}), 400

    portfolio = load_portfolio()
    iv_list = portfolio.get("intrinsicValues", [])

    item = {
        "ticker": ticker,
        "companyName": b.get("companyName", ""),
        "currentPrice": float(b.get("currentPrice", 0)),
        "intrinsicValue": float(b.get("intrinsicValue", 0)),
        "targetPrice": float(b.get("targetPrice", 0)),
        "distanceFromIntrinsic": float(b.get("distanceFromIntrinsic", 0)),
        "invtScore": b.get("invtScore", ""),
        "week52Low": float(b.get("week52Low", 0)),
        "week52High": float(b.get("week52High", 0)),
        "securityType": b.get("securityType", "Stocks"),
        "sector": b.get("sector", ""),
        "category": b.get("category", ""),
        "peRatio": float(b.get("peRatio", 0)),
        "eps": float(b.get("eps", 0)),
        "annualDividend": float(b.get("annualDividend", 0)),
        "dividendYield": float(b.get("dividendYield", 0)),
        "signal": b.get("signal", ""),
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    # Compute distanceFromIntrinsic if not provided
    if item["intrinsicValue"] > 0 and item["currentPrice"] > 0:
        item["distanceFromIntrinsic"] = round((item["currentPrice"] / item["intrinsicValue"]) - 1, 4)

    # Find existing entry by ticker
    found = -1
    for i, existing in enumerate(iv_list):
        if existing.get("ticker", "").upper() == ticker:
            found = i
            break

    if found >= 0:
        iv_list[found] = item
        action = "updated"
    else:
        iv_list.append(item)
        action = "added"

    portfolio["intrinsicValues"] = iv_list
    save_portfolio(portfolio)
    return jsonify({"ok": True, "action": action, "ticker": ticker})


# ── Super Investor Buys ────────────────────────────────────────────────
@app.route("/api/super-investor-buys")
def api_super_investor_buys():
    return crud_list("superInvestorBuys")

@app.route("/api/super-investor-buys/add", methods=["POST"])
def api_super_investor_buys_add():
    b = request.get_json()
    item = {
        "investor": b.get("investor", ""),
        "ticker": b.get("ticker", "").upper().strip(),
        "action": b.get("action", "Buy"),
        "shares": float(b.get("shares", 0)),
        "value": float(b.get("value", 0)),
        "date": b.get("date", ""),
        "quarter": b.get("quarter", ""),
        "source": b.get("source", "13F Filing"),
        "notes": b.get("notes", ""),
    }
    return crud_add("superInvestorBuys", item)

@app.route("/api/super-investor-buys/update", methods=["POST"])
def api_super_investor_buys_update():
    b = request.get_json()
    return crud_update("superInvestorBuys", int(b.get("index", -1)), b.get("updates", {}))

@app.route("/api/super-investor-buys/delete", methods=["POST"])
def api_super_investor_buys_delete():
    b = request.get_json()
    return crud_delete("superInvestorBuys", int(b.get("index", -1)))


# ── Super Investor 13F Routes ─────────────────────────────────────────

@app.route("/api/super-investors")
def api_super_investors_list():
    """List all available super investors."""
    return jsonify([
        {"key": k, "fund": v["fund"], "cik": v["cik"],
         "cached": k in _13f_cache,
         "fetchedAt": _13f_cache.get(k, {}).get("fetchedAt", ""),
         "quarter": _13f_cache.get(k, {}).get("quarter", ""),
         "holdingsCount": _13f_cache.get(k, {}).get("holdingsCount", 0)}
        for k, v in SUPER_INVESTORS.items()
    ])


@app.route("/api/super-investors/13f/<investor_key>")
def api_super_investor_13f(investor_key):
    """Fetch 13F holdings for one investor (uses cache if available)."""
    if investor_key in _13f_cache:
        return jsonify(_13f_cache[investor_key])
    try:
        result = _fetch_investor_13f(investor_key)
        if result is None:
            return jsonify({"error": f"Unknown investor: {investor_key}"}), 404
        return jsonify(result)
    except Exception as e:
        print(f"[13F] Error fetching {investor_key}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/super-investors/13f-all", methods=["POST"])
def api_super_investor_13f_all():
    """Fetch all investors in a background thread."""
    if _13f_progress.get("running"):
        return jsonify({"status": "already_running"})
    _13f_progress.update({"done": 0, "total": len(SUPER_INVESTORS),
                          "current": "", "results": {}, "running": True})
    def _bg():
        for i, key in enumerate(SUPER_INVESTORS):
            _13f_progress["current"] = key
            try:
                result = _fetch_investor_13f(key)
                _13f_progress["results"][key] = {
                    "ok": "error" not in (result or {}),
                    "holdingsCount": (result or {}).get("holdingsCount", 0),
                }
            except Exception as e:
                _13f_progress["results"][key] = {"ok": False, "error": str(e)}
            _13f_progress["done"] = i + 1
        _13f_progress["running"] = False
        _13f_progress["current"] = ""
    threading.Thread(target=_bg, daemon=True).start()
    return jsonify({"status": "started", "total": len(SUPER_INVESTORS)})



@app.route("/api/super-investors/13f-progress")
def api_super_investor_13f_progress():
    """Poll progress for the background 13F fetch."""
    return jsonify(_13f_progress)


@app.route("/api/super-investors/overlap", methods=["POST"])
def api_super_investor_overlap():
    """Compute ticker overlap across selected investors."""
    b = request.get_json()
    investors = b.get("investors", [])
    ticker_investors = {}  # ticker -> [{investor, value, shares}]
    for inv_key in investors:
        data = _13f_cache.get(inv_key)
        if not data or "holdings" not in data:
            continue
        for h in data["holdings"]:
            tk = h.get("ticker", h.get("cusip", ""))
            if tk not in ticker_investors:
                ticker_investors[tk] = []
            ticker_investors[tk].append({
                "investor": inv_key,
                "value": h["value"],
                "shares": h["shares"],
                "name": h.get("name", ""),
            })
    # Only tickers held by 2+ investors
    overlap = []
    for tk, entries in ticker_investors.items():
        if len(entries) >= 2:
            overlap.append({
                "ticker": tk,
                "name": entries[0]["name"],
                "heldBy": [e["investor"] for e in entries],
                "heldByCount": len(entries),
                "combinedValue": sum(e["value"] for e in entries),
            })
    overlap.sort(key=lambda x: x["heldByCount"], reverse=True)
    return jsonify(overlap)


@app.route("/api/super-investors/most-popular")
def api_super_investor_most_popular():
    """Top 50 most held stocks across all cached investors, ranked by investor count then value."""
    ticker_data = {}  # ticker -> {name, investors: set, totalValue, totalShares}
    for inv_key, data in _13f_cache.items():
        if "holdings" not in data:
            continue
        for h in data["holdings"]:
            tk = h.get("ticker", h.get("cusip", ""))
            if not tk or tk == h.get("cusip", ""):
                continue  # skip unresolved CUSIPs
            if tk not in ticker_data:
                ticker_data[tk] = {"name": h.get("name", ""), "investors": set(),
                                   "totalValue": 0, "totalShares": 0}
            ticker_data[tk]["investors"].add(inv_key)
            ticker_data[tk]["totalValue"] += h["value"]
            ticker_data[tk]["totalShares"] += h["shares"]
    # Convert to list, sort by investor count desc, then value desc
    popular = []
    for tk, d in ticker_data.items():
        popular.append({
            "ticker": tk, "name": d["name"],
            "investorCount": len(d["investors"]),
            "investors": sorted(d["investors"]),
            "totalValue": d["totalValue"],
            "totalShares": d["totalShares"],
        })
    popular.sort(key=lambda x: (x["investorCount"], x["totalValue"]), reverse=True)
    return jsonify({
        "popular": popular[:50],
        "cachedInvestors": len(_13f_cache),
        "totalInvestors": len(SUPER_INVESTORS),
    })


# ── Projections & Risk Scenarios ────────────────────────────────────────

def compute_projections(config, return_pct_override=None):
    """Compute year-by-year projection with monthly compounding."""
    start = config.get("startingValue", 0)
    monthly_add = config.get("monthlyContribution", 0)
    annual_return = (return_pct_override if return_pct_override is not None
                     else config.get("expectedReturnPct", 8)) / 100
    div_yield = config.get("dividendYieldPct", 0) / 100
    inflation = config.get("inflationPct", 0) / 100
    years = int(config.get("years", 20))
    monthly_rate = annual_return / 12

    rows = []
    balance = start
    total_contributions = start
    total_dividends = 0.0

    for year in range(0, years + 1):
        div_income = round(balance * div_yield, 2)
        real = round(balance / ((1 + inflation) ** year), 2) if inflation > 0 and year > 0 else round(balance, 2)
        rows.append({
            "year": year,
            "balance": round(balance, 2),
            "realBalance": real,
            "contributions": round(total_contributions, 2),
            "growth": round(balance - total_contributions, 2),
            "divIncome": div_income,
            "totalDividends": round(total_dividends, 2),
        })
        if year < years:
            for _ in range(12):
                balance = balance * (1 + monthly_rate) + monthly_add
            total_contributions += monthly_add * 12
            total_dividends += div_income

    return rows


def _projections_response(proj):
    """Build projections API response (no live yfinance calls)."""
    base_return = proj.get("expectedReturnPct", 8)
    table = {
        "base": compute_projections(proj),
        "bull": compute_projections(proj, return_pct_override=base_return + 2),
        "bear": compute_projections(proj, return_pct_override=max(base_return - 2, 0)),
    }
    return {"config": proj, "table": table}


def _normalize_proj_config(proj):
    """Normalize legacy projection config keys to standard names."""
    mapping = {
        "currentValue": "startingValue",
        "expectedGrowth": "expectedReturnPct",
        "dividendYield": "dividendYieldPct",
    }
    for old_key, new_key in mapping.items():
        if old_key in proj and new_key not in proj:
            val = proj.pop(old_key)
            # expectedGrowth was stored as decimal (0.07), convert to pct
            if old_key == "expectedGrowth":
                val = round(val * 100, 4) if val < 1 else val
            if old_key == "dividendYield":
                val = round(val * 100, 4) if val < 1 else val
            proj[new_key] = val
    # Defaults
    proj.setdefault("startingValue", 0)
    proj.setdefault("monthlyContribution", 0)
    proj.setdefault("expectedReturnPct", 8)
    proj.setdefault("dividendYieldPct", 0)
    proj.setdefault("inflationPct", 3)
    proj.setdefault("years", 20)
    return proj

@app.route("/api/projections")
def api_projections():
    portfolio = load_portfolio()
    proj = _normalize_proj_config(portfolio.get("projections", {}))
    return jsonify(_projections_response(proj))

@app.route("/api/projections/update", methods=["POST"])
def api_projections_update():
    b = request.get_json()
    portfolio = load_portfolio()
    proj = _normalize_proj_config(portfolio.get("projections", {}))
    for key in ["startingValue", "monthlyContribution", "expectedReturnPct", "years", "inflationPct", "dividendYieldPct"]:
        if key in b:
            proj[key] = float(b[key])
    portfolio["projections"] = proj
    save_portfolio(portfolio)
    return jsonify(_projections_response(proj))

@app.route("/api/risk-scenarios/update", methods=["POST"])
def api_risk_scenarios_update():
    b = request.get_json()
    portfolio = load_portfolio()
    portfolio["riskScenarios"] = b.get("scenarios", [])
    save_portfolio(portfolio)
    return jsonify({"ok": True})


# ── Stock Analyzer — Valuation Helpers ─────────────────────────────────

RISK_FREE_RATE = 0.0425      # 10Y Treasury
MARKET_RETURN  = 0.099        # S&P long-term avg
PERPETUAL_GROWTH = 0.025      # terminal growth
MARGIN_OF_SAFETY = 0.70
AAA_YIELD_BASELINE = 4.4      # Graham baseline
AAA_YIELD_CURRENT  = 5.3      # fallback; overridden by live FRED data
GRAHAM_BASE_PE     = 7.0      # P/E for no-growth company (Graham original: 8.5)
GRAHAM_CG          = 1.0      # growth multiplier (Graham original: 2.0)
GRAHAM_GROWTH_CAP  = 20.0     # max earnings growth % to avoid inflated IVs

SECTOR_AVERAGES = {
    "Technology":          {"pe": 30, "evEbitda": 20, "pb": 8},
    "Communication Services": {"pe": 18, "evEbitda": 12, "pb": 3},
    "Healthcare":          {"pe": 22, "evEbitda": 15, "pb": 4},
    "Financial Services":  {"pe": 14, "evEbitda": 10, "pb": 1.5},
    "Consumer Cyclical":   {"pe": 20, "evEbitda": 13, "pb": 4},
    "Consumer Defensive":  {"pe": 22, "evEbitda": 14, "pb": 5},
    "Industrials":         {"pe": 20, "evEbitda": 13, "pb": 4},
    "Energy":              {"pe": 12, "evEbitda": 6,  "pb": 1.8},
    "Utilities":           {"pe": 18, "evEbitda": 12, "pb": 2},
    "Real Estate":         {"pe": 35, "evEbitda": 20, "pb": 2},
    "Basic Materials":     {"pe": 15, "evEbitda": 9,  "pb": 2.5},
}

# ── SEC EDGAR API ──────────────────────────────────────────────────────

EDGAR_USER_AGENT = "InvToolkit ale.blancoglez91@gmail.com"
EDGAR_FACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
EDGAR_TICKERS_URL = "https://www.sec.gov/files/company_tickers.json"
_cik_map = {}  # ticker -> zero-padded CIK string, loaded once


def _load_cik_map():
    """Load SEC ticker->CIK mapping (~10k entries). Called once on first use."""
    global _cik_map
    if _cik_map:
        return
    try:
        r = http_requests.get(
            EDGAR_TICKERS_URL,
            headers={"User-Agent": EDGAR_USER_AGENT},
            timeout=15,
        )
        data = r.json()
        for entry in data.values():
            ticker = entry.get("ticker", "").upper()
            cik = str(entry.get("cik_str", "")).zfill(10)
            if ticker:
                _cik_map[ticker] = cik
        print(f"[EDGAR] Loaded {len(_cik_map)} ticker->CIK mappings")
    except Exception as e:
        print(f"[EDGAR] Failed to load CIK map: {e}")


def _get_cik(ticker):
    """Get zero-padded CIK for a ticker. Returns str or None."""
    if not _cik_map:
        _load_cik_map()
    return _cik_map.get(ticker.upper())


def _fetch_edgar_facts(ticker):
    """Fetch ALL XBRL data from SEC EDGAR companyfacts (1 call, ~3-7MB).
    Returns the 'facts' sub-dict or None on failure."""
    cik = _get_cik(ticker)
    if not cik:
        print(f"[EDGAR] No CIK found for {ticker}")
        return None
    try:
        url = EDGAR_FACTS_URL.format(cik=cik)
        r = http_requests.get(
            url,
            headers={"User-Agent": EDGAR_USER_AGENT},
            timeout=20,
        )
        if r.status_code != 200:
            print(f"[EDGAR] HTTP {r.status_code} for {ticker} (CIK {cik})")
            return None
        return r.json().get("facts")
    except Exception as e:
        print(f"[EDGAR] Request failed for {ticker}: {e}")
        return None


def _edgar_annual_values(facts, tag, unit="USD", ns="us-gaap", max_years=10):
    """Extract annual 10-K values for an XBRL tag.

    Filters: form=10-K, fp=FY. Deduplicates by end date (latest filed wins).
    Returns: {"2024": val, "2023": val, ...} limited to max_years.
    """
    entries = facts.get(ns, {}).get(tag, {}).get("units", {}).get(unit, [])
    by_end = {}
    for e in entries:
        if e.get("form") != "10-K" or e.get("fp") != "FY":
            continue
        end = e.get("end", "")
        if len(end) < 4:
            continue
        val = e.get("val")
        if val is None:
            continue
        filed = e.get("filed", "")
        if end not in by_end or filed > by_end[end][1]:
            by_end[end] = (val, filed)
    by_year = {end[:4]: v for end, (v, _) in by_end.items()}
    return {y: by_year[y] for y in sorted(by_year, reverse=True)[:max_years]}


def _edgar_latest(facts, tag, **kw):
    """Most recent annual value for a tag. Returns value or 0."""
    vals = _edgar_annual_values(facts, tag, max_years=1, **kw)
    return list(vals.values())[0] if vals else 0


def _edgar_with_fallbacks(facts, tags, **kw):
    """Try multiple XBRL tag names in order, return first non-empty."""
    for tag in tags:
        result = _edgar_annual_values(facts, tag, **kw)
        if result:
            return result
    return {}


def _edgar_merge_tags(facts, tags, **kw):
    """Merge data from multiple XBRL tags, filling gaps from later tags.
    Useful when companies change XBRL tags across years (e.g. REITs).
    Trims to max_years most recent entries."""
    max_years = kw.get("max_years", 11)
    merged = {}
    for tag in tags:
        result = _edgar_annual_values(facts, tag, **kw)
        for year, val in result.items():
            if year not in merged:
                merged[year] = val
    if len(merged) > max_years:
        keep = sorted(merged.keys())[-max_years:]
        merged = {y: merged[y] for y in keep}
    return merged


def _edgar_to_info(facts, yf_info):
    """Build unified info dict from EDGAR + yfinance. Mirrors _fmp_to_info()."""
    ocf = _edgar_latest(facts, "NetCashProvidedByUsedInOperatingActivities")
    capex = _edgar_latest(facts, "PaymentsToAcquirePropertyPlantAndEquipment")
    fcf = ocf - capex  # EDGAR capex is positive (payments)

    # Debt: try noncurrent + current, fallback to LongTermDebt
    debt_nc = _edgar_latest(facts, "LongTermDebtNoncurrent")
    debt_c = _edgar_latest(facts, "LongTermDebtCurrent")
    total_debt = debt_nc + debt_c
    if total_debt == 0:
        total_debt = _edgar_latest(facts, "LongTermDebt")

    total_cash = _edgar_latest(facts, "CashAndCashEquivalentsAtCarryingValue")
    equity = _edgar_latest(facts, "StockholdersEquity")
    eps = _edgar_latest(facts, "EarningsPerShareDiluted", unit="USD/shares")

    # Revenue: fallback chain for different accounting standards
    revenue_tags = [
        "RevenueFromContractWithCustomerExcludingAssessedTax",
        "Revenues",
        "SalesRevenueNet",
    ]
    rev_by_year = _edgar_with_fallbacks(facts, revenue_tags, max_years=1)
    revenue = list(rev_by_year.values())[0] if rev_by_year else 0

    # Shares outstanding
    shares = _edgar_latest(facts, "CommonStockSharesOutstanding", unit="shares")
    if not shares:
        shares = _edgar_latest(
            facts, "EntityCommonStockSharesOutstanding",
            unit="shares", ns="dei",
        )

    # Earnings growth: YoY from 2 years of EPS
    eps_by_year = _edgar_annual_values(
        facts, "EarningsPerShareDiluted", unit="USD/shares", max_years=3,
    )
    eps_years = sorted(eps_by_year.keys(), reverse=True)
    earnings_growth = 0
    if len(eps_years) >= 2:
        eps_curr = eps_by_year[eps_years[0]]
        eps_prev = eps_by_year[eps_years[1]]
        if eps_prev and eps_prev > 0:
            earnings_growth = (eps_curr - eps_prev) / eps_prev

    # Enterprise value: marketCap + debt - cash
    market_cap = yf_info.get("marketCap", 0) or 0
    ev_val = market_cap + total_debt - total_cash

    # EBITDA: operating income + D&A
    op_income = _edgar_latest(facts, "OperatingIncomeLoss")
    dda = _edgar_latest(facts, "DepreciationDepletionAndAmortization")
    ebitda = op_income + dda if (op_income or dda) else 0

    # Derived ratios
    price = yf_info.get("currentPrice") or yf_info.get("regularMarketPrice", 0)
    book_per_share = (equity / shares) if shares > 0 else 0
    ev_ebitda = (ev_val / ebitda) if ebitda > 0 else 0
    pe = (price / eps) if eps > 0 else 0
    pb = (price / book_per_share) if book_per_share > 0 else 0

    info = dict(yf_info)
    info.update({
        "totalDebt": total_debt,
        "totalCash": total_cash,
        "freeCashflow": fcf,
        "operatingCashflow": ocf,
        "trailingEps": eps,
        "totalRevenue": revenue,
        "sharesOutstanding": shares,
        "enterpriseValue": ev_val,
        "earningsGrowth": earnings_growth,
        "bookValue": round(book_per_share, 2),
        "enterpriseToEbitda": round(ev_ebitda, 2),
        "trailingPE": round(pe, 2),
        "priceToBook": round(pb, 2),
    })
    return info


def _edgar_to_financials(facts):
    """Translate EDGAR XBRL -> year-keyed dicts. Mirrors _fmp_to_financials().
    Returns (income, cashflow, balance) with up to 10 years of data."""
    # ── Income statement ──
    pretax_tags = [
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments",
    ]
    pretax_by_year = _edgar_with_fallbacks(facts, pretax_tags)
    tax_by_year = _edgar_annual_values(facts, "IncomeTaxExpenseBenefit")
    interest_by_year = _edgar_annual_values(facts, "InterestExpense")

    income = {}
    for year in pretax_by_year:
        income[year] = {
            "Pretax Income": pretax_by_year.get(year, 0),
            "Tax Provision": tax_by_year.get(year, 0),
            "Interest Expense": interest_by_year.get(year, 0),
        }

    # ── Cash flow statement ──
    ocf_by_year = _edgar_annual_values(
        facts, "NetCashProvidedByUsedInOperatingActivities",
    )
    capex_by_year = _edgar_annual_values(
        facts, "PaymentsToAcquirePropertyPlantAndEquipment",
    )
    cashflow = {}
    for year in set(ocf_by_year) | set(capex_by_year):
        raw_capex = capex_by_year.get(year, 0)
        cashflow[year] = {
            "Operating Cash Flow": ocf_by_year.get(year, 0),
            "Capital Expenditure": -abs(raw_capex),  # EDGAR positive -> code expects negative
        }

    # ── Balance sheet ──
    debt_nc = _edgar_annual_values(facts, "LongTermDebtNoncurrent")
    debt_c = _edgar_annual_values(facts, "LongTermDebtCurrent")
    debt_fb = _edgar_annual_values(facts, "LongTermDebt")
    cash_by_year = _edgar_annual_values(facts, "CashAndCashEquivalentsAtCarryingValue")
    equity_by_year = _edgar_annual_values(facts, "StockholdersEquity")

    balance = {}
    for year in set(cash_by_year) | set(equity_by_year):
        debt = debt_nc.get(year, 0) + debt_c.get(year, 0)
        if debt == 0:
            debt = debt_fb.get(year, 0)
        balance[year] = {
            "Total Debt": debt,
            "Cash And Cash Equivalents": cash_by_year.get(year, 0),
            "Stockholders Equity": equity_by_year.get(year, 0),
        }

    return income, cashflow, balance


# ── SEC EDGAR 13F — Super Investor Holdings ───────────────────────────

import xml.etree.ElementTree as ET

SUPER_INVESTORS = {
    "Warren Buffett":     {"cik": "0001067983", "fund": "Berkshire Hathaway"},
    "Michael Burry":      {"cik": "0001649339", "fund": "Scion Asset Management"},
    "Bill Ackman":        {"cik": "0001336528", "fund": "Pershing Square"},
    "Ray Dalio":          {"cik": "0001350694", "fund": "Bridgewater Associates"},
    "Seth Klarman":       {"cik": "0001061768", "fund": "Baupost Group"},
    "David Tepper":       {"cik": "0001656456", "fund": "Appaloosa Management"},
    "Howard Marks":       {"cik": "0000949509", "fund": "Oaktree Capital Management"},
    "Terry Smith":        {"cik": "0001569205", "fund": "Fundsmith LLP"},
    "Li Lu":              {"cik": "0001709323", "fund": "Himalaya Capital"},
}

_13F_CACHE_FILE = DATA_DIR / "13f_cache.json"
_13f_cache = {}  # investor_key -> {investor, fund, filingDate, quarter, holdings, totalValue}
_13f_progress = {"done": 0, "total": 0, "current": "", "results": {}, "running": False}


def _load_13f_cache():
    """Load persisted 13F data from disk on startup."""
    global _13f_cache
    if _13F_CACHE_FILE.exists():
        try:
            _13f_cache = json.loads(_13F_CACHE_FILE.read_text())
            print(f"[13F] Loaded {len(_13f_cache)} investors from disk cache")
        except Exception:
            _13f_cache = {}


def _save_13f_cache():
    """Persist 13F data to disk."""
    try:
        _13F_CACHE_FILE.write_text(json.dumps(_13f_cache, default=str))
    except Exception as e:
        print(f"[13F] Failed to save cache: {e}")


def _fetch_13f_latest(cik):
    """Get the most recent 13F-HR filing accession number and date."""
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    r = http_requests.get(url, headers={"User-Agent": EDGAR_USER_AGENT}, timeout=15)
    r.raise_for_status()
    data = r.json()
    recent = data.get("filings", {}).get("recent", {})
    forms = recent.get("form", [])
    accessions = recent.get("accessionNumber", [])
    dates = recent.get("filingDate", [])
    for i, form in enumerate(forms):
        if form in ("13F-HR", "13F-HR/A"):
            acc = accessions[i].replace("-", "")
            return {"accession": accessions[i], "accessionClean": acc, "filingDate": dates[i]}
    return None


def _fetch_13f_infotable(cik, accession_clean, accession_raw):
    """Download the infoTable XML from a 13F filing."""
    cik_num = cik.lstrip("0")
    index_url = f"https://www.sec.gov/Archives/edgar/data/{cik_num}/{accession_clean}/"
    r = http_requests.get(index_url, headers={"User-Agent": EDGAR_USER_AGENT}, timeout=15)
    r.raise_for_status()
    # Find the infotable XML filename in the index page
    matches = re.findall(r'href="([^"]*infotable[^"]*\.xml)"', r.text, re.IGNORECASE)
    if not matches:
        # Try alternative: primary_doc.xml pattern
        matches = re.findall(r'href="([^"]*\.xml)"', r.text, re.IGNORECASE)
    if not matches:
        return None
    xml_filename = matches[0]
    if xml_filename.startswith("http"):
        xml_url = xml_filename
    elif xml_filename.startswith("/"):
        xml_url = f"https://www.sec.gov{xml_filename}"
    else:
        xml_url = f"https://www.sec.gov/Archives/edgar/data/{cik_num}/{accession_clean}/{xml_filename}"
    r2 = http_requests.get(xml_url, headers={"User-Agent": EDGAR_USER_AGENT}, timeout=30)
    r2.raise_for_status()
    return r2.text


def _parse_13f_xml(xml_string):
    """Parse 13F infoTable XML into holdings list. Aggregates by CUSIP."""
    ns = {"ns": "http://www.sec.gov/edgar/document/thirteenf/informationtable"}
    root = ET.fromstring(xml_string)
    by_cusip = {}
    for entry in root.findall(".//ns:infoTable", ns):
        cusip = (entry.findtext("ns:cusip", "", ns) or "").strip()
        name = (entry.findtext("ns:nameOfIssuer", "", ns) or "").strip()
        value = int(entry.findtext("ns:value", "0", ns) or 0)  # value in dollars
        shares_el = entry.find("ns:shrsOrPrnAmt", ns)
        shares = int(shares_el.findtext("ns:sshPrnamt", "0", ns) or 0) if shares_el else 0
        put_call = (entry.findtext("ns:putCall", "", ns) or "").strip()
        if cusip in by_cusip:
            by_cusip[cusip]["value"] += value
            by_cusip[cusip]["shares"] += shares
        else:
            by_cusip[cusip] = {
                "cusip": cusip, "name": name, "value": value,
                "shares": shares, "putCall": put_call,
            }
    return list(by_cusip.values())


def _openfigi_batch(cusip_list, id_type="ID_CUSIP"):
    """Resolve a list of CUSIPs/CINS via OpenFIGI with rate limiting. Returns {cusip: ticker}."""
    ticker_map = {}
    batch_count = 0
    for i in range(0, len(cusip_list), 10):
        batch = cusip_list[i:i+10]
        body = [{"idType": id_type, "idValue": c} for c in batch]
        for attempt in range(3):
            try:
                r = http_requests.post(
                    "https://api.openfigi.com/v3/mapping",
                    json=body,
                    headers={"Content-Type": "application/json"},
                    timeout=30,
                )
                if r.status_code == 200:
                    results = r.json()
                    for j, item in enumerate(results):
                        if isinstance(item, dict) and "data" in item and item["data"]:
                            entries = item["data"]
                            us_entry = next((e for e in entries if e.get("exchCode") == "US"), None)
                            chosen = us_entry or entries[0]
                            ticker_map[batch[j]] = chosen.get("ticker", "")
                    break
                elif r.status_code == 429:
                    time.sleep(30)  # wait for rate limit reset
                else:
                    break
            except Exception:
                break
        batch_count += 1
        # Rate limit: 25 req/min for unauthenticated. Pause every 20 batches.
        if batch_count % 20 == 0:
            time.sleep(30)
    return ticker_map


def _resolve_cusips_to_tickers(holdings):
    """Batch resolve CUSIPs to tickers via OpenFIGI (free, no key, 10/batch, 25 req/min)."""
    cusips = list(dict.fromkeys(h["cusip"] for h in holdings if h.get("cusip")))
    if not cusips:
        return holdings
    ticker_map = _openfigi_batch(cusips, "ID_CUSIP")
    # Retry unresolved international CUSIPs (CINS codes starting with letter)
    unresolved_cins = [c for c in cusips if c not in ticker_map and c[0:1].isalpha()]
    if unresolved_cins:
        cins_map = _openfigi_batch(unresolved_cins, "ID_CINS")
        ticker_map.update(cins_map)
    for h in holdings:
        h["ticker"] = ticker_map.get(h["cusip"], h["cusip"])
    return holdings


def _fetch_investor_13f(investor_key):
    """Full 13F pipeline for one investor. Returns dict with holdings."""
    inv = SUPER_INVESTORS.get(investor_key)
    if not inv:
        return None
    cik = inv["cik"]
    # Step 1: Find latest 13F filing
    filing = _fetch_13f_latest(cik)
    if not filing:
        return {"investor": investor_key, "fund": inv["fund"], "error": "No 13F filing found"}
    # Step 2: Download infoTable XML
    xml = _fetch_13f_infotable(cik, filing["accessionClean"], filing["accession"])
    if not xml:
        return {"investor": investor_key, "fund": inv["fund"], "error": "Could not fetch infoTable XML"}
    # Step 3: Parse holdings
    holdings = _parse_13f_xml(xml)
    # Step 4: Resolve CUSIPs to tickers
    holdings = _resolve_cusips_to_tickers(holdings)
    # Sort by value descending
    holdings.sort(key=lambda h: h["value"], reverse=True)
    total_value = sum(h["value"] for h in holdings)
    # Add portfolio percentage
    for h in holdings:
        h["pctPortfolio"] = round(h["value"] / total_value * 100, 2) if total_value else 0
    # Derive quarter from filing date
    filing_date = filing["filingDate"]
    quarter = _derive_quarter(filing_date)
    result = {
        "investor": investor_key,
        "fund": inv["fund"],
        "filingDate": filing_date,
        "quarter": quarter,
        "holdings": holdings,
        "totalValue": total_value,
        "holdingsCount": len(holdings),
        "fetchedAt": datetime.now().isoformat(),
    }
    _13f_cache[investor_key] = result
    _save_13f_cache()
    return result


def _derive_quarter(filing_date):
    """Derive the reporting quarter from the filing date (filings are ~45 days after quarter end)."""
    try:
        dt = datetime.strptime(filing_date, "%Y-%m-%d")
        # 13F is due 45 days after quarter end, so filing in Feb = Q4 prev year, May = Q1, Aug = Q2, Nov = Q3
        month = dt.month
        year = dt.year
        if month <= 2:
            return f"Q4 {year - 1}"
        elif month <= 5:
            return f"Q1 {year}"
        elif month <= 8:
            return f"Q2 {year}"
        elif month <= 11:
            return f"Q3 {year}"
        else:
            return f"Q4 {year}"
    except:
        return ""


# ── FMP API (fallback) ─────────────────────────────────────────────────

FMP_API_KEY = "Yt3XCJh6dH3GNabskOSVMpQqKBzbSh70"
FMP_BASE = "https://financialmodelingprep.com/stable"


def _fmp_get(endpoint, **params):
    """Call FMP stable API. Returns parsed JSON or None on error."""
    params["apikey"] = FMP_API_KEY
    try:
        r = http_requests.get(f"{FMP_BASE}/{endpoint}", params=params, timeout=15)
        data = r.json()
        if isinstance(data, dict) and "Error Message" in data:
            print(f"[FMP] Error on {endpoint}: {data['Error Message'][:80]}")
            return None
        return data
    except Exception as e:
        print(f"[FMP] Request failed for {endpoint}: {e}")
        return None


def _fetch_fmp_dcf(symbol):
    """Fetch FMP's own DCF intrinsic value as an external benchmark."""
    data = _fmp_get("discounted-cash-flow", symbol=symbol)
    if data and isinstance(data, list) and len(data) > 0:
        return round(data[0].get("dcf", 0), 2)
    return None


def _fetch_fmp_benchmarks(symbol):
    """Fetch FMP key-metrics, ratings, and financial-scores in one call set."""
    result = {}
    # Key metrics → Graham Number, earnings yield, FCF yield, ROIC
    km = _fmp_get("key-metrics", symbol=symbol, period="annual")
    if km and isinstance(km, list) and len(km) > 0:
        latest = km[0]
        result["grahamNumber"] = round(latest.get("grahamNumber", 0) or 0, 2)
        result["earningsYield"] = round((latest.get("earningsYield", 0) or 0) * 100, 2)
        result["freeCashFlowYield"] = round((latest.get("freeCashFlowYield", 0) or 0) * 100, 2)
        result["roic"] = round((latest.get("returnOnInvestedCapital", 0) or 0) * 100, 2)

    # Ratings snapshot → overall letter grade + subscores
    rt = _fmp_get("ratings-snapshot", symbol=symbol)
    if rt and isinstance(rt, list) and len(rt) > 0:
        r = rt[0]
        result["rating"] = r.get("rating", "")
        result["ratingScore"] = r.get("overallScore", 0)
        result["ratingDcfScore"] = r.get("discountedCashFlowScore", 0)
        result["ratingPeScore"] = r.get("priceToEarningsScore", 0)
        result["ratingPbScore"] = r.get("priceToBookScore", 0)

    # Financial scores → Altman Z-Score, Piotroski Score
    fs = _fmp_get("financial-scores", symbol=symbol)
    if fs and isinstance(fs, list) and len(fs) > 0:
        f = fs[0]
        result["altmanZScore"] = round(f.get("altmanZScore", 0) or 0, 2)
        result["piotroskiScore"] = f.get("piotroskiScore", 0)

    return result


def _fetch_fred_aaa_yield():
    """Fetch latest AAA corporate bond yield from FRED (no API key needed).
    Returns (value, date_str) e.g. (5.30, '2026-02-01').
    """
    try:
        url = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=AAA&cosd=2025-01-01"
        r = http_requests.get(url, timeout=10)
        lines = r.text.strip().split("\n")
        if len(lines) >= 2:
            parts = lines[-1].split(",")
            return float(parts[1]), parts[0]
    except Exception as e:
        print(f"[FRED] Failed to fetch AAA yield: {e}")
    return AAA_YIELD_CURRENT, None  # fallback to constant


def _fetch_fmp_stock_data(ticker):
    """Fetch FMP financial data for valuation models (5 API calls).
    Profile/ratios come from yfinance to save FMP quota.
    """
    return {
        "income": _fmp_get("income-statement", symbol=ticker, period="annual"),
        "cashflow": _fmp_get("cash-flow-statement", symbol=ticker, period="annual"),
        "balance": _fmp_get("balance-sheet-statement", symbol=ticker, period="annual"),
        "ev": _fmp_get("enterprise-values", symbol=ticker, period="annual"),
        "growth": _fmp_get("financial-growth", symbol=ticker, period="annual"),
    }


def _fmp_to_info(fmp, yf_info):
    """Build unified info dict: FMP for financials, yfinance for profile/ratios.

    FMP provides 5 years of financial statements (vs yfinance 4) and
    is the uniform source for DCF inputs. yfinance fills profile fields,
    ratios, and supplementary data not available on FMP free tier.
    """
    inc0 = fmp["income"][0] if fmp.get("income") else {}
    cf0 = fmp["cashflow"][0] if fmp.get("cashflow") else {}
    bal0 = fmp["balance"][0] if fmp.get("balance") else {}
    ev0 = fmp["ev"][0] if fmp.get("ev") else {}
    gr0 = fmp["growth"][0] if fmp.get("growth") else {}

    shares = ev0.get("numberOfShares") or inc0.get("weightedAverageShsOut") or 0

    # Start with yfinance as base, then override financial fields from FMP
    info = dict(yf_info)
    # Derived ratios from FMP data (for Relative valuation)
    ebitda = inc0.get("ebitda", 0) or 0
    ev_val = ev0.get("enterpriseValue", 0) or 0
    equity = bal0.get("totalStockholdersEquity", 0) or 0
    eps = inc0.get("epsDiluted", 0) or 0
    price = yf_info.get("currentPrice") or yf_info.get("regularMarketPrice", 0)
    book_per_share = (equity / shares) if shares > 0 else 0
    ev_ebitda = (ev_val / ebitda) if ebitda > 0 else 0
    pe = (price / eps) if eps > 0 else 0
    pb = (price / book_per_share) if book_per_share > 0 else 0

    info.update({
        # FMP financial data (uniform source for DCF + Graham + Relative)
        "totalDebt": bal0.get("totalDebt", 0),
        "totalCash": bal0.get("cashAndCashEquivalents", 0),
        "freeCashflow": cf0.get("freeCashFlow", 0),
        "operatingCashflow": cf0.get("operatingCashFlow", 0),
        "trailingEps": eps,
        "totalRevenue": inc0.get("revenue", 0),
        "sharesOutstanding": shares,
        "enterpriseValue": ev_val,
        "earningsGrowth": gr0.get("epsgrowth", 0),  # FMP: decimal (e.g. 0.15 = 15%)
        # Derived ratios from FMP (override yfinance for source consistency)
        "bookValue": round(book_per_share, 2),
        "enterpriseToEbitda": round(ev_ebitda, 2),
        "trailingPE": round(pe, 2),
        "priceToBook": round(pb, 2),
    })
    return info


def _fmp_to_financials(fmp):
    """Translate FMP financial statements → yfinance-format dicts keyed by year."""
    income = {}
    for row in (fmp.get("income") or []):
        year = row.get("date", "")[:4]
        if not year:
            continue
        income[year] = {
            "Pretax Income": row.get("incomeBeforeTax", 0),
            "Tax Provision": row.get("incomeTaxExpense", 0),
            "Interest Expense": row.get("interestExpense", 0),
        }

    cashflow = {}
    for row in (fmp.get("cashflow") or []):
        year = row.get("date", "")[:4]
        if not year:
            continue
        cashflow[year] = {
            "Operating Cash Flow": row.get("operatingCashFlow", 0),
            "Capital Expenditure": row.get("capitalExpenditure", 0),
        }

    balance = {}
    for row in (fmp.get("balance") or []):
        year = row.get("date", "")[:4]
        if not year:
            continue
        balance[year] = {
            "Total Debt": row.get("totalDebt", 0),
            "Cash And Cash Equivalents": row.get("cashAndCashEquivalents", 0),
            "Stockholders Equity": row.get("totalStockholdersEquity", 0),
        }

    return income, cashflow, balance


# ── Finviz Peer Comparison ──────────────────────────────────────────────

def _finviz_fundamentals(ticker):
    """Fetch fundamentals for a single ticker from Finviz. Returns dict or None."""
    try:
        stock = finvizfinance(ticker)
        f = stock.ticker_fundament()
        pe_raw = f.get("P/E", "-")
        ev_raw = f.get("EV/EBITDA", "-")
        pb_raw = f.get("P/B", "-")
        return {
            "ticker": ticker,
            "name": f.get("Company", ticker),
            "price": _parse_finviz_num(f.get("Price", 0)),
            "mktCap": f.get("Market Cap", "-"),
            "pe": _parse_finviz_num(pe_raw),
            "forwardPE": _parse_finviz_num(f.get("Forward P/E", "-")),
            "evEbitda": _parse_finviz_num(ev_raw),
            "pb": _parse_finviz_num(pb_raw),
            "eps": _parse_finviz_num(f.get("EPS (ttm)", 0)),
            "sector": f.get("Sector", ""),
            "industry": f.get("Industry", ""),
        }
    except Exception as e:
        print(f"[Finviz] Error fetching {ticker}: {e}")
        return None


def _parse_finviz_num(val):
    """Parse Finviz value to float. Returns None for '-' or invalid."""
    if val is None or val == "-" or val == "":
        return None
    try:
        return float(str(val).replace(",", "").replace("%", ""))
    except (ValueError, TypeError):
        return None


def _fetch_peer_comparison(ticker):
    """Fetch peer list and fundamentals from Finviz (free, no API key).

    Returns dict with peer list and computed averages, or None on failure.
    Uses ThreadPoolExecutor for parallel peer fetching (~1-2s total).
    """
    try:
        stock = finvizfinance(ticker)
        peer_tickers = stock.ticker_peer()
        if not peer_tickers:
            return None

        # Limit to 8 peers max, fetch in parallel
        peer_tickers = peer_tickers[:8]
        peers = []
        with ThreadPoolExecutor(max_workers=4) as executor:
            results = list(executor.map(_finviz_fundamentals, peer_tickers))
        peers = [p for p in results if p is not None]

        if not peers:
            return None

        # Compute averages and medians from peers with valid data
        def _avg(key):
            vals = [p[key] for p in peers if p.get(key) is not None and p[key] > 0]
            return round(sum(vals) / len(vals), 2) if vals else None

        def _median(key):
            vals = sorted([p[key] for p in peers if p.get(key) is not None and p[key] > 0])
            if not vals:
                return None
            n = len(vals)
            mid = n // 2
            return round((vals[mid - 1] + vals[mid]) / 2, 2) if n % 2 == 0 else round(vals[mid], 2)

        return {
            "peers": peers,
            "averages": {
                "pe": _avg("pe"),
                "evEbitda": _avg("evEbitda"),
                "pb": _avg("pb"),
            },
            "medians": {
                "pe": _median("pe"),
                "evEbitda": _median("evEbitda"),
                "pb": _median("pb"),
            },
            "source": "Finviz",
        }
    except Exception as e:
        print(f"[Finviz] Peer comparison failed for {ticker}: {e}")
        return None


def _upside_signal(upside):
    """Map upside % to a valuation signal."""
    if upside > 50:
        return "Strong Buy"
    elif upside > 20:
        return "Buy"
    elif upside > -10:
        return "Hold"
    elif upside > -30:
        return "Expensive"
    return "Overrated"


def _trimmean(values, pct=0.2):
    """Trimmed mean — drop top/bottom pct of values."""
    if not values:
        return 0
    s = sorted(values)
    trim = max(1, int(len(s) * pct / 2))
    trimmed = s[trim:-trim] if len(s) > 2 * trim else s
    return sum(trimmed) / len(trimmed) if trimmed else 0


def _compute_wacc(info, income):
    """Compute WACC from CAPM. Returns (wacc_decimal, details_dict) or (None, None)."""
    beta = info.get("beta") or 1.0
    cost_of_equity = RISK_FREE_RATE + beta * (MARKET_RETURN - RISK_FREE_RATE)

    total_debt = info.get("totalDebt") or 0
    total_cash = info.get("totalCash") or 0
    net_debt = max(total_debt - total_cash, 0)
    market_cap = info.get("marketCap") or 0
    total_capital = net_debt + market_cap
    if total_capital <= 0:
        return None, None

    debt_weight = net_debt / total_capital
    equity_weight = market_cap / total_capital

    tax_rate = 0.21
    years_sorted = sorted(income.keys(), reverse=True) if income else []
    for yr in years_sorted:
        pretax = income[yr].get("Pretax Income", 0)
        tax_prov = income[yr].get("Tax Provision", 0)
        if pretax and pretax > 0:
            tax_rate = min(max(tax_prov / pretax, 0), 0.5)
            break

    interest_expense = 0
    for yr in years_sorted:
        ie = abs(income[yr].get("Interest Expense", 0))
        if ie > 0:
            interest_expense = ie
            break
    interest_rate = (interest_expense / total_debt) if total_debt > 0 else 0.04
    cost_of_debt = interest_rate * (1 - tax_rate)

    wacc = equity_weight * cost_of_equity + debt_weight * cost_of_debt
    wacc = max(wacc, 0.05)

    return wacc, {
        "beta": beta, "costOfEquity": cost_of_equity,
        "costOfDebt": cost_of_debt, "taxRate": tax_rate,
        "debtWeight": debt_weight, "equityWeight": equity_weight,
        "totalDebt": total_debt, "netDebt": net_debt,
    }


def _compute_historical_fcf(info, cashflow):
    """Extract historical FCF and avg growth from cashflow dict.
    Returns (historical_fcf_list, hist_avg_growth) or ([], 0).
    """
    historical_fcf = []
    cf_years = sorted(cashflow.keys())
    for yr in cf_years:
        ocf = cashflow[yr].get("Operating Cash Flow", 0)
        capex = cashflow[yr].get("Capital Expenditure", 0)
        fcf = ocf + capex
        historical_fcf.append({"year": yr, "fcf": round(fcf)})

    if not historical_fcf:
        current_fcf = info.get("freeCashflow") or info.get("operatingCashflow", 0)
        if current_fcf:
            historical_fcf = [{"year": "TTM", "fcf": round(current_fcf)}]

    growths = []
    for i in range(1, len(historical_fcf)):
        prev = historical_fcf[i-1]["fcf"]
        curr = historical_fcf[i]["fcf"]
        if prev and prev > 0 and curr > 0:
            g = (curr - prev) / prev
            growths.append(g)
            historical_fcf[i]["growth"] = round(g * 100, 1)

    hist_avg_growth = _trimmean(growths) if growths else 0.07
    return historical_fcf, hist_avg_growth


def compute_dcf(info, income, balance, cashflow):
    """Pure DCF valuation: WACC → single growth rate → Future FCF → discount → IV/share."""
    try:
        wacc, wacc_details = _compute_wacc(info, income)
        if wacc is None:
            return None

        beta = wacc_details["beta"]
        cost_of_equity = wacc_details["costOfEquity"]
        cost_of_debt = wacc_details["costOfDebt"]
        tax_rate = wacc_details["taxRate"]
        debt_weight = wacc_details["debtWeight"]
        equity_weight = wacc_details["equityWeight"]
        total_debt = wacc_details["totalDebt"]

        historical_fcf, hist_avg_growth = _compute_historical_fcf(info, cashflow)
        if not historical_fcf:
            return None

        # Single growth rate for projection (conservative: avg × 0.7)
        growth_rate = hist_avg_growth * 0.7
        growth_rate = max(-0.05, min(growth_rate, 0.30))  # cap

        base_fcf = historical_fcf[-1]["fcf"]
        if base_fcf <= 0:
            base_fcf = info.get("freeCashflow") or 0
        if base_fcf <= 0:
            return None

        price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
        total_cash = info.get("totalCash") or 0
        shares = info.get("sharesOutstanding") or 0
        if shares <= 0:
            return None

        # Project 9 years of future FCF
        projected_fcf = []
        pv_sum = 0
        fcf_val = base_fcf
        for yr in range(1, 10):
            fcf_val = fcf_val * (1 + growth_rate)
            discount = (1 + wacc) ** yr
            pv = fcf_val / discount
            pv_sum += pv
            projected_fcf.append({"year": yr, "fcf": round(fcf_val), "pvFcf": round(pv)})

        # Terminal value (Gordon Growth)
        if wacc <= PERPETUAL_GROWTH:
            return None
        terminal = fcf_val * (1 + PERPETUAL_GROWTH) / (wacc - PERPETUAL_GROWTH)
        pv_terminal = terminal / ((1 + wacc) ** 9)

        enterprise_val = pv_sum + pv_terminal
        equity_val = enterprise_val - total_debt + total_cash
        iv = equity_val / shares
        mos_iv = iv * MARGIN_OF_SAFETY
        upside = ((mos_iv - price) / price * 100) if price > 0 else 0

        return {
            "riskFreeRate": round(RISK_FREE_RATE * 100, 2),
            "marketReturn": round(MARKET_RETURN * 100, 2),
            "beta": round(beta, 2),
            "costOfEquity": round(cost_of_equity * 100, 2),
            "costOfDebt": round(cost_of_debt * 100, 2),
            "wacc": round(wacc * 100, 2),
            "taxRate": round(tax_rate * 100, 1),
            "debtToCapital": round(debt_weight * 100, 1),
            "equityToCapital": round(equity_weight * 100, 1),
            "historicalFcf": historical_fcf,
            "histAvgGrowth": round(hist_avg_growth * 100, 1),
            "growthRate": round(growth_rate * 100, 1),
            "projectedFcf": projected_fcf,
            "terminalValue": round(terminal),
            "pvTerminal": round(pv_terminal),
            "enterpriseValue": round(enterprise_val),
            "equityValue": round(equity_val),
            "ivPerShare": round(iv, 2),
            "marginOfSafetyIv": round(mos_iv, 2),
            "upside": round(upside, 1),
            "signal": _upside_signal(upside),
        }
    except Exception as e:
        print(f"[DCF] Error: {e}")
        return None


def _run_dcf_scenario(fcf_ps, growth1, growth2, terminal_factor, discount_rate):
    """Run one DCF scenario: two-phase growth (yr 1-5 + 6-10) + terminal multiple."""
    year_by_year = []
    current_fcf = fcf_ps
    pv_sum = 0

    for yr in range(1, 11):
        g = growth1 if yr <= 5 else growth2
        current_fcf = current_fcf * (1 + g)
        pv = current_fcf / ((1 + discount_rate) ** yr)
        pv_sum += pv
        year_by_year.append({"year": yr, "fcfPS": round(current_fcf, 4), "pv": round(pv, 4)})

    terminal_value = current_fcf * terminal_factor
    pv_terminal = terminal_value / ((1 + discount_rate) ** 10)
    iv = pv_sum + pv_terminal

    return {
        "yearByYear": year_by_year,
        "terminalValue": round(terminal_value, 2),
        "pvTerminal": round(pv_terminal, 2),
        "ivPerShare": round(iv, 2),
    }


def compute_dcf_scenarios(info, income, balance, cashflow):
    """DCF scenario-based two-phase growth valuation using FCF per share.

    Three weighted scenarios (Base 50%, Best 25%, Worst 25%),
    each with two-phase growth (years 1-5, years 6-10) and terminal multiple.
    """
    try:
        fcf = info.get("freeCashflow") or 0
        shares = info.get("sharesOutstanding") or 0
        if fcf <= 0 or shares <= 0:
            return None
        fcf_ps = fcf / shares
        price = info.get("currentPrice") or info.get("regularMarketPrice", 0)

        # WACC as default discount rate
        wacc, _ = _compute_wacc(info, income)
        if wacc is None:
            wacc = 0.10
        discount_rate = wacc

        # Historical growth for defaults
        _, hist_avg_growth = _compute_historical_fcf(info, cashflow)

        # Default scenario parameters
        # When growth is negative, swap best/worst factors so Best = least negative
        if hist_avg_growth >= 0:
            best_factor, worst_factor = 1.0, 0.3
        else:
            best_factor, worst_factor = 0.3, 1.0

        base_g1 = max(-0.05, min(hist_avg_growth * 0.7, 0.35))
        best_g1 = max(-0.05, min(hist_avg_growth * best_factor, 0.35))
        worst_g1 = max(-0.05, min(hist_avg_growth * worst_factor, 0.35))

        scenario_defs = {
            "base": {
                "growth1": base_g1, "growth2": base_g1 * 0.6,
                "terminalFactor": 15, "probability": 50,
            },
            "best": {
                "growth1": best_g1, "growth2": best_g1 * 0.8,
                "terminalFactor": 20, "probability": 25,
            },
            "worst": {
                "growth1": worst_g1, "growth2": worst_g1 * 0.5,
                "terminalFactor": 10, "probability": 25,
            },
        }

        # Cap growth2 values
        for sd in scenario_defs.values():
            sd["growth2"] = max(-0.05, min(sd["growth2"], 0.25))

        scenarios = {}
        composite_iv = 0
        for name, sd in scenario_defs.items():
            result = _run_dcf_scenario(
                fcf_ps, sd["growth1"], sd["growth2"],
                sd["terminalFactor"], discount_rate
            )
            scenarios[name] = {
                "growth1_5": round(sd["growth1"] * 100, 1),
                "growth6_10": round(sd["growth2"] * 100, 1),
                "terminalFactor": sd["terminalFactor"],
                "probability": sd["probability"],
                **result,
            }
            composite_iv += result["ivPerShare"] * (sd["probability"] / 100)

        mos_iv = composite_iv * MARGIN_OF_SAFETY
        upside = ((mos_iv - price) / price * 100) if price > 0 else 0

        return {
            "fcfPerShare": round(fcf_ps, 2),
            "price": round(price, 2),
            "discountRate": round(discount_rate * 100, 2),
            "wacc": round(wacc * 100, 2),
            "scenarios": scenarios,
            "compositeIv": round(composite_iv, 2),
            "marginOfSafetyIv": round(mos_iv, 2),
            "ivPerShare": round(composite_iv, 2),
            "upside": round(upside, 1),
            "signal": _upside_signal(upside),
        }
    except Exception as e:
        print(f"[DCF Scenarios] Error: {e}")
        return None


def compute_graham(info, aaa_yield_live=None, aaa_date=None):
    """Graham Revised Formula: IV = EPS × (basePE + Cg × g) × Y / C"""
    try:
        eps = info.get("trailingEps") or 0
        if eps <= 0:
            return {"negativeEps": True, "eps": round(eps, 2)}

        # Growth rate: earningsGrowth from FMP (decimal, e.g. 0.15 = 15%), cap to avoid inflated IVs
        raw_g = (info.get("earningsGrowth") or 0) * 100
        g = max(0, min(raw_g, GRAHAM_GROWTH_CAP)) if raw_g > 0 else 5.0

        base_pe = GRAHAM_BASE_PE
        cg = GRAHAM_CG
        aaa_current = aaa_yield_live or AAA_YIELD_CURRENT
        adjusted_multiple = base_pe + cg * g
        bond_adjustment = AAA_YIELD_BASELINE / aaa_current
        iv = eps * adjusted_multiple * bond_adjustment
        if iv <= 0:
            return {"negativeEps": True, "eps": round(eps, 2)}

        mos_iv = iv * MARGIN_OF_SAFETY
        price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
        upside = ((mos_iv - price) / price * 100) if price > 0 else 0

        return {
            "eps": round(eps, 2),
            "growthRate": round(g, 1),
            "basePE": base_pe,
            "cg": cg,
            "adjustedMultiple": round(adjusted_multiple, 1),
            "aaaYieldBaseline": AAA_YIELD_BASELINE,
            "aaaYieldCurrent": round(aaa_current, 2),
            "aaaYieldDate": aaa_date,
            "bondAdjustment": round(bond_adjustment, 4),
            "ivPerShare": round(iv, 2),
            "marginOfSafetyIv": round(mos_iv, 2),
            "upside": round(upside, 1),
            "signal": _upside_signal(upside),
        }
    except Exception as e:
        print(f"[Graham] Error: {e}")
        return None


def compute_relative(info):
    """Relative valuation using sector average multiples.

    All financial inputs (EPS, book value, EV, EBITDA, shares) come from FMP
    via the unified info dict. Sector averages are hardcoded defaults
    (editable in the frontend).
    """
    try:
        sector = info.get("sector", "")
        avgs = SECTOR_AVERAGES.get(sector)
        if not avgs:
            for k, v in SECTOR_AVERAGES.items():
                if k.lower() in sector.lower() or sector.lower() in k.lower():
                    avgs = v
                    break
        if not avgs:
            avgs = {"pe": 20, "evEbitda": 13, "pb": 3}

        eps = info.get("trailingEps") or 0
        book_val = info.get("bookValue") or 0
        price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
        shares = info.get("sharesOutstanding") or 0

        # EV/EBITDA → implied price
        ev = info.get("enterpriseValue") or 0
        ev_ebitda = info.get("enterpriseToEbitda") or 0
        ebitda = (ev / ev_ebitda) if ev_ebitda and ev_ebitda > 0 else 0
        ebitda_per_share = (ebitda / shares) if shares > 0 else 0

        metrics = []
        implied_prices = []

        # P/E implied
        stock_pe = info.get("trailingPE") or 0
        pe_implied = avgs["pe"] * eps if eps > 0 else 0
        metrics.append({
            "name": "P/E", "stockVal": round(stock_pe, 1),
            "sectorAvg": avgs["pe"],
            "impliedPrice": round(pe_implied, 2)
        })
        if pe_implied > 0:
            implied_prices.append(pe_implied)

        # EV/EBITDA implied
        stock_ev_ebitda = ev_ebitda
        ev_implied = avgs["evEbitda"] * ebitda_per_share if ebitda_per_share > 0 else 0
        metrics.append({
            "name": "EV/EBITDA", "stockVal": round(stock_ev_ebitda, 1),
            "sectorAvg": avgs["evEbitda"],
            "impliedPrice": round(ev_implied, 2)
        })
        if ev_implied > 0:
            implied_prices.append(ev_implied)

        # P/B implied
        stock_pb = info.get("priceToBook") or 0
        pb_implied = avgs["pb"] * book_val if book_val > 0 else 0
        metrics.append({
            "name": "P/B", "stockVal": round(stock_pb, 1),
            "sectorAvg": avgs["pb"],
            "impliedPrice": round(pb_implied, 2)
        })
        if pb_implied > 0:
            implied_prices.append(pb_implied)

        if not implied_prices:
            return None

        iv = sum(implied_prices) / len(implied_prices)
        mos_iv = iv * MARGIN_OF_SAFETY
        upside = ((mos_iv - price) / price * 100) if price > 0 else 0

        return {
            "sector": sector,
            "sectorDefaults": dict(avgs),
            "metrics": metrics,
            "eps": round(eps, 2),
            "bookValue": round(book_val, 2),
            "ebitdaPerShare": round(ebitda_per_share, 2),
            "ivPerShare": round(iv, 2),
            "marginOfSafetyIv": round(mos_iv, 2),
            "upside": round(upside, 1),
            "signal": _upside_signal(upside),
        }
    except Exception as e:
        print(f"[Relative] Error: {e}")
        return None


def compute_valuation_summary(dcf, graham, relative, dcf_scenarios, info):
    """Composite weighted IV based on stock category (Growth/Value/Blend)."""
    try:
        pe = info.get("trailingPE") or 0
        rev_growth = (info.get("revenueGrowth") or 0) * 100
        div_yield = info.get("dividendYield") or 0

        # Categorize stock to assign model weights
        if pe > 22 and rev_growth > 12:
            category = "Growth"
            weights = {"dcf": 0.30, "graham": 0.10, "relative": 0.10, "dcfScenarios": 0.50}
        elif (pe > 0 and pe < 24 and div_yield > 1.5) or (pe > 0 and pe < 16):
            category = "Value"
            weights = {"dcf": 0.15, "graham": 0.30, "relative": 0.25, "dcfScenarios": 0.30}
        else:
            category = "Blend"
            weights = {"dcf": 0.25, "graham": 0.20, "relative": 0.20, "dcfScenarios": 0.35}

        # Collect valid IVs
        models = {}
        if dcf and dcf.get("ivPerShare", 0) > 0:
            models["dcf"] = dcf["ivPerShare"]
        if graham and graham.get("ivPerShare", 0) > 0:
            models["graham"] = graham["ivPerShare"]
        if relative and relative.get("ivPerShare", 0) > 0:
            models["relative"] = relative["ivPerShare"]
        if dcf_scenarios and dcf_scenarios.get("ivPerShare", 0) > 0:
            models["dcfScenarios"] = dcf_scenarios["ivPerShare"]

        if not models:
            return None

        # Normalize weights for available models
        total_w = sum(weights[k] for k in models)
        if total_w <= 0:
            return None
        composite = sum(models[k] * weights[k] / total_w for k in models)
        mos_iv = composite * MARGIN_OF_SAFETY

        price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
        upside = ((mos_iv - price) / price * 100) if price > 0 else 0

        return {
            "category": category,
            "weights": weights,
            "models": {k: round(v, 2) for k, v in models.items()},
            "compositeIv": round(composite, 2),
            "marginOfSafetyIv": round(mos_iv, 2),
            "upside": round(upside, 1),
            "signal": _upside_signal(upside),
        }
    except Exception as e:
        print(f"[ValuationSummary] Error: {e}")
        return None


# ── Stock Analyzer ──────────────────────────────────────────────────────
ANALYZER_FILE = DATA_DIR / "analyzer.json"


def _load_analyzer_store():
    """Load persisted analyzer results from disk."""
    try:
        if ANALYZER_FILE.exists():
            return json.loads(ANALYZER_FILE.read_text())
    except Exception as e:
        print(f"[Analyzer] Failed to load {ANALYZER_FILE}: {e}")
    return {}


def _save_analyzer_store(store):
    """Persist analyzer results to disk."""
    try:
        ANALYZER_FILE.write_text(json.dumps(store, indent=2, default=str))
    except Exception as e:
        print(f"[Analyzer] Failed to save {ANALYZER_FILE}: {e}")


@app.route("/api/stock-analyzer/<ticker>")
def api_stock_analyzer(ticker):
    """Deep analysis: FMP for financials/ratios, yfinance for supplementary fields.

    Returns saved data from analyzer.json by default.
    Pass ?refresh=true to fetch fresh data from APIs and save.
    """
    ticker = ticker.upper().strip()
    refresh = request.args.get("refresh", "").lower() in ("true", "1", "yes")

    # Return saved data if not refreshing
    if not refresh:
        store = _load_analyzer_store()
        if ticker in store:
            return jsonify(store[ticker])

    try:
        # yfinance: profile, ratios, supplementary fields
        try:
            yf_info = yf.Ticker(ticker).info or {}
        except Exception:
            yf_info = {}
        if not yf_info.get("currentPrice") and not yf_info.get("regularMarketPrice"):
            return jsonify({"error": f"Ticker '{ticker}' not found"}), 404

        # Primary: SEC EDGAR (1 call, 10yr history, no daily limit)
        edgar_facts = _fetch_edgar_facts(ticker)
        data_source = None
        if edgar_facts:
            info = _edgar_to_info(edgar_facts, yf_info)
            income, cashflow, balance = _edgar_to_financials(edgar_facts)
            if income or cashflow:
                data_source = "SEC EDGAR"
                print(f"[Analyzer] {ticker}: SEC EDGAR ({len(income)} yr income, {len(cashflow)} yr cashflow)")
        if not data_source:
            # Fallback 1: FMP (5 API calls)
            print(f"[Analyzer] {ticker}: EDGAR unavailable, trying FMP")
            fmp = _fetch_fmp_stock_data(ticker)
            info = _fmp_to_info(fmp, yf_info)
            income, cashflow, balance = _fmp_to_financials(fmp)
            data_source = "FMP"
        if data_source == "FMP" and not income and not cashflow:
            # Fallback 2: yfinance (foreign ADRs not covered by EDGAR or FMP)
            print(f"[Analyzer] {ticker}: FMP empty, falling back to yfinance")
            info = dict(yf_info)
            t = yf.Ticker(ticker)
            income, cashflow, balance = {}, {}, {}
            try:
                cf = t.cashflow
                if cf is not None and not cf.empty:
                    for col in cf.columns:
                        yr = str(col.year) if hasattr(col, 'year') else str(col)[:4]
                        ocf = cf.at["Operating Cash Flow", col] if "Operating Cash Flow" in cf.index else 0
                        capex = cf.at["Capital Expenditure", col] if "Capital Expenditure" in cf.index else 0
                        cashflow[yr] = {
                            "Operating Cash Flow": int(ocf) if ocf == ocf else 0,
                            "Capital Expenditure": int(capex) if capex == capex else 0,
                        }
            except Exception as e:
                print(f"[Analyzer] yfinance cashflow error: {e}")
            try:
                inc = t.income_stmt
                if inc is not None and not inc.empty:
                    for col in inc.columns:
                        yr = str(col.year) if hasattr(col, 'year') else str(col)[:4]
                        pretax = inc.at["Pretax Income", col] if "Pretax Income" in inc.index else 0
                        tax = inc.at["Tax Provision", col] if "Tax Provision" in inc.index else 0
                        interest = inc.at["Interest Expense", col] if "Interest Expense" in inc.index else 0
                        income[yr] = {
                            "Pretax Income": int(pretax) if pretax == pretax else 0,
                            "Tax Provision": int(tax) if tax == tax else 0,
                            "Interest Expense": int(interest) if interest == interest else 0,
                        }
            except Exception as e:
                print(f"[Analyzer] yfinance income error: {e}")
            data_source = "Yahoo Finance"

        price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
        result = {
            "ticker": ticker,
            "name": info.get("longName", ticker),
            "sector": info.get("sector", ""),
            "industry": info.get("industry", ""),
            "price": price,
            "marketCap": info.get("marketCap", 0),
            "enterpriseValue": info.get("enterpriseValue", 0),
            "trailingPE": info.get("trailingPE", 0),
            "forwardPE": info.get("forwardPE", 0),
            "pegRatio": info.get("pegRatio", 0),
            "priceToBook": info.get("priceToBook", 0),
            "priceToSales": info.get("priceToSalesTrailing12Months", 0),
            "evToEbitda": info.get("enterpriseToEbitda", 0),
            "evToRevenue": info.get("enterpriseToRevenue", 0),
            "profitMargin": round((info.get("profitMargins") or 0) * 100, 2),
            "operatingMargin": round((info.get("operatingMargins") or 0) * 100, 2),
            "grossMargin": round((info.get("grossMargins") or 0) * 100, 2),
            "returnOnEquity": round((info.get("returnOnEquity") or 0) * 100, 2),
            "returnOnAssets": round((info.get("returnOnAssets") or 0) * 100, 2),
            "debtToEquity": info.get("debtToEquity", 0),
            "currentRatio": info.get("currentRatio", 0),
            "quickRatio": info.get("quickRatio", 0),
            "beta": info.get("beta", 0),
            "dividendYield": round(info.get("dividendYield") or 0, 2),
            "dividendRate": info.get("dividendRate", 0),
            "payoutRatio": round((info.get("payoutRatio") or 0) * 100, 2),
            "fiveYearAvgDivYield": info.get("fiveYearAvgDividendYield", 0),
            "revenueGrowth": round((info.get("revenueGrowth") or 0) * 100, 2),
            "earningsGrowth": round((info.get("earningsGrowth") or 0) * 100, 2),
            "targetMeanPrice": info.get("targetMeanPrice", 0),
            "targetHighPrice": info.get("targetHighPrice", 0),
            "targetLowPrice": info.get("targetLowPrice", 0),
            "recommendationKey": info.get("recommendationKey", ""),
            "numberOfAnalysts": info.get("numberOfAnalystOpinions", 0),
            "fiftyTwoWeekHigh": info.get("fiftyTwoWeekHigh", 0),
            "fiftyTwoWeekLow": info.get("fiftyTwoWeekLow", 0),
            "fiftyDayAvg": info.get("fiftyDayAverage", 0),
            "twoHundredDayAvg": info.get("twoHundredDayAverage", 0),
            "sharesOutstanding": info.get("sharesOutstanding", 0),
            "floatShares": info.get("floatShares", 0),
            "shortRatio": info.get("shortRatio", 0),
            "bookValue": info.get("bookValue", 0),
            "earningsPerShare": info.get("trailingEps", 0),
            "forwardEps": info.get("forwardEps", 0),
            "revenuePerShare": info.get("revenuePerShare", 0),
            "totalRevenue": info.get("totalRevenue", 0),
            "totalDebt": info.get("totalDebt", 0),
            "totalCash": info.get("totalCash", 0),
            "freeCashflow": info.get("freeCashflow", 0),
            "operatingCashflow": info.get("operatingCashflow", 0),
            "income": income,
            "balance": balance,
            "cashflow": cashflow,
            "analystConsensus": {
                "recommendation": info.get("recommendationKey", ""),
                "targetMean": info.get("targetMeanPrice", 0),
                "targetHigh": info.get("targetHighPrice", 0),
                "targetLow": info.get("targetLowPrice", 0),
                "numberOfAnalysts": info.get("numberOfAnalystOpinions", 0),
                "source": "Yahoo Finance",
            },
            "dataSources": {
                "financials": {
                    "SEC EDGAR": "SEC EDGAR (10-K XBRL filings)",
                    "FMP": "FMP (financial statements API)",
                    "Yahoo Finance": "Yahoo Finance (financial statements)",
                }.get(data_source, data_source),
                "profile": "Yahoo Finance (price, beta, analyst targets)",
                "bonds": "FRED (AAA corporate bond yield)",
                "ratios": f"{data_source}-derived (P/E, P/B, EV/EBITDA, book value/share)",
                "peers": "Finviz (peer companies, sector multiples)",
            },
            "lastUpdated": datetime.now().isoformat(),
        }

        # Fetch live AAA yield from FRED for Graham model
        aaa_yield_live, aaa_date = _fetch_fred_aaa_yield()

        # Fetch Finviz peers and FMP benchmarks in background threads
        peer_result = [None]
        fmp_dcf_result = [None]
        fmp_bench_result = [{}]
        def _bg_peers():
            peer_result[0] = _fetch_peer_comparison(ticker)
        def _bg_fmp_dcf():
            fmp_dcf_result[0] = _fetch_fmp_dcf(ticker)
        def _bg_fmp_bench():
            fmp_bench_result[0] = _fetch_fmp_benchmarks(ticker)
        peer_thread = threading.Thread(target=_bg_peers)
        fmp_dcf_thread = threading.Thread(target=_bg_fmp_dcf)
        fmp_bench_thread = threading.Thread(target=_bg_fmp_bench)
        peer_thread.start()
        fmp_dcf_thread.start()
        fmp_bench_thread.start()

        # Valuation models (run while peers fetch in background)
        dcf = compute_dcf(info, income, balance, cashflow)
        graham = compute_graham(info, aaa_yield_live=aaa_yield_live, aaa_date=aaa_date)
        relative = compute_relative(info)
        dcf_scenarios = compute_dcf_scenarios(info, income, balance, cashflow)
        summary = compute_valuation_summary(dcf, graham, relative, dcf_scenarios, info)

        # Wait for peers and FMP benchmarks (max 10s)
        peer_thread.join(timeout=10)
        fmp_dcf_thread.join(timeout=5)
        fmp_bench_thread.join(timeout=8)
        if relative and peer_result[0]:
            relative["peerComparison"] = peer_result[0]

        result["valuation"] = {
            "dcf": dcf,
            "graham": graham,
            "relative": relative,
            "dcfScenarios": dcf_scenarios,
            "summary": summary,
        }
        fmp_bench = fmp_bench_result[0] or {}
        result["benchmarks"] = {
            "fmpDcf": fmp_dcf_result[0],
            "fmpGrahamNumber": fmp_bench.get("grahamNumber", 0),
            "fmpRating": fmp_bench.get("rating", ""),
            "fmpRatingScore": fmp_bench.get("ratingScore", 0),
            "fmpRatingDcfScore": fmp_bench.get("ratingDcfScore", 0),
            "fmpRatingPeScore": fmp_bench.get("ratingPeScore", 0),
            "fmpRatingPbScore": fmp_bench.get("ratingPbScore", 0),
            "fmpAltmanZ": fmp_bench.get("altmanZScore", 0),
            "fmpPiotroski": fmp_bench.get("piotroskiScore", 0),
            "fmpEarningsYield": fmp_bench.get("earningsYield", 0),
            "fmpFcfYield": fmp_bench.get("freeCashFlowYield", 0),
            "fmpRoic": fmp_bench.get("roic", 0),
            "analystMean": info.get("targetMeanPrice", 0),
            "analystHigh": info.get("targetHighPrice", 0),
            "analystLow": info.get("targetLowPrice", 0),
            "analystCount": info.get("numberOfAnalystOpinions", 0),
        }

        # Persist to file and memory cache
        store = _load_analyzer_store()
        store[ticker] = result
        _save_analyzer_store(store)
        cache_set(f"analyzer_{ticker}", result)
        return jsonify(result)

    except Exception as e:
        print(f"[Analyzer] Error for {ticker}: {e}")
        return jsonify({"error": str(e)}), 500


# ── InvT Score — Fundamental Quality Scoring ───────────────────────────

INVT_THRESHOLDS = {
    # Growth (CAGR %): higher is better
    "revenue_cagr":   [(0, 0), (1, 1), (3, 3), (6, 5), (9, 7), (14, 9)],
    "eps_cagr":       [(-10, 0), (0, 1), (3, 3), (6, 5), (9, 7), (13, 9)],
    "fcf_share_cagr": [(-10, 0), (0, 1), (3, 3), (6, 5), (9, 7), (18, 9)],
    # Profitability (%): higher is better
    "gpm":            [(0, 0), (20, 1), (30, 3), (40, 5), (50, 7), (60, 9)],
    "npm":            [(0, 0), (3, 1), (6, 3), (10, 5), (15, 7), (20, 9)],
    "fcf_margin":     [(0, 0), (5, 1), (8, 3), (12, 5), (18, 7), (25, 9)],
    # Debt: lower/negative is better (inverted)
    "net_debt_cagr":  [(-25, 10), (-15, 9), (0, 7), (5, 5), (10, 3), (15, 1)],
    "net_debt_fcf":   [(0, 10), (1, 9), (2, 7), (3, 5), (4, 3), (5, 1)],
    # Debt: higher is better
    "interest_cov":   [(5, 0), (7, 1), (9, 3), (11, 5), (13, 7), (16, 9)],
    # Dividends: DPS growth — higher is better
    "dps_cagr":       [(-5, 0), (0, 1), (4, 3), (8, 5), (12, 7), (16, 9)],
    # Dividends: FCF Payout — lower is better (inverted)
    "fcf_payout":     [(30, 10), (40, 9), (50, 7), (60, 5), (70, 3), (80, 1)],
    # Buybacks: share count declining is better (inverted)
    "shares_cagr":    [(-2.5, 10), (-2, 9), (-1.5, 7), (-1, 5), (-0.5, 3), (0.5, 1), (2, 0)],
    # Capital Efficiency (%): higher is better
    "roa":            [(0, 0), (2, 1), (4, 3), (6, 5), (8, 7), (10, 9)],
    "roe":            [(0, 0), (5, 1), (10, 3), (15, 5), (20, 7), (25, 9)],
    "roic":           [(0, 0), (5, 1), (8, 3), (10, 5), (12, 7), (15, 9)],
}

# Categories that contribute to the overall InvT Score
INVT_CATEGORIES_SCORED = {
    "growth":        {"metrics": ["revenue_cagr", "eps_cagr", "fcf_share_cagr"],
                      "weights": [1/3, 1/3, 1/3], "label": "Growth"},
    "profitability": {"metrics": ["gpm", "npm", "fcf_margin"],
                      "weights": [1/3, 1/3, 1/3], "label": "Profitability"},
    "debt":          {"metrics": ["net_debt_cagr", "net_debt_fcf", "interest_cov"],
                      "weights": [1/3, 1/3, 1/3], "label": "Debt"},
    "efficiency":    {"metrics": ["roa", "roe", "roic"],
                      "weights": [1/3, 1/3, 1/3], "label": "Capital Efficiency"},
}

# Informational category — always computed, never in overall score
INVT_CATEGORIES_INFO = {
    "shareholder_returns": {"metrics": ["div_yield", "dps_cagr", "payout_ratio", "fcf_payout", "shares_cagr"],
                            "weights": [0.15, 0.25, 0.15, 0.25, 0.20], "label": "Dividend & Buyback"},
}

# Combined for iteration (display all 5)
INVT_CATEGORIES = {**INVT_CATEGORIES_SCORED, **INVT_CATEGORIES_INFO}

INVT_METRIC_NAMES = {
    "revenue_cagr": "Revenue CAGR", "eps_cagr": "EPS CAGR", "fcf_share_cagr": "FCF/Share CAGR",
    "gpm": "Gross Profit Margin", "npm": "Net Profit Margin", "fcf_margin": "FCF Margin",
    "net_debt_cagr": "Net Debt CAGR", "net_debt_fcf": "Net Debt / FCF", "interest_cov": "Interest Coverage",
    "div_yield": "Dividend Yield", "dps_cagr": "Div/Share CAGR", "payout_ratio": "Payout Ratio",
    "fcf_payout": "FCF Payout Ratio", "shares_cagr": "Shares Outstanding CAGR",
    "roa": "Return on Assets", "roe": "Return on Equity", "roic": "Return on Invested Capital",
}

INVT_METRIC_UNITS = {
    "revenue_cagr": "%", "eps_cagr": "%", "fcf_share_cagr": "%",
    "gpm": "%", "npm": "%", "fcf_margin": "%",
    "net_debt_cagr": "%", "net_debt_fcf": "x", "interest_cov": "x",
    "div_yield": "%", "dps_cagr": "%", "payout_ratio": "%",
    "fcf_payout": "%", "shares_cagr": "%",
    "roa": "%", "roe": "%", "roic": "%",
}


def _invt_cagr(start_val, end_val, years):
    """Compound Annual Growth Rate. Returns % (e.g. 12.5 for 12.5%)."""
    if years <= 0 or not start_val or not end_val:
        return None
    if start_val < 0 and end_val < 0:
        # Both negative: measure improvement (e.g. net debt shrinking)
        ratio = end_val / start_val
    elif start_val <= 0:
        return None  # Can't compute CAGR across zero crossing
    else:
        ratio = end_val / start_val
    if ratio <= 0:
        return None
    return (ratio ** (1 / years) - 1) * 100


def _invt_safe_avg(values):
    """Average of non-None values."""
    valid = [v for v in values if v is not None]
    return round(sum(valid) / len(valid), 2) if valid else None


_INVT_INVERTED = {"net_debt_cagr", "net_debt_fcf", "fcf_payout", "shares_cagr"}


def _invt_score_metric(value, key):
    """Score a metric 0-10 using its threshold table. Custom logic for div_yield and payout_ratio."""
    if value is None:
        return None
    # Custom scorers for non-monotonic metrics
    if key == "div_yield":
        if value <= 0: return 0
        if value < 1: return 3    # Low but paying
        if value < 2: return 5    # Moderate
        if value < 4: return 7    # Sweet spot
        if value < 6: return 5    # Getting high
        if value < 8: return 3    # Yield trap risk
        return 1                  # Distressed
    if key == "payout_ratio":
        if value >= 120: return 0
        if value >= 100: return 1
        if value >= 80: return 3
        if value >= 60: return 5
        if value >= 40: return 7
        if value >= 20: return 9   # Sweet spot
        if value >= 10: return 7
        return 5
    # Generic threshold scoring
    thresholds = INVT_THRESHOLDS.get(key)
    if not thresholds:
        return None
    for upper, score in thresholds:
        if value < upper:
            return score
    # Fall-through: inverted metrics get 0 (worst), normal metrics get max+1
    if key in _INVT_INVERTED:
        return 0
    return thresholds[-1][1] + 1  # Above all thresholds → max score (10)


def _invt_label(score):
    """Map overall score to classification label."""
    if score is None:
        return "Insufficient Data"
    if score >= 9: return "Elite \U0001f680"
    if score >= 8: return "High Quality \u2705"
    if score >= 6: return "Above Average \U0001f44d"
    if score >= 4: return "Below Average \U0001f4c9"
    return "Poor Quality \U0001f6a8"


def _fetch_invt_data(ticker):
    """Fetch 5+ years of financial data for InvT Score.
    Returns (yearly_data_list, data_source) or (None, None).
    Uses SEC EDGAR → FMP → yfinance cascade."""

    def _build_yearly(years, rev, gp, ni, ebit, eps, ocf, capex,
                      debt, cash, equity, assets, interest, pretax, tax,
                      divs_paid, shares):
        """Build sorted list of yearly data dicts from year-keyed dicts."""
        all_years = sorted(set(years) | set(rev.keys()))
        result = []
        for y in all_years:
            r = rev.get(y, 0)
            o = ocf.get(y, 0)
            c = capex.get(y, 0)
            fcf = o - abs(c)
            result.append({
                "year": y, "revenue": r, "grossProfit": gp.get(y, 0),
                "netIncome": ni.get(y, 0), "ebit": ebit.get(y, 0),
                "eps": eps.get(y, 0), "ocf": o, "capex": c, "fcf": fcf,
                "totalDebt": debt.get(y, 0), "cash": cash.get(y, 0),
                "equity": equity.get(y, 0), "totalAssets": assets.get(y, 0),
                "interestExpense": interest.get(y, 0),
                "pretaxIncome": pretax.get(y, 0), "taxProvision": tax.get(y, 0),
                "dividendsPaid": abs(divs_paid.get(y, 0)),  # Always positive
                "sharesOutstanding": shares.get(y, 0),
            })
        return result

    # Primary: SEC EDGAR
    facts = _fetch_edgar_facts(ticker)
    if facts:
        rev = _edgar_merge_tags(facts, [
            "RevenueFromContractWithCustomerExcludingAssessedTax",
            "Revenues", "SalesRevenueNet", "RealEstateRevenueNet",
        ], max_years=11)
        if rev:
            gp = _edgar_annual_values(facts, "GrossProfit", max_years=11)
            if not gp:
                cor = _edgar_merge_tags(facts, [
                    "CostOfGoodsAndServicesSold", "CostOfGoodsSold", "CostOfRevenue",
                ], max_years=11)
                if cor:
                    gp = {y: rev.get(y, 0) - cor.get(y, 0) for y in cor if y in rev}
            ni = _edgar_annual_values(facts, "NetIncomeLoss", max_years=11)
            ebit = _edgar_annual_values(facts, "OperatingIncomeLoss", max_years=11)
            eps = _edgar_annual_values(facts, "EarningsPerShareDiluted", unit="USD/shares", max_years=11)
            ocf = _edgar_annual_values(facts, "NetCashProvidedByUsedInOperatingActivities", max_years=11)
            capex = _edgar_annual_values(facts, "PaymentsToAcquirePropertyPlantAndEquipment", max_years=11)
            debt_nc = _edgar_annual_values(facts, "LongTermDebtNoncurrent", max_years=11)
            debt_c = _edgar_annual_values(facts, "LongTermDebtCurrent", max_years=11)
            debt_fb = _edgar_annual_values(facts, "LongTermDebt", max_years=11)
            debt = {}
            for y in set(debt_nc) | set(debt_c) | set(debt_fb):
                d = debt_nc.get(y, 0) + debt_c.get(y, 0)
                debt[y] = d if d else debt_fb.get(y, 0)
            cash = _edgar_merge_tags(facts, [
                "CashAndCashEquivalentsAtCarryingValue",
                "CashCashEquivalentsAndShortTermInvestments",
                "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents",
            ], max_years=11)
            equity = _edgar_annual_values(facts, "StockholdersEquity", max_years=11)
            assets = _edgar_annual_values(facts, "Assets", max_years=11)
            interest = _edgar_merge_tags(facts, [
                "InterestExpense", "InterestExpenseDebt",
                "InterestPaidNet",  # Cash flow fallback
                "InterestIncomeExpenseNonoperatingNet",
            ], max_years=11)
            pretax_tags = [
                "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
                "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments",
            ]
            pretax = _edgar_merge_tags(facts, pretax_tags, max_years=11)
            tax = _edgar_annual_values(facts, "IncomeTaxExpenseBenefit", max_years=11)
            divs = _edgar_merge_tags(facts, [
                "PaymentsOfDividends", "PaymentsOfDividendsCommonStock",
                "PaymentsOfOrdinaryDividends",
            ], max_years=11)
            shares = _edgar_merge_tags(facts, [
                "CommonStockSharesOutstanding",
                "WeightedAverageNumberOfShareOutstandingBasicAndDiluted",
                "WeightedAverageNumberOfDilutedSharesOutstanding",
            ], unit="shares", max_years=11)
            if not shares:
                shares = _edgar_annual_values(facts, "EntityCommonStockSharesOutstanding",
                                              unit="shares", ns="dei", max_years=11)
            yearly = _build_yearly(rev.keys(), rev, gp, ni, ebit, eps, ocf, capex,
                                   debt, cash, equity, assets, interest, pretax, tax,
                                   divs, shares)
            # Normalize stock splits: detect YoY share count jumps > 3x
            for i in range(1, len(yearly)):
                prev_s = yearly[i - 1].get("sharesOutstanding", 0)
                curr_s = yearly[i].get("sharesOutstanding", 0)
                if prev_s and curr_s:
                    ratio = curr_s / prev_s
                    if ratio > 3:  # Forward split detected
                        for j in range(i):
                            yearly[j]["sharesOutstanding"] *= ratio
                            if yearly[j].get("eps"):
                                yearly[j]["eps"] /= ratio
            if len(yearly) >= 2:
                print(f"[InvTScore] {ticker}: SEC EDGAR, {len(yearly)} years")
                return yearly, "SEC EDGAR"

    # Fallback: FMP
    try:
        fmp = _fetch_fmp_stock_data(ticker)
        inc_rows = fmp.get("income") or []
        cf_rows = fmp.get("cashflow") or []
        bal_rows = fmp.get("balance") or []
        ev_rows = fmp.get("ev") or []
        if inc_rows:
            rev, gp, ni, ebit, eps_d = {}, {}, {}, {}, {}
            interest, pretax, tax = {}, {}, {}
            for row in inc_rows[:11]:
                y = row.get("date", "")[:4]
                if not y: continue
                rev[y] = row.get("revenue", 0)
                gp[y] = row.get("grossProfit", 0)
                ni[y] = row.get("netIncome", 0)
                ebit[y] = row.get("operatingIncome", 0)
                eps_d[y] = row.get("epsDiluted", 0)
                interest[y] = row.get("interestExpense", 0)
                pretax[y] = row.get("incomeBeforeTax", 0)
                tax[y] = row.get("incomeTaxExpense", 0)
            ocf, capex_d, divs = {}, {}, {}
            for row in cf_rows[:11]:
                y = row.get("date", "")[:4]
                if not y: continue
                ocf[y] = row.get("operatingCashFlow", 0)
                capex_d[y] = row.get("capitalExpenditure", 0)
                divs[y] = row.get("dividendsPaid", 0)
            debt, cash_d, equity, assets = {}, {}, {}, {}
            for row in bal_rows[:11]:
                y = row.get("date", "")[:4]
                if not y: continue
                debt[y] = row.get("totalDebt", 0)
                cash_d[y] = row.get("cashAndCashEquivalents", 0)
                equity[y] = row.get("totalStockholdersEquity", 0)
                assets[y] = row.get("totalAssets", 0)
            shares = {}
            for row in ev_rows[:11]:
                y = row.get("date", "")[:4]
                if not y: continue
                shares[y] = row.get("numberOfShares", 0)
            yearly = _build_yearly(rev.keys(), rev, gp, ni, ebit, eps_d, ocf, capex_d,
                                   debt, cash_d, equity, assets, interest, pretax, tax,
                                   divs, shares)
            if len(yearly) >= 2:
                print(f"[InvTScore] {ticker}: FMP, {len(yearly)} years")
                return yearly, "FMP"
    except Exception as e:
        print(f"[InvTScore] FMP error for {ticker}: {e}")

    return None, None


def _compute_invt_metrics(yearly, mode="5yr"):
    """Compute all 16 InvT metric values from yearly data.
    mode='10yr': all available data (up to 10 years).
    mode='5yr': last 5 data points (4 CAGR periods)."""
    if not yearly or len(yearly) < 2:
        return {}
    if mode == "5yr":
        data = yearly[-5:] if len(yearly) >= 5 else yearly
    else:
        data = yearly  # Full range (10yr)

    first, last = data[0], data[-1]
    n = len(data) - 1  # Number of growth periods

    metrics = {}

    # ── Growth CAGRs ──
    metrics["revenue_cagr"] = _invt_cagr(first["revenue"], last["revenue"], n)
    metrics["eps_cagr"] = _invt_cagr(first["eps"], last["eps"], n) if first["eps"] and last["eps"] else None
    # FCF per share CAGR
    fcf_ps_first = first["fcf"] / first["sharesOutstanding"] if first.get("sharesOutstanding") else None
    fcf_ps_last = last["fcf"] / last["sharesOutstanding"] if last.get("sharesOutstanding") else None
    metrics["fcf_share_cagr"] = _invt_cagr(fcf_ps_first, fcf_ps_last, n) if fcf_ps_first and fcf_ps_last else None

    # ── Profitability Averages ──
    gpms = [d["grossProfit"] / d["revenue"] * 100 for d in data if d.get("revenue") and d.get("grossProfit")]
    metrics["gpm"] = _invt_safe_avg(gpms) if gpms else None
    npms = [d["netIncome"] / d["revenue"] * 100 for d in data if d.get("revenue") and d.get("netIncome")]
    metrics["npm"] = _invt_safe_avg(npms) if npms else None
    fcf_margins = [d["fcf"] / d["revenue"] * 100 for d in data if d.get("revenue") and d.get("fcf")]
    metrics["fcf_margin"] = _invt_safe_avg(fcf_margins) if fcf_margins else None

    # ── Debt ──
    net_debts = [(d["totalDebt"] - d["cash"]) for d in data if d.get("totalDebt") is not None]
    if len(net_debts) >= 2 and net_debts[0]:
        if net_debts[0] > 0 and net_debts[-1] <= 0:
            metrics["net_debt_cagr"] = -100  # Went from debt to net cash → best outcome
        elif net_debts[-1]:
            metrics["net_debt_cagr"] = _invt_cagr(net_debts[0], net_debts[-1], n)
        else:
            metrics["net_debt_cagr"] = None
    else:
        metrics["net_debt_cagr"] = None
    nd_fcf = []
    for d in data:
        nd = d.get("totalDebt", 0) - d.get("cash", 0)
        if d.get("fcf") and d["fcf"] > 0:
            nd_fcf.append(nd / d["fcf"])
    metrics["net_debt_fcf"] = _invt_safe_avg(nd_fcf) if nd_fcf else None
    int_covs = [d["ebit"] / d["interestExpense"]
                for d in data if d.get("ebit") and d.get("interestExpense") and d["interestExpense"] > 0]
    metrics["interest_cov"] = _invt_safe_avg(int_covs) if int_covs else None

    # ── Dividends & Buybacks ──
    div_yields = []
    for d in data:
        if d.get("dividendsPaid") and d.get("sharesOutstanding") and d.get("eps"):
            dps = d["dividendsPaid"] / d["sharesOutstanding"]
            # Approximate price from EPS × P/E ≈ use netIncome/equity as rough proxy
            # Better: use dividendsPaid / netIncome × 100 for yield-like metric
            # Actually compute DPS/EPS as a proxy since we lack historical prices
            # Use dividend yield = DPS / (EPS / earnings_yield) but we don't have yield
            # Simplest: if we have revenue per share, approximate. Let's use DPS directly.
            pass
    # For dividend yield, we compute from total dividends paid / market cap
    # Since we don't have historical market cap, use DPS as % of EPS as proxy
    dps_vals = []
    for d in data:
        if d.get("dividendsPaid") and d.get("sharesOutstanding"):
            dps_vals.append(d["dividendsPaid"] / d["sharesOutstanding"])
        else:
            dps_vals.append(0)
    # Use current yfinance dividend yield if available, otherwise estimate
    # For now, compute average DPS and we'll get yield from yfinance in the route
    metrics["_avg_dps"] = _invt_safe_avg(dps_vals) if dps_vals else 0
    metrics["div_yield"] = None  # Will be filled from yfinance in route

    if len(dps_vals) >= 2 and dps_vals[0] and dps_vals[-1]:
        metrics["dps_cagr"] = _invt_cagr(dps_vals[0], dps_vals[-1], n)
    else:
        metrics["dps_cagr"] = None if any(d.get("dividendsPaid", 0) > 0 for d in data) else 0

    payout_ratios = [d["dividendsPaid"] / d["netIncome"] * 100
                     for d in data if d.get("dividendsPaid") and d.get("netIncome") and d["netIncome"] > 0]
    metrics["payout_ratio"] = _invt_safe_avg(payout_ratios) if payout_ratios else 0

    fcf_payouts = [d["dividendsPaid"] / d["fcf"] * 100
                   for d in data if d.get("dividendsPaid") and d.get("fcf") and d["fcf"] > 0]
    metrics["fcf_payout"] = _invt_safe_avg(fcf_payouts) if fcf_payouts else 0

    shares_first = first.get("sharesOutstanding", 0)
    shares_last = last.get("sharesOutstanding", 0)
    metrics["shares_cagr"] = _invt_cagr(shares_first, shares_last, n) if shares_first and shares_last else None

    # ── Capital Efficiency Averages ──
    roas = [d["netIncome"] / d["totalAssets"] * 100
            for d in data if d.get("netIncome") and d.get("totalAssets") and d["totalAssets"] > 0]
    metrics["roa"] = _invt_safe_avg(roas) if roas else None

    roes = [d["netIncome"] / d["equity"] * 100
            for d in data if d.get("netIncome") and d.get("equity") and d["equity"] > 0]
    metrics["roe"] = _invt_safe_avg(roes) if roes else None

    roics = []
    for d in data:
        if d.get("ebit") and d.get("equity"):
            tax_rate = d["taxProvision"] / d["pretaxIncome"] if d.get("pretaxIncome") and d["pretaxIncome"] > 0 else 0.21
            invested = d.get("totalDebt", 0) + d["equity"] - d.get("cash", 0)
            if invested > 0:
                nopat = d["ebit"] * (1 - tax_rate)
                roics.append(nopat / invested * 100)
    metrics["roic"] = _invt_safe_avg(roics) if roics else None

    return metrics


def _compute_invt_category_scores(metric_scores, categories=None):
    """Compute category scores from individual metric scores.
    Requires ≥2 valid metrics for categories with 3+ metrics to prevent single-metric distortion."""
    cats = categories or INVT_CATEGORIES
    result = {}
    for cat_key, cat_def in cats.items():
        min_required = 2 if len(cat_def["metrics"]) >= 3 else 1
        valid_count = sum(1 for m in cat_def["metrics"] if metric_scores.get(m) is not None)
        if valid_count < min_required:
            result[cat_key] = None
            continue
        weighted_sum = 0
        weight_sum = 0
        for i, m_key in enumerate(cat_def["metrics"]):
            s = metric_scores.get(m_key)
            if s is not None:
                weighted_sum += s * cat_def["weights"][i]
                weight_sum += cat_def["weights"][i]
        result[cat_key] = round(weighted_sum / weight_sum, 1) if weight_sum > 0 else None
    return result


@app.route("/api/invt-score/<ticker>")
def api_invt_score(ticker):
    """InvT Score: 0-10 company quality score across 5 categories.
    Uses 10yr/5yr historical data with 70/30 hybrid weighting.
    Pass ?refresh=true to re-fetch from APIs."""
    ticker = ticker.upper().strip()
    refresh = request.args.get("refresh", "").lower() in ("true", "1", "yes")

    # Check cache
    if not refresh:
        store = _load_analyzer_store()
        cached = store.get(ticker, {}).get("invtScore")
        if cached and cached.get("version") == 3:
            return jsonify(cached)

    try:
        # 1. Fetch yearly data
        yearly, data_source = _fetch_invt_data(ticker)
        if not yearly or len(yearly) < 2:
            return jsonify({"error": "Insufficient financial data", "ticker": ticker}), 404

        # 2. Compute metrics for 10yr and 5yr modes
        metrics_10yr = _compute_invt_metrics(yearly, mode="10yr")
        metrics_5yr = _compute_invt_metrics(yearly, mode="5yr")

        # Fill dividend yield from yfinance (historical yield unavailable from EDGAR/FMP)
        yf_trailing_pe = None
        try:
            yf_info = yf.Ticker(ticker).info or {}
            div_yield = yf_info.get("dividendYield") or 0  # yfinance 1.2+: already % (0.9 = 0.9%)
            avg_div_yield = yf_info.get("fiveYearAvgDividendYield") or div_yield
            metrics_10yr["div_yield"] = round(avg_div_yield, 2)
            metrics_5yr["div_yield"] = round(div_yield, 2)
            yf_trailing_pe = yf_info.get("trailingPE")
        except Exception:
            metrics_10yr["div_yield"] = 0
            metrics_5yr["div_yield"] = 0

        # 3. Detect non-dividend payers (for informational note only)
        is_dividend_payer = any(d.get("dividendsPaid", 0) > 0 for d in yearly)

        # 4. Score each metric
        scores_10yr = {k: _invt_score_metric(v, k) for k, v in metrics_10yr.items() if not k.startswith("_")}
        scores_5yr = {k: _invt_score_metric(v, k) for k, v in metrics_5yr.items() if not k.startswith("_")}

        # 5. Category scores — scored categories for overall, info categories separate
        cats_10yr_scored = _compute_invt_category_scores(scores_10yr, INVT_CATEGORIES_SCORED)
        cats_5yr_scored = _compute_invt_category_scores(scores_5yr, INVT_CATEGORIES_SCORED)
        cats_10yr_info = _compute_invt_category_scores(scores_10yr, INVT_CATEGORIES_INFO)
        cats_5yr_info = _compute_invt_category_scores(scores_5yr, INVT_CATEGORIES_INFO)
        cats_10yr = {**cats_10yr_scored, **cats_10yr_info}
        cats_5yr = {**cats_5yr_scored, **cats_5yr_info}

        # 6. Hybrid category scores (all categories) — 70% 10yr + 30% 5yr
        hybrid_cats = {}
        for cat_key in INVT_CATEGORIES:
            s10 = cats_10yr.get(cat_key)
            s5 = cats_5yr.get(cat_key)
            if s10 is not None and s5 is not None:
                hybrid_cats[cat_key] = round(0.7 * s10 + 0.3 * s5, 1)
            else:
                hybrid_cats[cat_key] = s10 if s10 is not None else s5

        # 7. Overall scores — ONLY scored categories (Growth, Profitability, Debt, Efficiency)
        #    Require ≥3 of 4 scored categories to compute overall (refuse truncated scores)
        scored_10yr = [v for v in cats_10yr_scored.values() if v is not None]
        scored_5yr = [v for v in cats_5yr_scored.values() if v is not None]
        overall_10yr = _invt_safe_avg(scored_10yr) if len(scored_10yr) >= 3 else None
        overall_5yr = _invt_safe_avg(scored_5yr) if len(scored_5yr) >= 3 else None
        if overall_10yr is not None and overall_5yr is not None:
            overall = round(0.7 * overall_10yr + 0.3 * overall_5yr, 1)
        else:
            overall = overall_10yr if overall_10yr is not None else overall_5yr

        # 8. Build response — all 5 categories for display
        categories = {}
        for cat_key, cat_def in INVT_CATEGORIES.items():
            is_scored = cat_key in INVT_CATEGORIES_SCORED
            cat_metrics = []
            for m_key in cat_def["metrics"]:
                cat_metrics.append({
                    "name": INVT_METRIC_NAMES.get(m_key, m_key),
                    "key": m_key,
                    "value10yr": round(metrics_10yr.get(m_key, 0) or 0, 2) if metrics_10yr.get(m_key) is not None else None,
                    "value5yr": round(metrics_5yr.get(m_key, 0) or 0, 2) if metrics_5yr.get(m_key) is not None else None,
                    "score10yr": scores_10yr.get(m_key),
                    "score5yr": scores_5yr.get(m_key),
                    "unit": INVT_METRIC_UNITS.get(m_key, ""),
                })
            cat_entry = {
                "label": cat_def["label"],
                "score": hybrid_cats.get(cat_key),
                "score10yr": cats_10yr.get(cat_key),
                "score5yr": cats_5yr.get(cat_key),
                "metrics": cat_metrics,
                "scored": is_scored,
            }
            if cat_key == "shareholder_returns" and not is_dividend_payer:
                cat_entry["note"] = "Non-dividend payer"
            categories[cat_key] = cat_entry

        # Yearly data for per-metric charts (compact: only fields needed for charting)
        est_pe = yf_trailing_pe or 20  # Fallback P/E for estimating historical yield
        yearly_data = []
        prev_dps = None
        for d in yearly:
            s = d.get("sharesOutstanding", 0) or 1
            nd = d.get("totalDebt", 0) - d.get("cash", 0)
            r = d.get("revenue", 0) or 1
            tax_rate = d.get("taxProvision", 0) / d["pretaxIncome"] if d.get("pretaxIncome") and d["pretaxIncome"] > 0 else 0.21
            invested = d.get("totalDebt", 0) + d.get("equity", 0) - d.get("cash", 0)
            nopat = d.get("ebit", 0) * (1 - tax_rate)
            dps = round(d.get("dividendsPaid", 0) / s, 2) if s else 0
            div_growth = round((dps - prev_dps) / prev_dps * 100, 2) if prev_dps and prev_dps > 0 and dps else None
            prev_dps = dps
            # Estimate historical dividend yield: DPS / (EPS * P/E) * 100
            eps_val = d.get("eps", 0)
            est_price = abs(eps_val) * est_pe if eps_val else 0
            div_yield_est = round(dps / est_price * 100, 2) if dps and est_price > 0 else None
            shares_raw = d.get("sharesOutstanding", 0)
            yearly_data.append({
                "year": d["year"],
                "revenue": d.get("revenue", 0),
                "eps": eps_val,
                "fcfPerShare": round(d["fcf"] / s, 2) if d.get("fcf") and s else None,
                "gpm": round(d.get("grossProfit", 0) / r * 100, 2) if d.get("grossProfit") else None,
                "npm": round(d.get("netIncome", 0) / r * 100, 2) if d.get("netIncome") else None,
                "fcfMargin": round(d.get("fcf", 0) / r * 100, 2) if d.get("fcf") else None,
                "netDebt": nd,
                "netDebtFcf": round(nd / d["fcf"], 2) if d.get("fcf") and d["fcf"] > 0 else None,
                "interestCov": round(d["ebit"] / d["interestExpense"], 2) if d.get("ebit") and d.get("interestExpense") and d["interestExpense"] > 0 else None,
                "divYield": div_yield_est,
                "dps": dps,
                "divGrowth": div_growth,
                "payoutRatio": round(d["dividendsPaid"] / d["netIncome"] * 100, 2) if d.get("dividendsPaid") and d.get("netIncome") and d["netIncome"] > 0 else None,
                "fcfPayout": round(d["dividendsPaid"] / d["fcf"] * 100, 2) if d.get("dividendsPaid") and d.get("fcf") and d["fcf"] > 0 else None,
                "sharesOut": shares_raw if shares_raw else None,
                "roa": round(d["netIncome"] / d["totalAssets"] * 100, 2) if d.get("netIncome") and d.get("totalAssets") and d["totalAssets"] > 0 else None,
                "roe": round(d["netIncome"] / d["equity"] * 100, 2) if d.get("netIncome") and d.get("equity") and d["equity"] > 0 else None,
                "roic": round(nopat / invested * 100, 2) if d.get("ebit") and invested > 0 else None,
            })

        result = {
            "ticker": ticker,
            "score": overall,
            "label": _invt_label(overall),
            "score10yr": overall_10yr,
            "score5yr": overall_5yr,
            "shareholderReturnsScore": hybrid_cats.get("shareholder_returns"),
            "categories": categories,
            "years": [d["year"] for d in yearly],
            "yearlyData": yearly_data,
            "dataSource": data_source,
            "lastUpdated": datetime.now().isoformat(),
            "version": 3,
        }

        # 8. Cache
        store = _load_analyzer_store()
        if ticker not in store:
            store[ticker] = {}
        store[ticker]["invtScore"] = result
        _save_analyzer_store(store)

        return jsonify(result)

    except Exception as e:
        print(f"[InvTScore] Error for {ticker}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e), "ticker": ticker}), 500


# ── Salary & Retirement ────────────────────────────────────────────────

# Federal progressive tax brackets (2023 Single Filer — from Excel)
FEDERAL_BRACKETS = [
    (12400, 0.10), (50400, 0.12), (105700, 0.22),
    (201775, 0.24), (256225, 0.32), (640600, 0.35),
    (float('inf'), 0.37),
]

def compute_federal_tax(taxable_income):
    """Progressive federal tax using standard brackets."""
    if taxable_income <= 0:
        return 0
    tax = 0
    prev = 0
    for limit, rate in FEDERAL_BRACKETS:
        amt = min(taxable_income, limit) - prev
        if amt <= 0:
            break
        tax += amt * rate
        prev = limit
    return round(tax, 2)


def _default_taxes():
    return {
        "iraContributionPct": 0.03,
        "standardDeduction": 16100,
        "cityResidentTax": {"name": "City Tax (Resident)", "rate": 0.01, "enabled": True},
        "cityNonResidentTax": {"name": "City Tax (Non-Resident)", "rate": 0.003, "enabled": True},
        "stateTax": {"name": "State Tax", "rate": 0.0425, "enabled": True},
    }


def migrate_salary_data(salary):
    """Convert old flat salary format to new profiles structure. Idempotent."""
    if "profiles" in salary:
        return salary  # already migrated

    old = salary.copy()
    streams = []
    if old.get("w2Salary", 0) > 0:
        streams.append({"type": "W2", "amount": old["w2Salary"], "label": "Main Job"})
    if old.get("income1099", 0) > 0:
        streams.append({"type": "1099", "amount": old["income1099"], "label": "Freelance"})
    if not streams:
        streams.append({"type": "W2", "amount": 0, "label": "Main Job"})

    taxes = _default_taxes()
    if "iraContributionPct" in old:
        taxes["iraContributionPct"] = old["iraContributionPct"]
    if "lansingTaxPct" in old:
        taxes["cityResidentTax"]["rate"] = old["lansingTaxPct"]
    if "eLansingTaxPct" in old:
        taxes["cityNonResidentTax"]["rate"] = old["eLansingTaxPct"]
    if "michiganTaxPct" in old:
        taxes["stateTax"]["rate"] = old["michiganTaxPct"]

    profile = {
        "name": "Alejandro",
        "year": old.get("year", datetime.now().year),
        "incomeStreams": streams,
        "taxes": taxes,
        "projectedSalary": old.get("projectedW2", 140000),
        "history": old.get("history", []),
    }

    new_salary = {
        "activeProfile": "alejandro",
        "profiles": {"alejandro": profile},
        "savedMoney": old.get("savedMoney", 0),
        "pctSavingsToInvest": old.get("pctSavingsToInvest", 1.0),
        "pctIncomeCanSave": old.get("pctIncomeCanSave", 0.25),
        "yearsUntilRetirement": old.get("yearsUntilRetirement", 20),
        "desiredRetirementSalary": old.get("desiredRetirementSalary", 0),
        "annualInterestRate": old.get("annualInterestRate", 0),
        "returnRateRetirement": old.get("returnRateRetirement", 0.04),
    }
    return new_salary


def compute_salary_breakdown(profile):
    """Compute full tax breakdown for a salary profile."""
    streams = profile.get("incomeStreams", [])
    taxes = profile.get("taxes", _default_taxes())

    w2 = sum(s["amount"] for s in streams if s.get("type") == "W2")
    t1099 = sum(s["amount"] for s in streams if s.get("type") in ("1099", "Other"))
    total = w2 + t1099
    if total == 0:
        total = 0.01  # avoid division by zero

    ira_pct = taxes.get("iraContributionPct", 0.03)
    std_deduction = taxes.get("standardDeduction", 16100)
    se_factor = 0.9235
    ss_pct = 0.062
    medicare_pct = 0.0145

    # IRA only on W2
    w2_ira = round(w2 * ira_pct, 2)

    # Taxable base for local/state: salary minus IRA for W2, gross for 1099
    w2_local_base = w2 - w2_ira
    t1099_local_base = t1099

    # Build tax rows dynamically from config
    rows = []
    w2_deductions = w2_ira
    t1099_deductions = 0

    # Row: Annual Salary
    rows.append({"label": "Annual Salary", "total": round(total, 2), "totalMo": round(total/12, 2),
                 "w2": round(w2, 2), "w2Mo": round(w2/12, 2), "t1099": round(t1099, 2), "t1099Mo": round(t1099/12, 2), "isIncome": True})

    # Row: Pre-Tax Deductions (IRA)
    rows.append({"label": "Pre-Tax Deductions (IRA)", "total": round(w2_ira, 2), "totalMo": round(w2_ira/12, 2),
                 "w2": round(w2_ira, 2), "w2Mo": round(w2_ira/12, 2), "t1099": 0, "t1099Mo": 0,
                 "ratePct": ira_pct*100, "rateKey": "iraContributionPct"})

    # Configurable local/state taxes (toggleable)
    tax_lines = [
        ("cityResidentTax", w2_local_base, t1099_local_base),
        ("cityNonResidentTax", w2_local_base, t1099_local_base),
        ("stateTax", w2_local_base, t1099_local_base),
    ]
    w2_local_total = 0
    t1099_local_total = 0
    for key, w2_base, t1099_base in tax_lines:
        cfg = taxes.get(key, {})
        if not cfg.get("enabled", False):
            continue
        rate = cfg.get("rate", 0)
        name = cfg.get("name", key)
        w2_amt = round(w2_base * rate, 2)
        t1099_amt = round(t1099_base * rate, 2)
        total_amt = round(w2_amt + t1099_amt, 2)
        w2_local_total += w2_amt
        t1099_local_total += t1099_amt
        w2_deductions += w2_amt
        t1099_deductions += t1099_amt
        rows.append({"label": name, "total": total_amt, "totalMo": round(total_amt/12, 2),
                     "w2": w2_amt, "w2Mo": round(w2_amt/12, 2), "t1099": t1099_amt, "t1099Mo": round(t1099_amt/12, 2),
                     "ratePct": rate*100, "taxKey": key, "toggleable": True})

    # Federal tax — progressive brackets with standard deduction
    w2_fed_taxable = max(0, w2 - w2_ira - std_deduction)
    # 1099: deduct half of SE tax from federal taxable income
    t1099_se_tax = t1099 * se_factor * (ss_pct + medicare_pct) * 2
    t1099_fed_taxable = max(0, t1099 - round(t1099_se_tax / 2, 2))
    w2_federal = compute_federal_tax(w2_fed_taxable)
    t1099_federal = compute_federal_tax(t1099_fed_taxable)
    total_federal = round(w2_federal + t1099_federal, 2)
    # Compute effective federal rate for display
    fed_base = w2_fed_taxable + t1099_fed_taxable
    eff_fed_pct = round((total_federal / fed_base) * 100, 2) if fed_base > 0 else 0
    w2_deductions += w2_federal
    t1099_deductions += t1099_federal
    rows.append({"label": "Federal Tax", "total": total_federal, "totalMo": round(total_federal/12, 2),
                 "w2": w2_federal, "w2Mo": round(w2_federal/12, 2), "t1099": t1099_federal, "t1099Mo": round(t1099_federal/12, 2),
                 "effRate": eff_fed_pct, "isFederal": True})

    # Social Security
    w2_ss = round(w2 * ss_pct, 2)
    t1099_ss = round(t1099 * se_factor * ss_pct * 2, 2)
    total_ss = round(w2_ss + t1099_ss, 2)
    w2_deductions += w2_ss
    t1099_deductions += t1099_ss
    rows.append({"label": "Social Security", "total": total_ss, "totalMo": round(total_ss/12, 2),
                 "w2": w2_ss, "w2Mo": round(w2_ss/12, 2), "t1099": t1099_ss, "t1099Mo": round(t1099_ss/12, 2),
                 "fixedRate": round(ss_pct*100, 2)})

    # Medicare
    w2_med = round(w2 * medicare_pct, 2)
    t1099_med = round(t1099 * se_factor * medicare_pct * 2, 2)
    total_med = round(w2_med + t1099_med, 2)
    w2_deductions += w2_med
    t1099_deductions += t1099_med
    rows.append({"label": "Medicare", "total": total_med, "totalMo": round(total_med/12, 2),
                 "w2": w2_med, "w2Mo": round(w2_med/12, 2), "t1099": t1099_med, "t1099Mo": round(t1099_med/12, 2),
                 "fixedRate": round(medicare_pct*100, 2)})

    # Totals
    total_withheld = round(w2_deductions + t1099_deductions, 2)
    w2_takehome = round(w2 - w2_deductions, 2)
    t1099_takehome = round(t1099 - t1099_deductions, 2)
    total_takehome = round(total - total_withheld, 2)

    rows.append({"label": "Total Withheld", "total": total_withheld, "totalMo": round(total_withheld/12, 2),
                 "w2": round(w2_deductions, 2), "w2Mo": round(w2_deductions/12, 2),
                 "t1099": round(t1099_deductions, 2), "t1099Mo": round(t1099_deductions/12, 2), "isSummary": True})
    rows.append({"label": "Take-Home Pay", "total": total_takehome, "totalMo": round(total_takehome/12, 2),
                 "w2": w2_takehome, "w2Mo": round(w2_takehome/12, 2),
                 "t1099": t1099_takehome, "t1099Mo": round(t1099_takehome/12, 2), "isSummary": True, "isPositive": True})

    # Hourly / Eff Tax
    real_total = w2 + t1099  # use actual total, not the 0.01 guard
    total_hourly = round(real_total / (52 * 40), 2) if real_total > 0 else 0
    w2_hourly = round(w2 / (52 * 40), 2) if w2 > 0 else 0
    t1099_hourly = round(t1099 / (52 * 40), 2) if t1099 > 0 else 0
    total_eff = round(total_withheld / real_total, 4) if real_total > 0 else 0
    w2_eff = round(w2_deductions / w2, 4) if w2 > 0 else 0
    t1099_eff = round(t1099_deductions / t1099, 4) if t1099 > 0 else 0
    rows.append({"label": "Hourly Rate / Eff. Tax %", "total": total_hourly, "totalMo": total_eff,
                 "w2": w2_hourly, "w2Mo": w2_eff, "t1099": t1099_hourly, "t1099Mo": t1099_eff, "isRate": True})

    # Taxable income row (insert after Annual Salary)
    w2_taxable = round(w2 - w2_ira, 2)
    t1099_taxable = round(t1099, 2)
    total_taxable = round(w2_taxable + t1099_taxable, 2)
    rows.insert(1, {"label": "Taxable Income", "total": total_taxable, "totalMo": round(total_taxable/12, 2),
                     "w2": w2_taxable, "w2Mo": round(w2_taxable/12, 2), "t1099": t1099_taxable, "t1099Mo": round(t1099_taxable/12, 2)})

    # Employer cost (W2 only)
    emp_ira = round(w2 * ira_pct, 2)
    emp_futa = round(0.006 * 7000 + 0.027 * 9500, 2) if w2 > 0 else 0
    emp_ss = round(w2 * ss_pct, 2)
    emp_med = round(w2 * medicare_pct, 2)
    emp_total = round(emp_ira + emp_futa + emp_ss + emp_med, 2)
    employer = {
        "rows": [
            {"label": "IRA Match", "annual": emp_ira, "monthly": round(emp_ira/12, 2)},
            {"label": "Federal Unemployment (FUTA)", "annual": emp_futa, "monthly": round(emp_futa/12, 2)},
            {"label": "Social Security (6.2%)", "annual": emp_ss, "monthly": round(emp_ss/12, 2)},
            {"label": "Medicare (1.45%)", "annual": emp_med, "monthly": round(emp_med/12, 2)},
        ],
        "total": emp_total,
        "totalMonthly": round(emp_total/12, 2),
        "costToCompany": round(w2 + emp_total, 2),
        "costToCompanyMonthly": round((w2 + emp_total)/12, 2),
    }

    # Projected salary (W2 only, same tax config)
    proj_amount = profile.get("projectedSalary", 0)
    projected = None
    if proj_amount > 0:
        proj_profile = {
            "incomeStreams": [{"type": "W2", "amount": proj_amount, "label": "Projected"}],
            "taxes": taxes,
        }
        proj_bd = compute_salary_breakdown(proj_profile)
        projected = {
            "amount": proj_amount,
            "rows": proj_bd["rows"],
            "summary": proj_bd["summary"],
            "vsCurrent": {
                "deltaGross": round(proj_amount - real_total, 2),
                "deltaTakeHome": round(proj_bd["summary"]["takeHomePay"] - total_takehome, 2),
                "deltaEffRate": round((proj_bd["summary"]["effectiveTaxRate"] - total_eff) * 100, 2),
            }
        }

    return {
        "rows": rows,
        "summary": {
            "annualGross": round(real_total, 2), "w2Total": round(w2, 2), "t1099Total": round(t1099, 2),
            "takeHomePay": total_takehome, "totalWithhold": total_withheld,
            "effectiveTaxRate": total_eff, "hourlyRate": total_hourly,
            "monthlySalary": round(total_takehome/12, 2),
        },
        "employer": employer,
        "projected": projected,
    }


_TAX_NAME_MAP = {
    "Lansing Resident Tax": "City Tax (Resident)",
    "E Lansing Nonresident Tax": "City Tax (Non-Resident)",
    "Michigan State Tax": "State Tax",
}

def _get_salary_data(portfolio):
    """Get salary data, migrating if needed."""
    salary = portfolio.get("salary", {})
    if "profiles" not in salary:
        salary = migrate_salary_data(salary)
        portfolio["salary"] = salary
        save_portfolio(portfolio)
    # Fix legacy tax names + backfill missing effectiveTaxRate in history
    changed = False
    for pid, profile in salary.get("profiles", {}).items():
        taxes = profile.get("taxes", {})
        for tkey in ("cityResidentTax", "cityNonResidentTax", "stateTax"):
            cfg = taxes.get(tkey, {})
            if cfg.get("name") in _TAX_NAME_MAP:
                cfg["name"] = _TAX_NAME_MAP[cfg["name"]]
                changed = True
        for h in profile.get("history", []):
            if "effectiveTaxRate" not in h and h.get("annualPayroll", 0) > 0:
                h["effectiveTaxRate"] = round(1 - h["takeHomePay"] / h["annualPayroll"], 4)
                changed = True
    if changed:
        portfolio["salary"] = salary
        save_portfolio(portfolio)
    return salary


@app.route("/api/salary")
def api_salary():
    portfolio = load_portfolio()
    salary = _get_salary_data(portfolio)
    profile_id = request.args.get("profile", salary.get("activeProfile", "alejandro"))
    profile = salary.get("profiles", {}).get(profile_id, {})
    breakdown = compute_salary_breakdown(profile)
    # Household summary
    household = {"annualGross": 0, "takeHomePay": 0, "profileCount": 0}
    for pid, p in salary.get("profiles", {}).items():
        bd = compute_salary_breakdown(p)
        household["annualGross"] += bd["summary"]["annualGross"]
        household["takeHomePay"] += bd["summary"]["takeHomePay"]
        household["profileCount"] += 1
    household["annualGross"] = round(household["annualGross"], 2)
    household["takeHomePay"] = round(household["takeHomePay"], 2)
    return jsonify({
        "salary": salary,
        "profile": profile,
        "profileId": profile_id,
        "breakdown": breakdown,
        "household": household,
        "costOfLiving": portfolio.get("costOfLiving", []),
        "lastUpdated": datetime.now().isoformat(),
    })


@app.route("/api/salary/update", methods=["POST"])
def api_salary_update():
    b = request.get_json()
    portfolio = load_portfolio()
    salary = _get_salary_data(portfolio)
    profile_id = b.get("profileId", salary.get("activeProfile", "alejandro"))
    profile = salary.get("profiles", {}).get(profile_id, {})

    # Update income streams
    if "incomeStreams" in b:
        profile["incomeStreams"] = b["incomeStreams"]
    # Update tax config
    if "taxes" in b:
        profile["taxes"] = b["taxes"]
    # Update simple fields
    for key in ("year", "projectedSalary", "name"):
        if key in b:
            profile[key] = int(b[key]) if key == "year" else b[key]
    # Update shared fields
    for key in ("savedMoney", "pctSavingsToInvest", "pctIncomeCanSave"):
        if key in b:
            salary[key] = float(b[key])

    salary["profiles"][profile_id] = profile
    portfolio["salary"] = salary
    save_portfolio(portfolio)
    breakdown = compute_salary_breakdown(profile)
    return jsonify({"ok": True, "profile": profile, "breakdown": breakdown})


@app.route("/api/salary/profile", methods=["POST"])
def api_salary_profile_create():
    b = request.get_json()
    portfolio = load_portfolio()
    salary = _get_salary_data(portfolio)
    name = b.get("name", "New Profile")
    pid = name.lower().replace(" ", "_")
    # Avoid collisions
    base_pid = pid
    n = 1
    while pid in salary.get("profiles", {}):
        pid = f"{base_pid}_{n}"
        n += 1
    salary.setdefault("profiles", {})[pid] = {
        "name": name,
        "year": datetime.now().year,
        "incomeStreams": [{"type": "W2", "amount": 0, "label": "Main Job"}],
        "taxes": _default_taxes(),
        "projectedSalary": 0,
        "history": [],
    }
    salary["activeProfile"] = pid
    portfolio["salary"] = salary
    save_portfolio(portfolio)
    return jsonify({"ok": True, "profileId": pid, "salary": salary})


@app.route("/api/salary/profile/<pid>", methods=["DELETE"])
def api_salary_profile_delete(pid):
    portfolio = load_portfolio()
    salary = _get_salary_data(portfolio)
    profiles = salary.get("profiles", {})
    if pid in profiles and len(profiles) > 1:
        del profiles[pid]
        if salary.get("activeProfile") == pid:
            salary["activeProfile"] = next(iter(profiles))
        portfolio["salary"] = salary
        save_portfolio(portfolio)
        return jsonify({"ok": True, "salary": salary})
    return jsonify({"ok": False, "error": "Cannot delete last profile"}), 400


@app.route("/api/salary/history/save", methods=["POST"])
def api_salary_history_save():
    b = request.get_json()
    portfolio = load_portfolio()
    salary = _get_salary_data(portfolio)
    profile_id = b.get("profileId", salary.get("activeProfile", "alejandro"))
    profile = salary.get("profiles", {}).get(profile_id, {})
    bd = compute_salary_breakdown(profile)
    summ = bd["summary"]
    year = profile.get("year", datetime.now().year)
    history = profile.get("history", [])
    # Replace if year exists, else append
    history = [h for h in history if h.get("year") != year]
    history.append({
        "year": year,
        "annualPayroll": summ["annualGross"],
        "monthlyPayroll": round(summ["annualGross"] / 12, 2),
        "takeHomePay": summ["takeHomePay"],
        "effectiveTaxRate": summ["effectiveTaxRate"],
    })
    history.sort(key=lambda h: h["year"])
    profile["history"] = history
    salary["profiles"][profile_id] = profile
    portfolio["salary"] = salary
    save_portfolio(portfolio)
    return jsonify({"ok": True, "history": history})


@app.route("/api/salary/history/<int:year>", methods=["DELETE"])
def api_salary_history_delete(year):
    b = request.get_json() or {}
    portfolio = load_portfolio()
    salary = _get_salary_data(portfolio)
    profile_id = b.get("profileId", salary.get("activeProfile", "alejandro"))
    profile = salary.get("profiles", {}).get(profile_id, {})
    history = profile.get("history", [])
    profile["history"] = [h for h in history if h.get("year") != year]
    salary["profiles"][profile_id] = profile
    portfolio["salary"] = salary
    save_portfolio(portfolio)
    return jsonify({"ok": True, "history": profile["history"]})

@app.route("/api/cost-of-living")
def api_cost_of_living():
    return crud_list("costOfLiving")

@app.route("/api/cost-of-living/add", methods=["POST"])
def api_cost_of_living_add():
    b = request.get_json()
    item = {
        "city": b.get("city", ""),
        "state": b.get("state", ""),
        "rent": float(b.get("rent", 0)),
        "food": float(b.get("food", 0)),
        "transport": float(b.get("transport", 0)),
        "utilities": float(b.get("utilities", 0)),
        "insurance": float(b.get("insurance", 0)),
        "other": float(b.get("other", 0)),
        "notes": b.get("notes", ""),
    }
    item["total"] = round(sum(item[k] for k in ["rent", "food", "transport", "utilities", "insurance", "other"]), 2)
    return crud_add("costOfLiving", item)

@app.route("/api/cost-of-living/update", methods=["POST"])
def api_cost_of_living_update():
    b = request.get_json()
    return crud_update("costOfLiving", int(b.get("index", -1)), b.get("updates", {}))

@app.route("/api/cost-of-living/delete", methods=["POST"])
def api_cost_of_living_delete():
    b = request.get_json()
    return crud_delete("costOfLiving", int(b.get("index", -1)))


# ── Passive Income Tracking ────────────────────────────────────────────
@app.route("/api/passive-income")
def api_passive_income():
    return crud_list("passiveIncome")

@app.route("/api/passive-income/add", methods=["POST"])
def api_passive_income_add():
    b = request.get_json()
    item = {
        "source": b.get("source", ""),
        "type": b.get("type", "Dividend"),
        "amount": float(b.get("amount", 0)),
        "frequency": b.get("frequency", "Monthly"),
        "startDate": b.get("startDate", ""),
        "active": b.get("active", True),
        "notes": b.get("notes", ""),
    }
    freq_mult = {"Monthly": 12, "Quarterly": 4, "Semi-Annual": 2, "Annual": 1, "Weekly": 52}
    item["annualized"] = round(item["amount"] * freq_mult.get(item["frequency"], 12), 2)
    return crud_add("passiveIncome", item)

@app.route("/api/passive-income/update", methods=["POST"])
def api_passive_income_update():
    b = request.get_json()
    return crud_update("passiveIncome", int(b.get("index", -1)), b.get("updates", {}))

@app.route("/api/passive-income/delete", methods=["POST"])
def api_passive_income_delete():
    b = request.get_json()
    return crud_delete("passiveIncome", int(b.get("index", -1)))


# ── Rule 4% ────────────────────────────────────────────────────────────
@app.route("/api/rule4pct")
def api_rule4pct():
    portfolio = load_portfolio()
    return jsonify({
        "rule4Pct": portfolio.get("rule4Pct", {}),
        "lastUpdated": datetime.now().isoformat(),
    })

@app.route("/api/rule4pct/update", methods=["POST"])
def api_rule4pct_update():
    b = request.get_json()
    portfolio = load_portfolio()
    r4 = portfolio.get("rule4Pct", {})
    for key in ["annualExpenses", "inflationPct", "withdrawalPct", "currentPortfolio", "monthlyContribution", "expectedReturnPct"]:
        if key in b:
            r4[key] = float(b[key])
    portfolio["rule4Pct"] = r4
    save_portfolio(portfolio)
    return jsonify({"ok": True, "rule4Pct": r4})


# ── Historic S&P 500 Data ──────────────────────────────────────────────
def load_historic_data():
    """Load S&P 500 historic data from portfolio or import from Excel."""
    portfolio = load_portfolio()
    historic = portfolio.get("historicData")
    if historic and len(historic) > 50:
        return historic

    # Try importing from Excel
    xlsx_path = os.path.join(DATA_DIR, "Investments Toolkit-v1.0.xlsx")
    if not os.path.exists(xlsx_path):
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["Historic_Data"]
        data = []
        for r in range(2, ws.max_row + 1):
            year = ws.cell(r, 1).value
            if year is None or not isinstance(year, (int, float)) or year < 1900:
                continue
            data.append({
                "year": int(year),
                "avgClosing": ws.cell(r, 2).value or 0,
                "yearOpen": ws.cell(r, 3).value or 0,
                "yearHigh": ws.cell(r, 4).value or 0,
                "yearLow": ws.cell(r, 5).value or 0,
                "yearClose": ws.cell(r, 6).value or 0,
                "annualReturn": ws.cell(r, 7).value or 0,
                "cpi": ws.cell(r, 8).value or 0,
            })
        wb.close()
        # Cache in portfolio
        portfolio["historicData"] = data
        save_portfolio(portfolio)
        return data
    except Exception as e:
        print(f"Error importing historic data: {e}")
        return []


@app.route("/api/historic-data")
def api_historic_data():
    data = load_historic_data()
    return jsonify({
        "historicData": data,
        "lastUpdated": datetime.now().isoformat(),
    })


def _run_simulation(returns_by_year, cpi_by_year, all_years, max_year,
                     starting_balance, withdrawal_rate, horizon,
                     strategy="fixed", guardrail_floor=None, guardrail_ceiling=None,
                     cash_buffer_years=0, div_yield=0, div_growth=0):
    """Core simulation engine supporting multiple strategies."""
    scenarios = []
    success_count = 0
    total_count = 0

    for start_year in all_years:
        end_year = start_year + horizon - 1
        if end_year > max_year:
            break

        total_count += 1
        balance = starting_balance
        base_withdrawal = starting_balance * withdrawal_rate
        annual_withdrawal = base_withdrawal
        cash_reserve = base_withdrawal * cash_buffer_years
        yearly_data = []
        survived = True
        cumulative_inflation = 1.0

        for yr_offset in range(horizon):
            yr = start_year + yr_offset
            ret = returns_by_year.get(yr, 0)
            cpi = cpi_by_year.get(yr, 0.03)
            cumulative_inflation *= (1 + cpi)

            if strategy == "dividend":
                # Dividend strategy: yield on current balance, no selling
                div_income = balance * div_yield
                balance = balance * (1 + ret)
                actual_withdrawal = div_income
            elif strategy == "combined":
                # Combined: dividend income + sell remainder
                div_income = balance * div_yield
                balance = balance * (1 + ret)
                sell_amount = max(0, annual_withdrawal - div_income)
                balance -= sell_amount
                actual_withdrawal = div_income + sell_amount
            elif strategy == "guardrails":
                # Guardrails: adjust withdrawal based on portfolio performance
                balance = balance * (1 + ret)
                floor_amount = base_withdrawal * cumulative_inflation * (guardrail_floor or 0.8)
                ceiling_amount = base_withdrawal * cumulative_inflation * (guardrail_ceiling or 1.2)
                # Target: withdrawal_rate of current balance
                target = balance * withdrawal_rate
                annual_withdrawal = max(floor_amount, min(ceiling_amount, target))
                # Use cash buffer during down years
                if ret < -0.1 and cash_reserve > 0:
                    from_cash = min(cash_reserve, annual_withdrawal)
                    cash_reserve -= from_cash
                    balance -= (annual_withdrawal - from_cash)
                else:
                    balance -= annual_withdrawal
                actual_withdrawal = annual_withdrawal
            else:
                # Fixed (classic Rule 4%)
                balance = balance * (1 + ret)
                # Use cash buffer during down years
                if cash_buffer_years > 0 and ret < -0.1 and cash_reserve > 0:
                    from_cash = min(cash_reserve, annual_withdrawal)
                    cash_reserve -= from_cash
                    balance -= (annual_withdrawal - from_cash)
                else:
                    balance -= annual_withdrawal
                actual_withdrawal = annual_withdrawal

            yearly_data.append({
                "year": yr,
                "retirementYear": yr_offset + 1,
                "balance": round(balance, 2),
                "returnPct": ret,
                "withdrawalAmount": round(actual_withdrawal, 2),
                "inflationPct": cpi,
                "cumulativeInflation": round(cumulative_inflation, 4),
                "cashReserve": round(cash_reserve, 2) if cash_buffer_years > 0 else None,
            })

            if balance <= 0:
                survived = False
                for remaining in range(yr_offset + 1, horizon):
                    yearly_data.append({
                        "year": start_year + remaining,
                        "retirementYear": remaining + 1,
                        "balance": 0,
                        "returnPct": returns_by_year.get(start_year + remaining, 0),
                        "withdrawalAmount": 0,
                        "inflationPct": cpi_by_year.get(start_year + remaining, 0),
                        "cumulativeInflation": round(cumulative_inflation, 4),
                        "cashReserve": 0,
                    })
                break

            # Adjust withdrawal for inflation (fixed & combined strategies)
            if strategy in ("fixed", "combined"):
                annual_withdrawal *= (1 + cpi)
            elif strategy == "dividend":
                # Dividend growth replaces inflation adjustment
                div_yield_adj = div_yield  # yield stays same, applied to growing balance

        if survived:
            success_count += 1

        scenarios.append({
            "startYear": start_year,
            "endYear": end_year,
            "survived": survived,
            "finalBalance": round(yearly_data[-1]["balance"], 2),
            "data": yearly_data,
        })

    success_rate = round(success_count / total_count * 100, 1) if total_count > 0 else 0
    avg_final = round(sum(s["finalBalance"] for s in scenarios if s["survived"]) / max(success_count, 1), 2)
    worst = min(scenarios, key=lambda s: s["finalBalance"]) if scenarios else None
    best = max(scenarios, key=lambda s: s["finalBalance"]) if scenarios else None

    return {
        "horizon": horizon,
        "totalScenarios": total_count,
        "successCount": success_count,
        "failureCount": total_count - success_count,
        "successRate": success_rate,
        "avgFinalBalance": avg_final,
        "worstStartYear": worst["startYear"] if worst else None,
        "worstFinalBalance": worst["finalBalance"] if worst else None,
        "bestStartYear": best["startYear"] if best else None,
        "bestFinalBalance": best["finalBalance"] if best else None,
        "scenarios": scenarios,
    }


@app.route("/api/rule4pct/simulate")
def api_rule4pct_simulate():
    """Run Rule 4% historical simulation across all possible starting years."""
    starting_balance = float(request.args.get("balance", 1000000))
    withdrawal_rate = float(request.args.get("rate", 4)) / 100
    strategy = request.args.get("strategy", "fixed")  # fixed, guardrails, dividend, combined
    cash_buffer = int(request.args.get("cashBuffer", 0))
    div_yield = float(request.args.get("divYield", 4)) / 100
    div_growth = float(request.args.get("divGrowth", 5.6)) / 100
    guardrail_floor = float(request.args.get("grFloor", 80)) / 100
    guardrail_ceiling = float(request.args.get("grCeiling", 120)) / 100

    historic = load_historic_data()
    if not historic:
        return jsonify({"error": "No historic data available"}), 404

    returns_by_year = {h["year"]: h["annualReturn"] for h in historic}
    cpi_by_year = {h["year"]: h["cpi"] for h in historic}
    all_years = sorted(returns_by_year.keys())
    max_year = all_years[-1]

    results = {}
    for horizon in [20, 30, 40]:
        results[str(horizon)] = _run_simulation(
            returns_by_year, cpi_by_year, all_years, max_year,
            starting_balance, withdrawal_rate, horizon,
            strategy=strategy,
            guardrail_floor=guardrail_floor,
            guardrail_ceiling=guardrail_ceiling,
            cash_buffer_years=cash_buffer,
            div_yield=div_yield,
            div_growth=div_growth,
        )

    return jsonify({
        "startingBalance": starting_balance,
        "withdrawalRate": withdrawal_rate * 100,
        "strategy": strategy,
        "results": results,
        "lastUpdated": datetime.now().isoformat(),
    })


@app.route("/api/rule4pct/compare")
def api_rule4pct_compare():
    """Compare multiple strategies side by side for a specific horizon and starting year."""
    starting_balance = float(request.args.get("balance", 1000000))
    rate = float(request.args.get("rate", 4)) / 100
    horizon = int(request.args.get("horizon", 20))
    div_yield = float(request.args.get("divYield", 4)) / 100
    cash_buffer = int(request.args.get("cashBuffer", 0))

    historic = load_historic_data()
    if not historic:
        return jsonify({"error": "No historic data available"}), 404

    returns_by_year = {h["year"]: h["annualReturn"] for h in historic}
    cpi_by_year = {h["year"]: h["cpi"] for h in historic}
    all_years = sorted(returns_by_year.keys())
    max_year = all_years[-1]

    comparison = {}
    for strat in ["fixed", "guardrails", "dividend", "combined"]:
        comparison[strat] = _run_simulation(
            returns_by_year, cpi_by_year, all_years, max_year,
            starting_balance, rate, horizon,
            strategy=strat,
            cash_buffer_years=cash_buffer,
            div_yield=div_yield,
        )

    return jsonify({
        "startingBalance": starting_balance,
        "withdrawalRate": rate * 100,
        "horizon": horizon,
        "divYield": div_yield * 100,
        "comparison": comparison,
        "lastUpdated": datetime.now().isoformat(),
    })


@app.route("/api/status")
def api_status():
    """Health check."""
    return jsonify({
        "status": "ok",
        "dataSource": "yfinance",
        "dataDir": str(DATA_DIR),
        "cacheEntries": len(_cache),
        "portfolioFile": PORTFOLIO_FILE.exists(),
        "timestamp": datetime.now().isoformat(),
    })


# ── Main ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    load_disk_cache()
    _load_13f_cache()
    print("\n" + "=" * 55)
    print("  InvToolkit — Investment Dashboard")
    print("=" * 55)
    print("  Data source: Yahoo Finance (yfinance)")
    print(f"  Data dir:    {DATA_DIR}")
    print(f"  Portfolio:   {PORTFOLIO_FILE}")
    print(f"  Dashboard:   http://localhost:5050")
    print("=" * 55 + "\n")
    app.run(host="0.0.0.0", port=5050, debug=True)
