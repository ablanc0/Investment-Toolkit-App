"""Export routes — CSV and Excel download endpoints."""

import csv
import io
from datetime import datetime

from flask import Blueprint, jsonify, make_response

from services.data_store import load_portfolio
from services.yfinance_svc import fetch_all_quotes

bp = Blueprint("export", __name__)


def _csv_response(rows, headers, filename):
    """Build a CSV download response."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _xlsx_response(rows, headers, filename, sheet_name="Data"):
    """Build an Excel download response."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1d2e", end_color="1a1d2e", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@bp.route("/api/export/holdings", methods=["GET"])
def export_holdings():
    """Export current holdings as CSV."""
    from flask import request
    fmt = request.args.get("format", "csv")

    portfolio = load_portfolio()
    positions = portfolio.get("positions", [])
    quotes = fetch_all_quotes([p["ticker"] for p in positions])

    headers = ["Ticker", "Shares", "Avg Cost", "Current Price", "Cost Basis",
               "Market Value", "Return $", "Return %", "Day Change %",
               "Div Yield %", "Annual Div Income", "Category", "Sector"]

    rows = []
    for pos in positions:
        ticker = pos["ticker"]
        q = quotes.get(ticker, {})
        shares = float(pos.get("shares", 0))
        avg_cost = float(pos.get("avgCost", 0))
        price = q.get("price", 0)
        cost_basis = round(shares * avg_cost, 2)
        market_value = round(shares * price, 2)
        market_return = round(market_value - cost_basis, 2)
        return_pct = round((market_return / cost_basis * 100) if cost_basis > 0 else 0, 2)
        day_chg = q.get("dayChangePercent", 0)
        div_yield = q.get("dividendYield", 0)
        div_rate = q.get("dividendRate", 0) or 0
        annual_div = round(div_rate * shares, 2)
        category = pos.get("category", "")
        sector = q.get("sector", "")

        rows.append([ticker, shares, avg_cost, price, cost_basis, market_value,
                     market_return, return_pct, round(day_chg, 2),
                     round(div_yield, 2), annual_div, category, sector])

    ts = datetime.now().strftime("%Y%m%d")
    if fmt == "xlsx":
        return _xlsx_response(rows, headers, f"holdings_{ts}.xlsx", "Holdings")
    return _csv_response(rows, headers, f"holdings_{ts}.csv")


@bp.route("/api/export/dividends", methods=["GET"])
def export_dividends():
    """Export dividend log as CSV."""
    from flask import request
    fmt = request.args.get("format", "csv")

    portfolio = load_portfolio()
    div_log = portfolio.get("dividendLog", [])

    headers = ["Month", "Ticker", "Amount"]
    rows = []
    for entry in div_log:
        month = entry.get("month", "")
        for ticker, amount in entry.get("dividends", {}).items():
            rows.append([month, ticker, amount])

    ts = datetime.now().strftime("%Y%m%d")
    if fmt == "xlsx":
        return _xlsx_response(rows, headers, f"dividends_{ts}.xlsx", "Dividends")
    return _csv_response(rows, headers, f"dividends_{ts}.csv")


@bp.route("/api/export/watchlist", methods=["GET"])
def export_watchlist():
    """Export watchlist as CSV."""
    from flask import request
    fmt = request.args.get("format", "csv")

    portfolio = load_portfolio()
    watchlist = portfolio.get("watchlist", [])
    tickers = [w["ticker"] for w in watchlist]
    quotes = fetch_all_quotes(tickers) if tickers else {}

    headers = ["Ticker", "Target Price", "Current Price", "Category", "Notes"]
    rows = []
    for w in watchlist:
        ticker = w["ticker"]
        q = quotes.get(ticker, {})
        rows.append([
            ticker,
            w.get("targetPrice", ""),
            q.get("price", ""),
            w.get("category", ""),
            w.get("notes", ""),
        ])

    ts = datetime.now().strftime("%Y%m%d")
    if fmt == "xlsx":
        return _xlsx_response(rows, headers, f"watchlist_{ts}.xlsx", "Watchlist")
    return _csv_response(rows, headers, f"watchlist_{ts}.csv")


@bp.route("/api/export/monthly", methods=["GET"])
def export_monthly():
    """Export monthly portfolio data as CSV."""
    from flask import request
    fmt = request.args.get("format", "csv")

    portfolio = load_portfolio()
    monthly = portfolio.get("monthlyData", [])

    headers = ["Month", "Portfolio Value", "Contributions", "Accumulated Investment",
               "Dividend Income"]
    rows = []
    for m in monthly:
        div_income = sum(
            sum(float(v) for v in entry.get("dividends", {}).values())
            for entry in portfolio.get("dividendLog", [])
            if entry.get("month", "").startswith(m.get("month", "???").split(" ")[0])
        ) if "month" in m else 0
        rows.append([
            m.get("month", ""),
            m.get("portfolioValue", 0),
            m.get("contributions", 0),
            m.get("accumulatedInvestment", 0),
            round(div_income, 2),
        ])

    ts = datetime.now().strftime("%Y%m%d")
    if fmt == "xlsx":
        return _xlsx_response(rows, headers, f"monthly_{ts}.xlsx", "Monthly Data")
    return _csv_response(rows, headers, f"monthly_{ts}.csv")
